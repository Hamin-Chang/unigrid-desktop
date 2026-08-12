# -*- coding: utf-8 -*-
"""A1 조정 열을 붙여도 **옛 계통이 그대로인가** (§7 5단계 1번 — 열 설계).

가장 무서운 실패가 하나 있다. 엔진은 `size(표, 2) >= N` 으로 기능 유무를 가리므로
**표가 넓어지는 것만으로 없던 기능이 켜진다.** A1 은 `AC Line Data` 를 13 → 19열,
`AC Bus Data` 를 17 → 22열로 늘리는 일이라 정면으로 그 함정 위에 있다.

빠져나가는 길은 `read_v2` 의 폭 줄이기다 — 새 열이 **전부 비어 있으면** 폭을 옛 폭으로
되돌린다. 그래서 이 시험이 보는 것은 딱 둘이다:

    · 값을 안 넣으면 엔진이 받는 폭이 **옛날 그대로**인가 (13 · 17)
    · 값을 넣으면 폭이 **늘고 그 값이 제 자리에 꽂히는가** (19 · 22)

    python tests/test_a1_columns.py
"""

import sys
import tempfile
from pathlib import Path

import numpy as np
from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
REPO = HERE.parent
SRC = REPO / "src"
if SRC.is_dir() and str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))

import convert_case                      # noqa: E402
import format_v2 as F                    # noqa: E402
import read_v2                           # noqa: E402

V1_CASE = HERE / "cases_v1" / "AConly_case118.xlsx"
# (시트 이름, 옛 폭, 새 폭, 첫 조정 열의 v1 자리)
TARGETS = [("AC Line Data", 13, 19, 14), ("AC Bus Data", 17, 22, 18)]

ok = ng = 0


def check_small(label: str, got: float, limit: float = 1e-12) -> None:
    """티끌만큼은 봐준다. 이 저장소의 기존 기준과 같은 값이다(`test_format_v2.py`).

    🚨 0 을 요구하면 안 된다 — 소수를 이진수로 담는 데서 오는 차이가 남는다
       (`102.36149999999999` ↔ `102.3615`). 자리가 밀리는 진짜 실패는 이것보다
       한참 크게 나오므로 이 잣대로도 갈린다.
    """
    global ok, ng
    if got <= limit:
        print(f"  ✅ {label:<52} {got:.3e}")
        ok += 1
    else:
        print(f"  ❌ {label:<52} {got:.3e}  (봐주는 한계 {limit:.0e})")
        ng += 1


def check(label: str, got, want) -> None:
    global ok, ng
    if got == want:
        print(f"  ✅ {label:<52} {got}")
        ok += 1
    else:
        print(f"  ❌ {label:<52} {got}  (바라던 값 {want})")
        ng += 1


def sheet_of(name: str) -> F.Sheet:
    return next(s for s in F.SHEETS if s.name == name)


def engine_table(path: Path, name: str) -> np.ndarray:
    """엔진이 실제로 받는 표 (= `read_v2` 가 옛 자리로 되돌린 것)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        return read_v2.read_sheet(sheet_of(name), wb[name])
    finally:
        wb.close()


def main() -> int:
    if not V1_CASE.exists():
        print(f"시험 케이스가 없습니다: {V1_CASE}")
        return 2

    tmp = Path(tempfile.mkdtemp(prefix="a1_cols_"))
    out, notes = convert_case.convert_file(V1_CASE, tmp)
    if out is None:
        print(f"변환 실패: {notes}")
        return 2

    print("\n[1] 바꾼 파일에 조정 열이 만들어졌나 (사람이 값을 채울 자리)")
    wb = load_workbook(out, read_only=True, data_only=True)
    for name, _old, new_w, _first in TARGETS:
        head = [c for c in next(wb[name].iter_rows(min_row=1, max_row=1,
                                                   values_only=True))]
        while head and head[-1] in (None, ""):
            head.pop()
        check(f"{name} 머리글 수", len(head), new_w)
    wb.close()

    print("\n[2] 값을 안 넣으면 엔진이 받는 폭이 **옛날 그대로**인가")
    for name, old_w, _new, _first in TARGETS:
        arr = engine_table(out, name)
        check(f"{name} 엔진 폭", arr.shape[1], old_w)

    print("\n[3] 값을 넣으면 폭이 늘고 그 값이 제 자리에 꽂히나")
    for name, _old, new_w, first in TARGETS:
        wb = load_workbook(out)
        ws = wb[name]
        ws.cell(row=2, column=first, value=1)          # Ctrl Mode / Shunt Ctrl Mode = 1
        ws.cell(row=2, column=first + 2, value=1.02)   # Ctrl Target / Shunt Bmin 자리
        wb.save(out)
        wb.close()

        arr = engine_table(out, name)
        check(f"{name} 엔진 폭 (값을 넣은 뒤)", arr.shape[1], new_w)
        check(f"{name} 모드 칸 {first}열", float(arr[0, first - 1]), 1.0)
        check(f"{name} 그다음 값 {first + 2}열", float(arr[0, first + 1]), 1.02)

        # 되돌려 놓고 다음 시트로
        wb = load_workbook(out)
        ws = wb[name]
        # 🚨 `cell(..., value=None)` 으로는 안 지워진다 — openpyxl 이 None 이면 대입을
        #    건너뛴다. 반드시 `.value = None` 로 써야 한다.
        ws.cell(row=2, column=first).value = None
        ws.cell(row=2, column=first + 2).value = None
        wb.save(out)
        wb.close()

    print("\n[4] 되돌리면 폭도 되돌아오나 (한 번 넓어진 채 굳지 않는가)")
    for name, old_w, _new, _first in TARGETS:
        arr = engine_table(out, name)
        check(f"{name} 엔진 폭 (지운 뒤)", arr.shape[1], old_w)

    print("\n[5] 조정 열이 **옛 열을 밀어내지 않았나** (자리 어긋남이 제일 조용한 실패다)")
    for name, old_w, _new, _first in TARGETS:
        sheet = sheet_of(name)
        arr_v2 = engine_table(out, name)
        v1 = convert_case.load_workbook(V1_CASE, read_only=True, data_only=True)
        rows = convert_case._rows(v1[sheet.source_name])
        v1.close()
        want = np.array([[convert_case._num(v) for v in r[:old_w]] for r in rows],
                        dtype=float)
        got = arr_v2[:, :old_w]
        # 단위를 바꾸는 열은 v2 가 되돌린 값과 v1 값이 같아야 한다
        d = np.abs(np.nan_to_num(got) - np.nan_to_num(want[:got.shape[0], :]))
        check_small(f"{name} 옛 {old_w}열 최대 차이", float(np.max(d)) if d.size else 0.0)

    print(f"\n같음 {ok} · 다름 {ng}")
    return 1 if ng else 0


if __name__ == "__main__":
    sys.exit(main())
