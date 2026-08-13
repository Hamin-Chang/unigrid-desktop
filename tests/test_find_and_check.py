# -*- coding: utf-8 -*-
"""버스 번호로 찾기 · 점검 엑셀 내보내기 (2026-08-06 신설).

  🚨 찾기에서 제일 위험한 곳: **화면 줄 번호 ≠ 진짜 줄 번호**가 된다.
     좁혀 놓고 첫 줄을 껐는데 엉뚱한 선로가 꺼지면 조용히 틀린 답을 본다.

      python tests/test_find_and_check.py
"""
import os
import sys
import tempfile
from pathlib import Path

import numpy as np

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from PySide6.QtWidgets import QApplication, QMessageBox      # noqa: E402

qapp = QApplication.instance() or QApplication([])
import app as APP                                            # noqa: E402
for _n in ("warning", "information", "critical"):
    setattr(QMessageBox, _n, staticmethod(lambda *a, **k: None))
import scenario as SC                                        # noqa: E402
import checks                                                # noqa: E402
import exporter                                              # noqa: E402
import app_engine                                            # noqa: E402
from load_case import load_case                              # noqa: E402
from openpyxl import load_workbook                           # noqa: E402

bad = 0


def ok(cond, what, note=""):
    global bad
    print(f"  {'✅' if cond else '🚨'} {what}" + (f"  — {note}" if note else ""))
    if not cond:
        bad += 1


class Fake:
    def __init__(self, case):
        self.loaded_case = case


win = APP.Proto()
case = load_case(str(REPO / "cases/ACDC_71bus_3IC_parallel.xlsx"))
sol = app_engine.solve(case)
win.thread = Fake(case)
win._last_path = str(REPO / "cases/ACDC_71bus_3IC_parallel.xlsx")
win._solved(sol)

# ── 1. 찾기가 고르는 줄 ────────────────────────────────────────────────
print("1) 버스 번호로 찾기")
win.grid_key = "AC_Line_dat"
lines = SC._values(case, "AC_Line_dat")

win.grid_find = ""
ok(win.find_rows("AC_Line_dat", lines) is None, "비우면 전부 (좁히지 않는다)")

win.grid_find = "34"
got = win.find_rows("AC_Line_dat", lines)
names = [SC.describe_row(case, "AC_Line_dat", i) for i in got]
ok(all("34" in n for n in names), "고른 줄이 전부 34번을 낀다", " · ".join(names))
ok(len(got) < lines.shape[0], "좁혀졌다", f"{lines.shape[0]}줄 → {len(got)}줄")

win.grid_find = "34 8"
got2 = win.find_rows("AC_Line_dat", lines)
ok(set(got) <= set(got2), "번호를 더 주면 줄이 는다", f"{len(got)} → {len(got2)}줄")

win.grid_find = "999999"
ok(win.find_rows("AC_Line_dat", lines) == [], "없는 번호면 0줄")

# IC 는 AC·DC 양쪽 버스로 찾힌다
win.grid_key = "IC_dat"
ics = SC._values(case, "IC_dat")
win.grid_find = "39"                       # DC 쪽 버스
ok(len(win.find_rows("IC_dat", ics)) == 3, "IC 는 DC 쪽 번호로도 찾힌다 (병렬 3대)")
win.grid_find = "38"                       # AC 쪽 버스
ok(len(win.find_rows("IC_dat", ics)) == 3, "AC 쪽 번호로도 같은 3대")

# ── 2. 🚨 좁힌 채로 끄면 **진짜 그 줄**이 꺼지나 ────────────────────────
print("\n2) 좁힌 채로 끄기 — 화면 줄 ≠ 진짜 줄")
win.grid_key = "AC_Line_dat"
win.grid_find = "34"
win.rebuild()
seen = list(win._grid_rows)
ok(seen == win.find_rows("AC_Line_dat", lines), "표가 고른 줄만 들고 있다",
   f"{len(seen)}줄")
real = seen[0]
want = SC.describe_row(case, "AC_Line_dat", real)
win.changes = []
win.flip_row(real)
made = [c for c in win.changes if isinstance(c, SC.Cell)]
ok(len(made) == 1 and made[0].row == real, "끈 것이 진짜 그 줄", f"줄 {real} · {want}")
ok(want in SC.describe(win.changes), "이름도 그 줄", SC.describe(win.changes))

# 좁힌 채로 값을 고쳐도 진짜 줄에 들어가나 (발전기 지정전압)
win.changes = []
win.grid_key = "AC_gen_dat"
win.grid_find = "36"
win.rebuild()
seen_g = list(win._grid_rows)
ok(len(seen_g) >= 1, "발전기도 번호로 찾힌다", f"{len(seen_g)}대")
# 🚨 첫 번호 열을 왼쪽에 고정하면서 이 함수는 **표가 아니라 담는 상자**를 돌려준다.
#    미는 쪽 표는 앱이 `_grid_tb` 로, 화면 열 자리는 `_grid_off` 로 들고 있다(2026-08-13).
#    ⚠️ 돌려준 상자를 **붙잡고 있어야 한다** — 놓으면 그 안의 표까지 지워진다.
box = win.grid_table_widget()
tb, off = win._grid_tb, win._grid_off
# ⚠️ Vg 는 **데이터 7열**이다. 옛 시험은 주석에 "Vg" 라 써 놓고 4열(Droop Q-Vac)을
#    고치고 있었다 — 둘 다 고칠 수 있는 열이라 통과해서 안 드러났다(2026-08-13).
VG = 7
assert APP.GRID_HEADERS["AC_gen_dat"][VG].startswith("Vg"), \
    APP.GRID_HEADERS["AC_gen_dat"][VG]
item = tb.item(0, VG + off)
if item is not None:
    before = float(SC._values(case, "AC_gen_dat")[seen_g[0], VG])
    item.setText(f"{before + 0.02:.4f}")
    win.grid_edited("AC_gen_dat", item, off, APP.GRID_SCALES.get("AC_gen_dat", {}))
    cells = [c for c in win.changes if isinstance(c, SC.Cell)]
    ok(len(cells) == 1 and cells[0].row == seen_g[0],
       "값을 고쳐도 진짜 줄에 들어간다", f"줄 {seen_g[0]}")
win.grid_find = ""
win.changes = []

# ── 3. 점검이 엑셀로 나가나 ────────────────────────────────────────────
print("\n3) 점검 엑셀 내보내기")
names_x = exporter.table_names(sol)
ok("점검" in names_x, "내보낼 것 목록에 점검이 있다", " · ".join(names_x))
with tempfile.TemporaryDirectory() as d:
    files = exporter.save_tables(sol, Path(d), names_x)
    chk = [f for f in files if f.name == "Check_results.xlsx"]
    ok(len(chk) == 1, "Check_results.xlsx 가 나온다")
    wb = load_workbook(chk[0])
    ok(len(wb.sheetnames) == int(sol.n_time), "한 시각에 시트 하나",
       f"{len(wb.sheetnames)}장 · 시각 {sol.n_time}")
    ws = wb[wb.sheetnames[0]]
    rows = [[c for c in r] for r in ws.iter_rows(values_only=True)]
    head = rows[0]
    n_screen = checks.violation_count(checks.real_violations(sol, 0))
    ok(str(head[1]) == f"모두 {n_screen}건", "첫 줄 건수가 화면과 같다",
       f"{head[1]} vs 화면 {n_screen}건")
    titles = [str(r[0]) for r in rows if r and str(r[0]).startswith("▶")]
    ok(len(titles) == 4, "네 가지가 다 들어간다", " · ".join(titles))
    n_data = sum(1 for r in rows[1:]
                 if r and r[0] not in (None, "") and not str(r[0]).startswith("▶")
                 and str(r[0]) not in ("버스", "From", "변환기", "발전기"))
    ok(n_data == n_screen, "적힌 줄 수도 화면과 같다", f"{n_data}줄")

print(f"\n{'✅ 전부 통과' if bad == 0 else f'🚨 {bad}건 틀림'}")
sys.stdout.flush()
app_engine.shutdown()
os._exit(1 if bad else 0)
