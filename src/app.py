"""UNIGRID 데스크톱 — 앱 본체 (창·탭·표·그래프·계통도).

계산은 컴파일된 MATLAB 엔진이 한다(`app_engine` → `app_worker` → `engine/`).
케이스를 열기 전까지는 화면 모양을 보이려고 **가짜 값**을 그리고,
케이스를 열면 그 자리에 실제 결과가 들어온다.

실행:  ~/venvs/unigrid-acdc/bin/python src/app.py

⚠️ 아직 안 만든 것은 화면에서 그렇게 말한다 — 엑셀로 만들기(§7 3단계).
"""
import os
import sys
import time
import math
import re
import random
from pathlib import Path

import numpy as np
from PySide6.QtCore import Qt, QThread, QTimer, Signal
from PySide6.QtGui import QColor
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QGridLayout,
    QLabel, QPushButton, QFrame, QTabWidget, QTableWidget, QTableWidgetItem,
    QComboBox, QSpinBox, QDialog, QCheckBox, QLineEdit, QButtonGroup,
    QHeaderView, QScrollArea, QSizePolicy, QSplitter, QFileDialog,
    QProgressDialog, QMessageBox, QInputDialog, QSlider,
)

import json
import app_engine as ENGINE
import engine_path
import charts
import checks
from checks import (col_index, gen_limit_rows, real_violations,   # noqa: F401
                    unrated_lines, violation_count, GEN_LIMIT_COLS)
import exporter

# 케이스 읽기는 **이 저장소 안**에 있다 (2026-08-05 들여옴).
# 예전에는 옆 폴더의 공개 파이썬 패키지를 빌려 썼는데, 그 폴더는 이 맥에만 있어
# 남의 컴퓨터에서는 앱이 떠도 파일을 하나도 못 열었다 — 공개 배포가 안 되는 구조였다.
try:
    from load_case import load_case
except Exception:                      # 그래도 앱은 뜨게 (무엇이 없는지는 불러올 때 알린다)
    load_case = None

import scenario as SC          # 계통 조건 바꾸기 (PDR §7 2단계)


def _grid_headers():
    """계통 데이터 탭에 쓸 열 이름 — `format_v2` 의 v2 머리글을 엔진 자리에 맞춰 편다.

    이렇게 하면 서식 정의가 **한 곳**(format_v2)에만 있고, 화면은 그것을 빌려 쓴다.
    """
    try:
        import format_v2 as F
        import read_v2
    except Exception:
        return {}
    out = {}
    for sheet in F.SHEETS:
        key = read_v2.TABLE_OF_SHEET.get(sheet.name)
        if key is None:
            continue
        cols = [c for c in sheet.cols if c.v1_col]
        if not cols:
            continue
        wide = max(c.v1_col for c in cols)
        names = [""] * wide
        for c in cols:
            names[c.v1_col - 1] = c.header
        out[key] = names
    return out


GRID_HEADERS = _grid_headers()


def _grid_scales():
    """엔진 값 → 화면 값 배율. 🚨 이게 없으면 W 를 [MW] 머리글 아래 찍는다.

    (실제로 그랬다 — 발전기 `P_gen [MW]` 칸에 10000000 이 찍혔다. 10이어야 한다.)
    """
    try:
        import format_v2 as F
        import read_v2
    except Exception:
        return {}
    out = {}
    for sheet in F.SHEETS:
        key = read_v2.TABLE_OF_SHEET.get(sheet.name)
        if key is None:
            continue
        m = {c.v1_col - 1: c.scale for c in sheet.cols
             if c.v1_col and c.scale != F.KEEP}
        if m:
            out[key] = m
    return out


GRID_SCALES = _grid_scales()

# ③ 운전 조건 — **여기만 고칠 수 있다** (0부터 센 열 번호).
# 나머지 칸은 회색으로 두어 "④ 계통 자체는 엑셀에서" 라는 선을 화면으로 보여 준다(PDR §4.3).
# 머리글은 있는데 파일엔 아직 없는 열까지 **빈 칸으로 보여 줄** 표.
# A1 조정 열(14~19)은 옛 파일에 없다 — 안 보여 주면 켤 방법이 없다.
GRID_PAD_TO_HEADERS = {"AC_Line_dat", "AC_Bus_dat"}

GRID_EDITABLE = {
    "AC_Line_dat": {13, 14, 15, 16, 17, 18},   # A1 조정 — Mode·Bus·Target·Min·Max·Steps
    "AC_Bus_dat": {17, 18, 19, 20, 21},        # A1 ④ SVC — Shunt Mode·Target·Bmin·Bmax·Steps
    "AC_gen_dat": {2, 3, 4, 7},        # 운전모드 · P-f droop · Q-V droop · 지정전압
    "DC_gen_dat": {2, 3, 5},           # 운전모드 · P-Vdc droop · 지정전압
    "IC_dat": {2, 3, 4, 5, 6, 7, 8},   # AC/DC 제어모드 · droop 셋 · P·Q 동작점
    "DCDC_Conv_dat": {5, 6, 7, 9},     # droop 둘 · 동작점 · 운전모드
}

# 🚨 **비울 수 있는 칸** (2026-08-14 사용자 요청). 표에서 값을 지우면 NaN 이 되는데,
#   그게 뜻을 갖는 칸은 **A1 조정 칸뿐**이다 — 비어 있음 = 「안 걸었다」·「한계를 안 적어
#   자동으로 잡는다」. 엔진이 이미 그렇게 읽는다(`isnan` 검사가 다 있다).
#   ⚠️ 나머지 편집 칸(발전기 운전모드·droop·지정전압 등)은 **비우면 어떻게 되는지
#      확인하지 않았으므로 그대로 막는다.** 넓히려면 그 칸이 NaN 일 때 엔진이 무엇을
#      하는지 먼저 확인하고 여기 더한다.
#   계기: 계단을 켠 뒤 한계를 다시 비워 자동(0.9~1.1)으로 돌리려는데 **방법이 없었다** —
#         `float("")` 이 걸려 "숫자가 아닙니다" 로 되돌아갔다.
GRID_CLEARABLE = {
    "AC_Line_dat": {13, 14, 15, 16, 17, 18},   # Ctrl Mode·Bus·Target·Min·Max·Step Size
    "AC_Bus_dat": {17, 18, 19, 20, 21},        # Shunt Ctrl Mode·Target·Bmin·Bmax·Step Size
}

# 이 버스 수를 넘으면 **그래프를 접은 채로 연다** (2026-08-06 사용자 확정).
#   이유는 속도가 아니라 **읽기 어려워서**다 — 6,495버스면 점이 6,495개라 빨간 덩어리가 된다.
#   ⚠️ 접어도 별로 안 빨라진다(실측 0.86~1.49배, `실측_R4_버스수대시간.csv`). 표가 그만큼
#      넓어져 줄을 더 그리기 때문이다. 그러니 "빨라진다"고 말하지 않는다.
#   숫자는 점이 겹쳐 보이기 시작하는 지점으로 잡았다. 바꾸려면 여기 한 줄만 고치면 된다.
BIG_BUSES = 1000
# 시나리오 목록에서 **한 번에 보여 줄 줄 수**. 넘으면 목록 안에서 스크롤한다.
# 안 씌우면 한 줄에 25px 씩 카드가 계속 자라 그래프와 표를 먹는다(2026-08-15 실측).
SCENARIO_ROWS = 4
SCENARIO_ROW_H = 42   # 실측 줄 간격(렌더에서 잼 — sizeHint 증가분 25 와 다르다)

# 계통 데이터 탭에 보여 줄 표 (차례대로). 켜고 끌 수 있는 것이 앞에 온다.
GRID_TABLES = [
    ("AC_Line_dat", "AC 선로"), ("AC_gen_dat", "AC 발전기"),
    ("DC_Line_dat", "DC 선로"), ("DC_gen_dat", "DC 발전기"),
    ("IC_dat", "IC"), ("DCDC_Conv_dat", "DC/DC"),
    ("AC_Bus_dat", "AC 버스"), ("DC_Bus_dat", "DC 버스"),
]

RECENT_FILE = _HERE_RECENT = Path(__file__).resolve().parent / ".recent.json"


def load_recent() -> list:
    try:
        return json.loads(RECENT_FILE.read_text(encoding="utf-8"))[:6]
    except Exception:
        return []


def save_recent(path: str, info: str) -> None:
    items = [x for x in load_recent() if x.get("path") != path]
    items.insert(0, {"path": path, "info": info})
    try:
        RECENT_FILE.write_text(json.dumps(items[:6], ensure_ascii=False),
                               encoding="utf-8")
    except Exception:
        pass


class WarmThread(QThread):
    """시작 화면에 있는 동안 계산 엔진(MATLAB Runtime)을 미리 띄워둔다."""
    ready = Signal(bool)

    def run(self):
        try:
            ENGINE.warmup()
            self.ready.emit(True)
        except Exception:
            self.ready.emit(False)


class SolveThread(QThread):
    """조류계산을 화면과 별도로 돌린다 (7초쯤 걸려 UI가 멈추면 안 되므로)."""
    done = Signal(object)
    failed = Signal(str)
    # 계산 엔진(MATLAB Runtime)이 없는 것은 **계산 실패가 아니라 설치 안내**다.
    # 남의 컴퓨터에서 가장 먼저 만나는 화면이라 따로 받는다 (PDR §4.1 · R2).
    engine_missing = Signal(str)

    def __init__(self, path, case=None, method="nr"):
        super().__init__()
        self.path = path
        self.case = case          # 있으면 이것을 푼다 (조건을 바꿔 다시 풀 때)
        self.method = method      # "nr" Newton-Raphson · "gs" Gauss-Seidel (2026-08-12)
        self.loaded_case = None   # 파일에서 읽은 원본 — 창이 받아 들고 있는다

    def run(self):
        try:
            warm = ENGINE.solved_before()   # 첫 계산이면 준비 시간이 섞인다
            t0 = time.perf_counter()
            # 위험한 형태 거르기는 `load_case` 안으로 들어갔다 (2026-08-06)
            case = self.case if self.case is not None else load_case(self.path)
            self.loaded_case = case
            sol = ENGINE.solve(case, method=self.method)
            sol.seconds = time.perf_counter() - t0   # 엑셀 읽는 시간까지 포함
            sol.warm_start = warm
            sol.method = self.method                 # 어느 해법으로 푼 결과인지 (2026-08-12)
            self.done.emit(sol)
        except engine_path.EngineNotFound as exc:
            self.engine_missing.emit(str(exc))     # 안내문을 그대로 들고 있다
        except Exception as exc:
            self.failed.emit(str(exc))

class CurveThread(QThread):
    """PV·QV 곡선을 화면과 별도로 돌린다 (14버스도 6초, 큰 계통은 몇 분)."""
    done = Signal(object)
    failed = Signal(str)
    engine_missing = Signal(str)

    def __init__(self, case, load_buses, curve_buses):
        super().__init__()
        self.case = case
        self.load_buses = load_buses
        self.curve_buses = curve_buses

    def run(self):
        try:
            cur = ENGINE.curve(self.case, self.load_buses, self.curve_buses)
            self.done.emit(cur)
        except engine_path.EngineNotFound as exc:
            self.engine_missing.emit(str(exc))
        except Exception as exc:
            self.failed.emit(str(exc))


# ─────────────────────────────────────────── 색
LIGHT = dict(
    bg="#eef1f5", surface="#ffffff", border="#d5dae2", text="#1b2430",
    muted="#6b7684", accent="#0b6ab8", accent_soft="#e3eefb",
    ok="#1a7f4b", warn="#c2570e", plot="#f7f9fc",
)
DARK = dict(
    bg="#161a20", surface="#1e242c", border="#333c47", text="#e6ebf2",
    muted="#95a1b1", accent="#4da3ff", accent_soft="#22303f",
    ok="#3ecf8e", warn="#e0873f", plot="#232a33",
)

MODES = ["스냅샷", "다이나믹", "비교"]
# 최상위 갈래 — 조류계산과 곡선은 **서로 기대지 않는다** (F1d · 2026-08-12)
TASKS = ["조류계산", "PV·QV 곡선"]

GRAPHS = {
    "스냅샷": [
        ("전압·위상", ["전압  [pu]  ·  x축 = 버스", "위상각  [deg]  ·  x축 = 버스"], "v"),
        ("조류 (P·Q)", ["유효전력 P  (3D)", "무효전력 Q  (3D)"], "h"),
        ("부하율", ["선로 부하율  [%]"], "v"),
        ("토폴로지", ["계통 단선도"], "v"),
    ],
    "다이나믹": [
        ("전압·위상", ["전압  [pu]  ·  x축 = 시간", "위상각  [deg]  ·  x축 = 시간"], "v"),
        ("주파수", ["주파수  [Hz]  ·  x축 = 시간"], "v"),
        ("토폴로지", ["계통 단선도"], "v"),
    ],
}

# 열 이름은 result_columns.py 그대로. show=기본으로 보이는 열
TABLE_SPECS = {
    "AC 결과": [
        ("Bus", 1), ("VM[pu]", 1), ("Freq[pu]", 0), ("Angle[deg]", 1),
        ("Gen_P[MW]", 1), ("Gen_Q[MVAR]", 1), ("Load_P[MW]", 1),
        ("Load_Q[MVAR]", 1), ("toAC_P[MW]", 0), ("toAC_Q[MVAR]", 0),
        ("baseKV[kV]", 0), ("Vmin[pu]", 0), ("Vmax[pu]", 0),
    ],
    "DC 결과": [
        ("Bus", 1), ("VM[pu]", 1), ("VM_norm[pu]", 0), ("Gen_P[MW]", 1),
        ("Load_P[MW]", 1), ("toDC_P[MW]", 1), ("baseKV[kV]", 0),
        ("Vmin[pu]", 0), ("Vmax[pu]", 0),
    ],
    "선로 조류": [
        ("From", 1), ("To", 1), ("From_P[MW]", 1), ("To_P[MW]", 1),
        ("From_Q[MVAR]", 0), ("To_Q[MVAR]", 0), ("Loss_P[MW]", 1),
        ("Loss_Q[MVAR]", 0), ("Capacity[MVA]", 0), ("Loading[%]", 1),
        ("Status", 0),
    ],
    "손실": [
        ("Time[h]", 1), ("Ploss[W]", 1), ("Qloss[Var]", 1),
        ("Ploss[%]", 1), ("Qloss[%]", 1),
    ],
    "VSC 버스": [
        ("BusAC", 1), ("BusDC", 1), ("VSC_VM[pu]", 1), ("VSC_Angle[deg]", 1),
        ("Inj_P[MW]", 1), ("Inj_Q[MVAR]", 1), ("Loss[MW]", 1),
    ],
    "VSC 그리드전력": [
        ("BusAC", 1), ("BusDC", 1), ("Grid_P[MW]", 1), ("Grid_Q[MVAR]", 1),
        ("TrafFilter_P[MW]", 0), ("TrafFilter_Q[MVAR]", 0), ("Filter_Q[MVAR]", 0),
        ("VSCFilter_Q[MVAR]", 0), ("VSC_P[MW]", 1), ("VSC_Q[MVAR]", 1),
    ],
    "VSC 손실": [
        ("BusAC", 1), ("BusDC", 1), ("VSC_P[MW]", 0), ("VSC_Q[MVAR]", 0),
        ("Filter_Q[MVAR]", 0), ("TransfoLoss_P[MW]", 1), ("TransfoLoss_Q[MVAR]", 0),
        ("ReactorLoss_P[MW]", 1), ("ReactorLoss_Q[MVAR]", 0), ("VSCLoss_P[MW]", 1),
    ],
}
VSC_TABLES = ["VSC 버스", "VSC 그리드전력", "VSC 손실"]


def tables_for(mode, show_vsc):
    names = ["AC 결과", "DC 결과", "선로 조류"]
    if mode == "다이나믹":          # 손실 = 계통 전체·시간축 데이터
        names.append("손실")
    if show_vsc:
        names += VSC_TABLES
    return names


# 점검(위반) 가짜 데이터 — 실제로는 결과 표에서 걸러낸다
VIOLATIONS = {
    "전압 위반": (["Bus", "V[pu]", "한계", "초과량"],
        [["7", "1.0642", "Vmax 1.05", "+0.0142"],
         ["12", "0.9385", "Vmin 0.95", "-0.0115"]]),
    "과부하 선로": (["From", "To", "Loading[%]", "용량[MVA]"],
        [["3", "7", "112.4", "25.0"]]),
    "변환기 한계": (["BusAC", "BusDC", "상태", "S_max[MVA]"],
        [["4", "2", "용량곡선 도달", "10.0"]]),
    "발전기 한계": (["발전기", "항목", "걸린 한계", "출력", "한계값"],
        [["AC 3", "무효 Q", "용량 원 (S_N)", "18.4", "18.4"]]),
}
# 수렴 (실측값: CIGRE 케이스)
CONV = dict(
    converged=True, iters=2, threshold=1e-3,
    mis=[0.481353, 0.013484, 0.000377],
    blocks=["AC_P_mis", "AC_Q_mis", "DC_P_mis", "F_P_conv",
            "F_Q_conv", "F_P_AC", "F_Q_AC", "F_P_DC"],
    block_hist=[[0.505421, 0.362611, 0.16, 0, 0, 0, 0, 0],
                [0.380797, 0.019848, 0.001058, 1.8e-05, 0, 0, 0, 0],
                [0.413095, 0.052252, 9.5e-05, 2.2e-05, 0, 0, 0, 0]],
    dominant=["AC_P_mis", "AC_P_mis", "AC_P_mis"],
    seconds=1.39,
)



def dynamic_table(sol, bus_row):
    """다이나믹 — 고른 버스 하나의 시간별 값 (지금 앱에는 없던 표)."""
    n_ac = sol.AC.shape[0] if sol.AC.size else 0
    if bus_row < n_ac and sol.AC.size:
        arr, cols, which = sol.AC, sol.cols("AC"), "AC"
        row = bus_row
    elif sol.DC.size:
        arr, cols, which = sol.DC, sol.cols("DC"), "DC"
        row = bus_row - n_ac
    else:
        return "시간별 값", [], np.zeros((0, 0))
    row = max(0, min(row, arr.shape[0] - 1))
    keep = [c for c in ("VM[pu]", "Angle[deg]", "Gen_P[MW]", "Gen_Q[MVAR]",
                        "Load_P[MW]", "Load_Q[MVAR]", "toAC_P[MW]", "toDC_P[MW]")
            if c in cols]
    idx = [cols.index(c) for c in keep]
    T = arr.shape[2]
    out = np.zeros((T, 1 + len(idx)))
    out[:, 0] = np.arange(1, T + 1)
    for j, ci in enumerate(idx):
        out[:, j + 1] = arr[row, ci, :]
    bus_no = int(arr[row, 0, 0])
    return f"{which} {bus_no} 시간별", ["Time[h]"] + keep, out


def _tab_base(text: str) -> str:
    """탭 이름에서 건수를 뗀다 — "점검 (3)" · "계통 데이터 (2)" 는 같은 탭이다."""
    return text.split(" (")[0].strip()


def real_tables(sol, mode, t, show_vsc):
    """(탭이름, 열이름들, 값행렬) 목록 — 실제 결과에서."""
    out = []
    if mode == "다이나믹":
        label, cols, arr = dynamic_table(sol, t)   # t 자리에 bus_row가 온다
        if arr.size:
            out.append((label, cols, arr))
    for which, label in [("AC", "AC 결과"), ("DC", "DC 결과"),
                         ("Branch", "선로 조류")]:
        arr = sol.at(which, t)
        if arr.size:
            out.append((label, sol.cols(which), arr))
    if mode == "다이나믹" and sol.loss.size:
        out.append(("손실", sol.cols("Loss"), sol.loss))
    if show_vsc and sol.VSC_bus is not None and sol.VSC_bus.size:
        out.append(("VSC 버스", sol.cols("VSC_bus"), sol.VSC_bus))
    return out


COMPARE_ITEMS = [
    ("전압 크기", True), ("위상각", True),
    ("주파수", False), ("손실", False),      # False = 시간끼리 비교에서만
]


def fake(col, row):
    random.seed(col * 97 + row * 31)
    if col == 0:
        return str(row + 1)
    return f"{random.uniform(0.94, 1.06):.4f}" if col == 1 else \
           f"{random.uniform(-8, 8):.2f}"


def hline_soft(c):
    f = QFrame()
    f.setFrameShape(QFrame.HLine)
    f.setStyleSheet(f"color:{c['border']};background:{c['border']};max-height:1px;")
    return f


def _scrollable(page):
    """카드를 세로로 쌓은 화면을 **스크롤에 담는다**.

    담지 않으면 그 화면의 키가 곧 창의 최소 높이가 되어, 화면보다 큰 창이
    만들어지고 **아래가 잘린 채 줄일 수도 없다**(2026-08-13 실측: 점검 탭 490px
    + 그래프 470px 로 창 최소 높이가 1201px 이었다).
    """
    sa = QScrollArea()
    sa.setWidgetResizable(True)
    sa.setFrameShape(QFrame.NoFrame)
    sa.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
    sa.setWidget(page)
    sa.setMinimumHeight(0)
    return sa


# ─────────────────────────────────────────── 조각
class NumItem(QTableWidgetItem):
    """숫자로 견주는 표 칸 (2026-08-18).

    `QTableWidgetItem` 은 **보이는 글자로** 견준다. 그대로 정렬하면 버스가
    1 · 10 · 100 · 11 · 2 … 로 늘어서고, 전압도 `1.0993` 이 `0.983` 보다 앞선다.
    ⇒ 원래 숫자를 들고 있다가 그것으로 견준다.
    """

    __slots__ = ("_v",)

    def __init__(self, text, value):
        super().__init__(text)
        self._v = float(value)

    def __lt__(self, other):
        ov = getattr(other, "_v", None)
        if ov is None:
            return super().__lt__(other)
        a, b = self._v, ov
        if a != a:            # NaN 은 맨 뒤로 — 위로 올라오면 표가 못 쓰게 된다
            return False
        if b != b:
            return True
        return a < b


class _ClickLabel(QLabel):
    """누르면 알려 주는 라벨.

    한 번 = 그 시나리오로 가기 · 두 번 = 이름 고치기.
    (2026-08-15에 한 번 누르기를 더했다 — [결과 보기] 단추를 없애고 이름이 그 일을 맡는다.
     단추 하나가 줄 높이를 42px 로 만들어 목록이 자리를 크게 먹고 있었다.)
    """

    double_clicked = Signal(object)
    clicked = Signal(object)

    def mouseReleaseEvent(self, ev):
        # 🚨 두 번 누르기와 안 부딪히게 — Qt 는 두 번 누를 때도 뗌을 먼저 보낸다.
        #    잠깐 기다렸다가 그 사이에 두 번째가 안 오면 그때 한 번으로 친다.
        self._pending = getattr(self, "_pending", None) or QTimer(self)
        self._pending.setSingleShot(True)
        try:
            self._pending.timeout.disconnect()
        except (RuntimeError, TypeError):
            pass
        self._pending.timeout.connect(lambda: self.clicked.emit(self))
        self._pending.start(QApplication.doubleClickInterval())

    def mouseDoubleClickEvent(self, ev):
        if getattr(self, "_pending", None) is not None:
            self._pending.stop()             # 한 번 누르기로 새지 않게
        self.double_clicked.emit(self)


class Card(QFrame):
    def __init__(self, c):
        super().__init__()
        self.setObjectName("card")
        self.v = QVBoxLayout(self)
        self.v.setContentsMargins(14, 12, 14, 14)
        self.v.setSpacing(9)


class PlotBox(QFrame):
    """그래프 자리 (실제로 안 그림)."""

    def __init__(self, name, c):
        super().__init__()
        self.setObjectName("plot")
        self.setMinimumHeight(120)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        v = QVBoxLayout(self)
        v.setContentsMargins(12, 10, 12, 10)
        t = QLabel(name)
        t.setStyleSheet(f"color:{c['text']};font-size:14px;font-weight:600;")
        v.addWidget(t)
        v.addStretch()
        m = QLabel("〔 그래프 자리 〕")
        m.setAlignment(Qt.AlignCenter)
        m.setStyleSheet(f"color:{c['muted']};font-size:16px;")
        v.addWidget(m)
        v.addStretch()


# ─────────────────────────────────────────── 창들
class ConvertDialog(QDialog):
    def __init__(self, parent, c):
        super().__init__(parent)
        self.setWindowTitle("UNIGRID 엑셀로 만들기")
        self.setMinimumWidth(520)
        self.setStyleSheet(parent.styleSheet())
        v = QVBoxLayout(self)
        v.setContentsMargins(20, 18, 20, 18)
        v.setSpacing(11)
        t = QLabel("UNIGRID 엑셀로 만들기")
        t.setStyleSheet(f"color:{c['text']};font-size:18px;font-weight:700;")
        v.addWidget(t)
        s = QLabel("계산만 하려면 이 창은 필요 없습니다 — 파일을 바로 열면 됩니다.\n"
                   "여기서 만든 엑셀에 DC 버스·변환기·24시간 부하를 직접 넣어\n"
                   "AC/DC 혼합 케이스를 만들 수 있습니다.")
        s.setStyleSheet(f"color:{c['muted']};font-size:13px;line-height:150%;")
        v.addWidget(s)
        s2 = QLabel("어떤 형식에서 만들까요?")
        s2.setStyleSheet(f"color:{c['text']};font-size:14px;font-weight:600;")
        v.addWidget(s2)
        for name, desc, patt in [
                ("MATPOWER  (.m)", "AC 계통", "MATPOWER 케이스 (*.m)"),
                ("PSS/E  (.raw)", "AC 계통 · 3권선 포함", "PSS/E 계통 (*.raw *.RAW)"),
                ("MatACDC", "AC/DC 혼합 계통", "")]:
            b = QPushButton(f"{name}\n{desc}")
            b.setMinimumHeight(64)
            b.clicked.connect(lambda _, n=name, p=patt: self._pick(n, p))
            v.addWidget(b)
        note = QLabel("고르면 파일 선택 창 → 변환 → 저장 위치를 묻습니다")
        note.setStyleSheet(f"color:{c['muted']};font-size:13px;")
        v.addWidget(note)
        row = QHBoxLayout()
        row.addStretch()
        cancel = QPushButton("닫기")
        cancel.clicked.connect(self.reject)
        row.addWidget(cancel)
        v.addLayout(row)

    def _pick(self, name, patt):
        """고른 형식의 파일을 골라 → 읽어서 → v2 엑셀로 저장한다 (PDR §7 3단계 X1 (b))."""
        if not patt:
            # MatACDC 는 `.m` 을 **두 개** 받는다 (AC 계통 + DC 계통).
            # 🚨 순서가 중요하다 — 두 파일 다 확장자가 `.m` 이라 바꿔 고르면
            #    "busdc 를 찾을 수 없습니다" 같은 엉뚱한 말만 나온다. 먼저 알려 준다.
            if QMessageBox.question(
                    self, "파일을 두 개 고릅니다",
                    "MatACDC 계통은 파일이 둘로 나뉘어 있습니다.\n\n"
                    "    첫 번째 — AC 계통 (MATPOWER, bus·gen·branch)\n"
                    "    두 번째 — DC 계통 (MatACDC, busdc·convdc·branchdc)\n\n"
                    "이 순서로 고르셔야 합니다. 둘 다 확장자가 .m 이라\n"
                    "바꿔 고르면 읽지 못합니다.\n\n계속할까요?",
                    QMessageBox.Ok | QMessageBox.Cancel) != QMessageBox.Ok:
                return
            ac, _ = QFileDialog.getOpenFileName(
                self, "1/2 — 먼저 AC 계통 파일 (MATPOWER .m)", "", "MATPOWER 케이스 (*.m)")
            if not ac:
                return
            dc, _ = QFileDialog.getOpenFileName(
                self, "2/2 — 이제 DC 계통 파일 (MatACDC .m)", str(Path(ac).parent),
                "MatACDC 케이스 (*.m)")
            if not dc:
                return
            src = Path(ac)
            try:
                import unigrid_convert
                case = unigrid_convert.matacdc_to_case(ac, dc)
            except Exception as exc:
                QMessageBox.warning(self, "읽지 못했습니다",
                                    f"MatACDC 케이스를 읽지 못했습니다.\n\n{exc}")
                return
        else:
            src, _ = QFileDialog.getOpenFileName(self, f"{name} 파일 고르기", "", patt)
            if not src:
                return
            src = Path(src)
            try:
                case = load_case(str(src))
            except Exception as exc:
                QMessageBox.warning(self, "읽지 못했습니다",
                                    f"{src.name} 을 읽지 못했습니다.\n\n{exc}")
                return

        out, _ = QFileDialog.getSaveFileName(
            self, "UNIGRID 엑셀로 저장", str(src.with_name(f"{src.stem}_unigrid.xlsx")),
            "UNIGRID 케이스 (*.xlsx)")
        if not out:
            return

        try:
            import write_v2
            saved = write_v2.write_case(case, out)
        except Exception as exc:
            QMessageBox.warning(self, "저장하지 못했습니다", str(exc))
            return

        n_ac = len(case.tables.get("AC_Bus_dat", []))
        n_dc = len(case.tables.get("DC_Bus_dat", []))
        QMessageBox.information(
            self, "만들었습니다",
            f"{saved.name}\n\nAC 버스 {n_ac}개" + (f" · DC 버스 {n_dc}개" if n_dc else "") +
            "\n\n이 파일을 그대로 불러와 계산할 수 있습니다.\n"
            "DC 버스·변환기·24시간 부하는 엑셀에서 직접 채워 넣으세요.")
        self.accept()


class ImportDialog(QDialog):
    def __init__(self, parent, c):
        super().__init__(parent)
        self.setWindowTitle("불러오기")
        self.setMinimumWidth(560)
        self.setStyleSheet(parent.styleSheet())
        self.chosen = None
        v = QVBoxLayout(self)
        v.setContentsMargins(20, 18, 20, 18)
        v.setSpacing(11)
        t = QLabel("불러오기 — 계통 파일 선택")
        t.setStyleSheet(f"color:{c['text']};font-size:18px;font-weight:700;")
        v.addWidget(t)

        drop = QFrame()
        drop.setObjectName("plot")
        drop.setMinimumHeight(90)
        dv = QVBoxLayout(drop)
        d1 = QLabel("여기로 파일을 끌어다 놓거나")
        d1.setAlignment(Qt.AlignCenter)
        d1.setStyleSheet(f"color:{c['text']};font-size:16px;font-weight:600;")
        d2 = QLabel(".xlsx  ·  .m  ·  .raw     (형식은 자동으로 알아냄)")
        d2.setAlignment(Qt.AlignCenter)
        d2.setStyleSheet(f"color:{c['muted']};font-size:13px;")
        dv.addWidget(d1)
        dv.addWidget(d2)
        v.addWidget(drop)

        s = QLabel("최근에 연 파일")
        s.setStyleSheet(f"color:{c['muted']};font-size:14px;")
        v.addWidget(s)
        for f, info in [("ACDC_CIGRE_MVACMVDCLVDC.xlsx", "AC/DC 혼합 · AC 14 / DC 11"),
                        ("ACDC_71bus_3IC_parallel.xlsx", "AC/DC 혼합 · AC 38 / DC 33"),
                        ("matpower_ieee14.m", "AC only · 14")]:
            b = QPushButton(f"{f}\n{info}")
            b.setMinimumHeight(58)
            b.clicked.connect(lambda _, n=f, i=info: self._choose(n, i))
            v.addWidget(b)
        row = QHBoxLayout()
        row.addStretch()
        cancel = QPushButton("취소")
        cancel.clicked.connect(self.reject)
        pick = QPushButton("파일 찾기")
        pick.setObjectName("primary")
        pick.clicked.connect(lambda: self._choose("ACDC_CIGRE_MVACMVDCLVDC.xlsx",
                                                  "AC/DC 혼합 · AC 14 / DC 11"))
        row.addWidget(cancel)
        row.addWidget(pick)
        v.addLayout(row)

    def _choose(self, name, info):
        self.chosen = (name, info)
        self.accept()


class ExportDialog(QDialog):
    """내보내기 — 무엇을 어디에 저장할지 고르고 **실제로 파일을 쓴다**.

    파일 이름·시트 구성은 원본 MATLAB 앱과 똑같이 맞췄다(exporter.py 참고).
    원본은 묻지 않고 전부 저장했지만, 24시간 × 파일 4개는 시간이 걸려서
    고를 수 있게 뒀다.
    """

    def __init__(self, parent, c, mode, picked_items):
        super().__init__(parent)
        self.setWindowTitle("내보내기")
        self.setMinimumWidth(520)
        self.setStyleSheet(parent.styleSheet())
        self.win, self.c, self.mode = parent, c, mode
        self.sol = getattr(parent, "sol", None)
        self.tabs, self.compares = [], []       # (이름, 체크박스)
        self.folder = exporter.default_folder(
            self.sol.case_name if self.sol is not None else "결과")

        v = QVBoxLayout(self)
        v.setContentsMargins(20, 18, 20, 18)
        v.setSpacing(9)

        if self.sol is None:
            t = QLabel("내보내기")
            t.setStyleSheet(f"color:{c['text']};font-size:18px;font-weight:700;")
            v.addWidget(t)
            n = QLabel("저장할 결과가 없습니다. 먼저 케이스를 불러와 계산하세요.")
            n.setStyleSheet(f"color:{c['muted']};font-size:14px;")
            v.addWidget(n)
        elif mode == "비교":
            t = QLabel("내보내기 — 비교 표만 (그림은 여기서 저장하지 않습니다)")
            t.setWordWrap(True)
            t.setStyleSheet(f"color:{c['text']};font-size:17px;font-weight:700;")
            v.addWidget(t)
            # 여기서는 **표만** 저장한다. 비교 그림은 왼쪽 아래
            # "이 비교 그림 저장" 버튼이 따로 맡는다(사용자가 그렇게 나누기로 함).
            n = QLabel("여기서는 비교 표를 엑셀로 저장합니다. "
                       "그림은 왼쪽 아래 “이 비교 그림 저장” 버튼을 쓰세요.")
            n.setWordWrap(True)
            n.setStyleSheet(f"color:{c['muted']};font-size:13px;")
            v.addWidget(n)
            for name, always in COMPARE_ITEMS:
                on = name in picked_items and (
                    always or parent.compare_axis == "시간끼리")
                cb = QCheckBox(name if on else f"{name}   (화면에서 안 고름)")
                cb.setChecked(on)
                cb.setEnabled(on)
                v.addWidget(cb)
                self.compares.append((name, cb))
        else:
            t = QLabel("내보내기 — 무엇을 저장할까요?")
            t.setStyleSheet(f"color:{c['text']};font-size:18px;font-weight:700;")
            v.addWidget(t)
            l1 = QLabel("엑셀   (한 시간에 시트 하나)")
            l1.setStyleSheet(f"color:{c['muted']};font-size:14px;")
            v.addWidget(l1)
            for name in exporter.table_names(self.sol):
                fname = exporter.TABLE_FILES[name][1]
                cb = QCheckBox(f"{name}     {fname}")
                cb.setChecked(True)
                v.addWidget(cb)
                self.tabs.append((name, cb))
            l2 = QLabel("그림   (PNG · PDF 둘 다)")
            l2.setStyleSheet(f"color:{c['muted']};font-size:14px;")
            v.addWidget(l2)
            for tab, fname in exporter.figure_names(self.sol, mode):
                # 🚨 예전에는 그래프가 접혀 있으면 이 칸을 **꺼서 못 고르게** 했다
                #    (2026-08-18 삭제). 그런데 `exporter.save_figures` 는 화면의
                #    그래프를 쓰지 않고 **결과에서 새로 그린다** — 접혀 있어도 그림 8개가
                #    멀쩡히 나오는 것을 실측으로 확인했다. 막을 까닭이 없었다.
                #    ⚠️ 게다가 **저절로 접혔을 때도** 막혔다 — 선로 하나 껐을 뿐인데
                #       *"지금은 숫자 모드라 그림을 저장할 수 없습니다"* 가 떴다.
                #       고른 적 없는 모드를 이유로 대는 셈이었다.
                cb = QCheckBox(f"{tab}     {fname}.png / .pdf")
                cb.setChecked(True)
                v.addWidget(cb)
                self.tabs.append((tab, cb))

        path = QFrame()
        path.setObjectName("plot")
        pv = QHBoxLayout(path)
        pv.setContentsMargins(12, 8, 12, 8)
        self.path_label = QLabel()
        self.path_label.setStyleSheet(f"color:{c['text']};font-size:13px;")
        self._show_folder()
        pv.addWidget(self.path_label)
        pv.addStretch()
        pb = QPushButton("바꾸기")
        pb.clicked.connect(self.pick_folder)
        pv.addWidget(pb)
        v.addWidget(path)

        row = QHBoxLayout()
        row.addStretch()
        cancel = QPushButton("취소")
        cancel.clicked.connect(self.reject)
        self.save_btn = QPushButton("저장")
        self.save_btn.setObjectName("primary")
        self.save_btn.setEnabled(self.sol is not None)
        self.save_btn.clicked.connect(self.do_save)
        row.addWidget(cancel)
        row.addWidget(self.save_btn)
        v.addLayout(row)

    # ── 저장 위치 ──
    def _show_folder(self):
        """길면 홈 폴더를 ~ 로 줄여 보여 준다."""
        s = str(self.folder)
        home = str(Path.home())
        if s.startswith(home):
            s = "~" + s[len(home):]
        self.path_label.setText(f"저장 위치:   {s}")

    def pick_folder(self):
        got = QFileDialog.getExistingDirectory(
            self, "저장할 폴더를 고르세요", str(self.folder.parent))
        if got:
            self.folder = Path(got)
            self._show_folder()

    # ── 실제 저장 ──
    def do_save(self):
        picked = {n for n, cb in self.tabs + self.compares if cb.isChecked()}
        if not picked:
            QMessageBox.information(self, "내보내기", "저장할 항목을 하나 이상 고르세요.")
            return
        self.save_btn.setEnabled(False)
        bar = QProgressDialog("저장하는 중…", None, 0, len(picked), self)
        bar.setWindowTitle("내보내기")
        bar.setWindowModality(Qt.WindowModal)
        bar.setCancelButton(None)
        bar.setMinimumDuration(0)
        done = [0]

        def step(name):
            done[0] += 1
            bar.setValue(min(done[0], len(picked)))
            bar.setLabelText(f"{name} 저장했습니다")
            QApplication.processEvents()

        try:
            files = self._write(picked, step)
        except Exception as exc:            # 저장 실패를 조용히 넘기지 않는다
            bar.close()
            self.save_btn.setEnabled(True)
            QMessageBox.critical(self, "내보내기 실패", f"{exc}")
            return
        bar.close()
        if not files:
            self.save_btn.setEnabled(True)
            QMessageBox.information(self, "내보내기", "저장된 것이 없습니다.")
            return
        names = "\n".join(f"  · {p.name}" for p in files[:12])
        more = f"\n  … 그 밖에 {len(files) - 12}개" if len(files) > 12 else ""
        QMessageBox.information(
            self, "내보내기 완료",
            f"{len(files)}개 파일을 저장했습니다.\n\n{self.folder}\n\n{names}{more}")
        self.accept()

    def _write(self, picked, step):
        """고른 것을 실제로 쓴다. 돌려주는 값은 만들어진 파일 목록."""
        if self.mode == "비교":
            return self._write_compare(picked, step)
        files = []
        tables = [n for n in exporter.table_names(self.sol) if n in picked]
        if tables:
            files += exporter.save_tables(self.sol, self.folder, tables, step)
        figs = {t for t, _ in exporter.figure_names(self.sol, self.mode)} & picked
        if figs:
            files += exporter.save_figures(
                self.c, self.sol, self.mode, self.folder, figs,
                self.win.t, self.win.bus_row, step)
        return files

    def _write_compare(self, picked, step):
        """비교 표를 엑셀 한 권에 — 항목마다 시트 하나."""
        from openpyxl import Workbook
        targets = [t.strip() for t in self.win.compare_targets.split(",")
                   if t.strip()]
        axis = "Bus" if self.win.compare_axis == "버스끼리" else "Time"
        wb = Workbook()
        wb.remove(wb.active)
        for name, _ in self.compares:
            if name not in picked:
                continue
            got = self.win.compare_rows(name, targets)
            if got is None:
                continue
            head, rows = got
            ws = wb.create_sheet(name)
            ws.append([str(h) for h in head])
            for r in rows:
                ws.append([float(x) for x in r])
            step(name)
        if not wb.sheetnames:
            return []
        self.folder.mkdir(parents=True, exist_ok=True)
        p = self.folder / f"Comparison by {axis}.xlsx"
        wb.save(p)
        return [p]


# ─────────────────────────────────────────── 본 창
DROP_EXT = {".xlsx", ".m", ".raw"}


def dropped_path(event):
    """끌어다 놓은 것 중 우리가 읽을 수 있는 첫 파일 (없으면 None)."""
    md = event.mimeData()
    if not md.hasUrls():
        return None
    for url in md.urls():
        p = url.toLocalFile()
        if p and Path(p).suffix.lower() in DROP_EXT:
            return p
    return None


class Proto(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setAcceptDrops(True)     # 파일을 창에 끌어다 놓으면 바로 연다
        self.dropzone = None
        self.drop_label = None
        self.dark = False
        # ── 무엇을 할까 (PDR §7 4단계 F1d · 2026-08-12 사용자 확정 "조류계산이랑 아예 분리")
        #    곡선은 조류계산 결과(`self.sol`)에 **기대지 않는다.** 조류계산을 한 번도
        #    안 돌려도 곡선만 돌릴 수 있다. 공유하는 것은 케이스와 계통 조건뿐이다.
        self.task = "조류계산"        # "조류계산" · "PV·QV 곡선"
        self.cur = None               # 곡선 결과 (app_engine.Curve)
        self.curve_load = ""          # 부하를 늘릴 버스 (비면 부하가 있는 버스 전부)
        self.curve_pick = ""          # 곡선을 그릴 버스 (비면 늘린 버스와 같게)
        self.curve_x = "MW"           # 가로축 — "MW" 합계 부하 · "lambda" 배수
        self.curve_busy = False
        self.curve_err = ""
        self.numbers = False
        self.numbers_auto = False     # 큰 계통이라 **자동으로** 접힌 것인가
        self.graph_kept = False       # 사용자가 직접 펼친 적이 있나 (있으면 자동으로 안 접는다)
        self.numbers_why = ""         # 왜 접혔나 — "big"(큰 계통) · "changed"(조건을 바꿔 품)
        self.mode = "스냅샷"
        self.compare_axis = "버스끼리"
        self.overlay = set()          # 겹쳐 볼 시나리오 (Book 안 자리 번호) — 비면 전부
        self.compare_targets = "3, 7, 12"
        self.picked = {"전압 크기"}
        self.case = ("ACDC_CIGRE_MVACMVDCLVDC.xlsx", "AC/DC 혼합 · AC 14 / DC 11")
        self.sol = None               # 실제 계산 결과 (없으면 화면 모양용 가짜값)
        self.t = 0                    # 보고 있는 시간대 (0부터)
        self.bus_row = 0              # 다이나믹에서 고른 버스 (행 번호)
        self.case_has_vsc = True      # AC-only 케이스면 False
        self.show_vsc = False
        self.show_violations = False  # 계통도 '위반 보기' 켜짐 여부
        self.graph_tab = 0            # 보고 있던 그래프 탭 (재생성 때 되돌리려고)
        # 아래쪽 표 탭도 같이 기억한다. **번호가 아니라 이름**으로 — 모드·VSC 표시에 따라
        # 탭 개수가 달라지고, 이름에도 건수가 붙는다("점검 (3)"·"계통 데이터 (2)").
        self.table_tab = "AC 결과"
        # ── 계통 조건 (PDR §7 2단계) ──
        # 원본 케이스는 읽고 나면 바뀌지 않는다. 그 위에 "바꾼 것" 목록만 얹는다.
        self.base_case = None         # 파일에서 읽은 원본 (scenario.apply 의 바탕)
        self.applied = []             # **지금 화면의 결과**를 만든 조건 (이미 계산된 것)
        self.changes = []             # 그 위에 얹었지만 **아직 계산 안 한** 것
        self.book = SC.Book()         # 담아 둔 시나리오
        # 결과 표를 어떤 열로 늘어놓고 보고 있나 — {표 이름: (열 번호, 오름/내림)}.
        # 🚨 **계산해도 유지한다** (2026-08-18 사용자 확정). 「전압 낮은 순」으로 보다가
        #    선로를 끄고 다시 계산하면 또 전압 낮은 순으로 보여야, 무엇이 달라졌는지
        #    바로 견줄 수 있다. 매번 풀리면 그때마다 다시 눌러야 한다.
        self.sort_by = {}
        self._applying_sort = False   # 되돌아 부르는 것을 막는 빗장 (아래 _sort_changed)
        self._strips = {}             # 표 위 띠 — 정렬만 바뀌면 이것만 다시 채운다
        self.grid_key = "AC_Line_dat" # 계통 데이터 탭에서 보고 있는 표
        self.grid_find = ""           # 계통 데이터 탭에서 찾는 버스 번호
        self._grid_rows = []          # 화면 줄 → 진짜 줄 (찾기로 좁혔을 때)
        # 부하 일괄 증감 슬라이더 — 끌고 있는 동안 화면을 다시 그리면 손잡이가 사라진다.
        # 그래서 놓았을 때(sliderReleased) 또는 잠깐 멈췄을 때만 반영한다.
        self._load_pending = 1.0
        self._load_timer = QTimer(self)
        self._load_timer.setSingleShot(True)
        self._load_timer.timeout.connect(lambda: self.scale_loads(self._load_pending))
        self.visible = {k: {n for n, d in v if d} for k, v in TABLE_SPECS.items()}
        # 위아래 나눔 자리 — 탭 갈래마다 따로 ("grid" = 계통 데이터 · "other" = 나머지)
        self.split_sizes = {}
        # 계통 데이터 표를 다시 그릴 때 보던 자리로 되돌리려고 들고 있는 것
        # (2026-08-13 사용자: "숫자 하나 넣을 때마다 가로 스크롤을 다시 해야 한다")
        self._grid_tb = None           # 지금 화면의 표 (미는 쪽)
        self._grid_frozen = None       # 왼쪽에 고정한 표 (상태 + 첫 번호 열)
        self._grid_off = 0             # 화면 열 = 데이터 열 + off
        self._grid_view = None         # (가로, 세로) 스크롤 자리
        self._grid_focus = None        # 다음에 고를 칸 (화면 줄, 화면 열)
        self.setWindowTitle("UNIGRID")
        # 🚨 화면보다 큰 창으로 열면 **아래가 잘린 채 줄일 수도 없다**(2026-08-13).
        #    맥북 화면은 세로가 900 언저리라 950 도 넘친다. 쓸 수 있는 넓이에 맞춘다.
        w0, h0 = 1440, 950
        scr = QApplication.primaryScreen()
        if scr is not None:
            a = scr.availableGeometry()
            w0 = min(w0, a.width() - 40)
            h0 = min(h0, a.height() - 40)
        self.resize(w0, h0)
        self.build()

    # ── 테마 ──
    @property
    def c(self):
        return DARK if self.dark else LIGHT

    def qss(self):
        c = self.c
        return f"""
        QMainWindow, QWidget, QDialog {{ background:{c['bg']}; color:{c['text']};
            font-family:'Apple SD Gothic Neo','Helvetica Neue',sans-serif; }}
        QLabel, QCheckBox {{ background:transparent; }}
        /* 툴팁 — 전용 규칙이 없으면 macOS 는 배경을 어둡게 그리는데 위 QWidget
           규칙이 글자색까지 어둡게 강제해 '어두운 글자+어두운 배경'으로 안 보였다
           (계통도 위반 요소 손말풍선이 빈 상자로 떴다). 배경·글자·테두리를
           카드와 같은 팔레트로 못 박아 라이트·다크 양쪽에서 읽히게 한다. */
        QToolTip {{ background:{c['surface']}; color:{c['text']};
            border:1px solid {c['border']}; padding:6px 9px; font-size:13px; }}
        #dropzone {{ background:{c['surface']}; border:2px dashed {c['border']};
            border-radius:14px; }}
        #card, #plot {{ background:{c['surface']};
            border:1px solid {c['border']}; border-radius:9px; }}
        #plot {{ background:{c['plot']}; }}
        #topbar {{ background:{c['surface']}; border-bottom:1px solid {c['border']}; }}
        #sidebar {{ background:{c['surface']}; border-right:1px solid {c['border']}; }}
        #statusbar {{ background:{c['surface']}; border-top:1px solid {c['border']}; }}
        QPushButton {{ background:{c['surface']}; color:{c['text']};
            border:1px solid {c['border']}; border-radius:7px;
            padding:10px 17px; font-size:14px; }}
        QPushButton:hover {{ border-color:{c['accent']}; }}
        QPushButton#primary {{ background:{c['accent']}; color:#ffffff;
            border:none; font-weight:600; }}
        QPushButton#seg_on {{ background:{c['accent']}; color:#ffffff;
            border:none; font-weight:700; font-size:14px;
            border-radius:8px; padding:8px 14px; }}
        QPushButton#seg_off {{ background:transparent; color:{c['muted']};
            border:none; font-size:14px; border-radius:8px; padding:8px 14px; }}
        QPushButton#seg_off:hover {{ background:{c['accent_soft']};
            color:{c['accent']}; }}
        QPushButton#accentline {{ border:1px solid {c['accent']};
            color:{c['accent']}; font-weight:600; }}
        QPushButton#accentline:hover {{ background:{c['accent_soft']}; }}
        QPushButton#link {{ background:transparent; border:none;
            color:{c['muted']}; font-size:13px; padding:4px 6px;
            text-decoration:underline; }}
        QPushButton#link:hover {{ color:{c['accent']}; }}
        QComboBox, QSpinBox, QLineEdit {{ background:{c['surface']}; color:{c['text']};
            border:1px solid {c['border']}; border-radius:6px;
            padding:9px 12px; font-size:16px; }}
        QTabWidget::pane {{ border:1px solid {c['border']}; border-radius:8px;
            background:{c['surface']}; top:-1px; }}
        QTabBar::tab {{ background:transparent; color:{c['muted']};
            padding:10px 18px; font-size:14px; border:none;
            margin-right:3px; min-width:96px; }}
        QTabBar::tab:selected {{ color:{c['accent']};
            border-bottom:2px solid {c['accent']}; font-weight:700; }}
        QTableWidget {{ background:{c['surface']}; border:none;
            gridline-color:{c['border']}; font-size:14px; }}
        QHeaderView::section {{ background:{c['bg']}; color:{c['muted']};
            border:none; border-bottom:1px solid {c['border']};
            padding:9px; font-size:14px; font-weight:600; }}
        QCheckBox {{ font-size:14px; color:{c['text']}; spacing:7px; }}
        QScrollArea {{ border:none; background:transparent; }}
        """

    # ── 전체 다시 그리기 ──
    def build(self):
        self.setStyleSheet(self.qss())
        # 화면을 갈아끼우면 옛 위젯 참조는 버린다 (지워진 위젯을 만지면 죽는다)
        self.dropzone = self.drop_label = self._tabs = self._split = None
        if self.sol is None:
            self.setCentralWidget(self.start_page())
            return
        root = QWidget()
        self.setCentralWidget(root)
        v = QVBoxLayout(root)
        v.setContentsMargins(0, 0, 0, 0)
        v.setSpacing(0)
        v.addWidget(self.topbar())
        mid = QHBoxLayout()
        mid.setContentsMargins(0, 0, 0, 0)
        mid.setSpacing(0)
        # 왼쪽 줄도 스크롤에 담는다 — 카드를 다섯 장 쌓아 665px 이라, 안 담으면
        # 이것 하나로 창 최소 높이가 781px 이 된다(2026-08-13 실측).
        side = QScrollArea()
        side.setWidgetResizable(True)
        side.setFrameShape(QFrame.NoFrame)
        side.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        side.setWidget(self.sidebar())
        side.setFixedWidth(280)
        side.setMinimumHeight(0)
        mid.addWidget(side)
        mid.addWidget(self.center(), 1)
        v.addLayout(mid, 1)
        v.addWidget(self.statusbar())

    def rebuild(self):
        self._strips = {}       # 옛 띠는 곧 지워진다 — 죽은 위젯을 붙들지 않는다
        self._save_grid_view()
        # 그리기 **전에** 접힘 상태를 자리에 맞춘다. `_table_tab_changed` 에서도 부르지만
        # 그 길로만 오는 게 아니다 — 새 파일을 열 때 앞 파일의 탭이 그대로 남아 있고,
        # 표 한 칸을 고쳐도 여기로 온다. 여기서 맞춰 두면 어느 길로 와도 같아진다.
        # (다시 그리는 중이므로 돌아온 값은 안 본다 — 곧 그릴 화면이 이미 새 상태다.)
        self._fold_for_room()
        self.build()

    # ── 계통 데이터 표의 "보던 자리" ───────────────────────────────────
    # 값을 한 칸 고칠 때마다 화면을 통째로 다시 그리므로, 아무것도 안 하면 표가
    # 맨 왼쪽으로 돌아간다. 조정 열은 19열 중 14~19열이라 **매번 오른쪽 끝까지
    # 다시 밀어야 했다**(2026-08-13 사용자 지적). 자리를 기억했다가 되돌린다.

    def _save_grid_view(self):
        """지금 표가 어디를 보고 있었는지 적어 둔다."""
        tb = getattr(self, "_grid_tb", None)
        if tb is None:
            return
        try:
            self._grid_view = (tb.horizontalScrollBar().value(),
                               tb.verticalScrollBar().value())
        except RuntimeError:        # 이미 지워진 위젯
            pass
        self._grid_tb = self._grid_frozen = None

    def _restore_grid_view(self, tb):
        """되돌린다 — 자리를 먼저 맞추고, 다음에 칠 칸을 고른다."""
        try:
            if self._grid_view:
                h, v = self._grid_view
                tb.horizontalScrollBar().setValue(h)
                tb.verticalScrollBar().setValue(v)
            spot = getattr(self, "_grid_focus", None)
            if spot:
                r, cc = spot
                if r < tb.rowCount() and cc < tb.columnCount():
                    tb.setCurrentCell(r, cc)
                    it = tb.item(r, cc)
                    if it is not None:
                        tb.scrollToItem(it)   # 이미 보이면 안 움직인다
                    tb.setFocus()
                self._grid_focus = None
        except RuntimeError:
            pass

    # ── 상단 ──
    def topbar(self):
        c = self.c
        bar = QFrame()
        bar.setObjectName("topbar")
        bar.setFixedHeight(70)
        h = QHBoxLayout(bar)
        h.setContentsMargins(20, 0, 20, 0)
        h.setSpacing(9)

        logo = QLabel("UNIGRID")
        logo.setStyleSheet(
            f"color:{c['text']};font-size:23px;font-weight:800;letter-spacing:1.4px;")
        h.addWidget(logo)
        h.addSpacing(14)

        b1 = QPushButton("불러오기")
        b1.setObjectName("accentline")   # 가장 자주 쓰는 버튼이라 한 단계 위로
        b1.clicked.connect(self.do_import)
        # 비교 모드에서는 이 버튼이 **표만** 담당한다(비교 그림은 사이드바의
        # "이 비교 그림 저장" 버튼이 맡는다) → 누르기 전에 알 수 있게 이름에 박는다.
        b2 = QPushButton("내보내기 (표만)" if self.mode == "비교" else "내보내기")
        if self.mode == "비교":
            b2.setToolTip("비교 모드에서는 표(엑셀)만 내보냅니다.\n"
                          "비교 그림은 왼쪽 아래 “이 비교 그림 저장” 버튼을 쓰세요.")
        b2.clicked.connect(self.do_export)
        b3 = QPushButton("엑셀로 만들기")   # 케이스 만들 때만 쓰는 도구 → 맨 뒤
        b3.clicked.connect(lambda: ConvertDialog(self, c).exec())
        for b in (b1, b2, b3):
            h.addWidget(b)
        h.addStretch()

        # 🚨 여기 있던 「표시 모드 — 숫자만 / 그림 포함」 을 **없앴다** (2026-08-18 사용자
        #    지적 *"그래프 펼치기랑 숫자만/그림 포함 이게 좀 겹치는거 같은데"*).
        #    스위치는 `self.numbers` **하나**인데 이름이 넷이었다 —
        #    위쪽 [숫자만]/[그림 포함] · 그래프 자리 [그래프 접기]/[그래프 펼치기].
        #    2026-08-15 에 그래프 자리에 접기를 넣으면서 겹침을 **툴팁으로 때웠고**
        #    (*"위쪽 [숫자만] 과 같은 일입니다"*), 그 땜질이 결국 이 지적으로 돌아왔다.
        #    남긴 쪽은 **그래프 자리**다 — 그래프 바로 옆이라 무엇을 접는지 분명하고,
        #    접혀 있을 때는 안내 띠가 상태와 이유를 함께 말해 준다. 위쪽은 그래프에서
        #    멀고 `숫자만` 이라는 이름만으로는 무슨 일이 나는지 눌러 봐야 알았다.
        #    ⚠️ 함께 사라진 것 = 그 단추의 **거짓 툴팁** *"5~7배 빠릅니다"*. 같은 파일
        #       아래쪽(안내 띠)이 이미 *"사실이 아니다 — 실측 0.86~1.49배"* 라고 적어
        #       놨는데 툴팁만 옛 문구를 달고 남아 있었다.

        return bar

    # ── 좌측 ──
    def sidebar(self):
        c = self.c
        sb = QFrame()
        sb.setObjectName("sidebar")
        sb.setFixedWidth(280)
        v = QVBoxLayout(sb)
        v.setContentsMargins(14, 16, 14, 14)
        v.setSpacing(8)

        v.addWidget(self.case_card())
        v.addSpacing(12)

        # ── 무엇을 할까 — 케이스 다음으로 큰 갈림이라 맨 위에 둔다 (F1d)
        tl = QLabel("무엇을 할까")
        tl.setStyleSheet(f"color:{c['muted']};font-size:13px;font-weight:700;")
        v.addWidget(tl)
        tseg = QFrame()
        tseg.setObjectName("segwrap")
        tseg.setStyleSheet(
            f"#segwrap {{ background:{c['bg']};border:1px solid {c['border']};"
            f"border-radius:8px; }}")
        th = QVBoxLayout(tseg)
        th.setContentsMargins(3, 3, 3, 3)
        th.setSpacing(3)
        why = self.curve_why()
        for name in TASKS:
            b = QPushButton(name)
            b.setObjectName("seg_on" if self.task == name else "seg_off")
            b.setCursor(Qt.PointingHandCursor)
            if name == "PV·QV 곡선" and why:
                b.setEnabled(False)
                b.setToolTip(why)
            else:
                b.clicked.connect(lambda _, x=name: self.set_task(x))
            th.addWidget(b)
        v.addWidget(tseg)
        if why and self.task != "PV·QV 곡선":
            wn = QLabel(why)
            wn.setWordWrap(True)
            wn.setStyleSheet(f"color:{c['muted']};font-size:12px;")
            v.addWidget(wn)
        v.addSpacing(12)

        if self.task == "PV·QV 곡선":
            self.curve_controls(v)
            v.addStretch()
            return sb

        # 모드 3분할 — 가장 큰 선택이라 이름표를 붙인다
        ml = QLabel("보기")
        ml.setStyleSheet(f"color:{c['muted']};font-size:13px;font-weight:700;")
        v.addWidget(ml)
        seg = QFrame()
        seg.setObjectName("segwrap")
        seg.setStyleSheet(
            f"#segwrap {{ background:{c['bg']};border:1px solid {c['border']};"
            f"border-radius:8px; }}")
        sh = QHBoxLayout(seg)
        sh.setContentsMargins(3, 3, 3, 3)
        sh.setSpacing(3)
        for m in MODES:
            b = QPushButton(m)
            b.setObjectName("seg_on" if self.mode == m else "seg_off")
            b.setCursor(Qt.PointingHandCursor)
            b.clicked.connect(lambda _, x=m: self.set_mode(x))
            sh.addWidget(b)
        v.addWidget(seg)
        v.addSpacing(12)

        if self.mode == "스냅샷":
            lb = QLabel("시간 선택")
            lb.setStyleSheet(f"color:{c['muted']};font-size:13px;font-weight:700;")
            v.addWidget(lb)
            cb = QComboBox()
            n_t = self.sol.n_time if self.sol is not None else 24
            cb.addItems([f"{i} H" for i in range(1, n_t + 1)])
            cb.setCurrentIndex(min(self.t, n_t - 1))
            cb.currentIndexChanged.connect(self.set_time)
            v.addWidget(cb)
            n = QLabel("그래프와 표가 이 시간을 같이 따라갑니다")
            n.setWordWrap(True)
            n.setStyleSheet(f"color:{c['muted']};font-size:12px;")
            v.addWidget(n)
            v.addSpacing(14)
            v.addWidget(self.freq_card())

        elif self.mode == "다이나믹":
            lb = QLabel("버스 선택")
            lb.setStyleSheet(f"color:{c['muted']};font-size:13px;font-weight:700;")
            v.addWidget(lb)
            cb = QComboBox()
            if self.sol is not None and self.sol.AC.size:
                buses = [f"AC {int(b)}" for b in self.sol.AC[:, 0, 0]]
                if self.sol.DC.size:
                    buses += [f"DC {int(b)}" for b in self.sol.DC[:, 0, 0]]
            else:
                buses = [f"AC {i}" for i in range(1, 15)]
            cb.addItems(buses)
            cb.setCurrentIndex(min(self.bus_row, len(buses) - 1))
            cb.currentIndexChanged.connect(self.set_bus)
            v.addWidget(cb)
            n = QLabel("그래프와 표가 이 버스를 같이 따라갑니다")
            n.setWordWrap(True)
            n.setStyleSheet(f"color:{c['muted']};font-size:12px;")
            v.addWidget(n)

        else:  # 비교
            lb = QLabel("무엇끼리 비교")
            lb.setStyleSheet(f"color:{c['muted']};font-size:13px;font-weight:700;")
            v.addWidget(lb)
            seg2 = QFrame()
            seg2.setObjectName("segwrap")
            seg2.setStyleSheet(
                f"#segwrap {{ background:{c['bg']};border:1px solid {c['border']};"
                f"border-radius:8px; }}")
            s2 = QVBoxLayout(seg2)          # 세 개는 한 줄에 안 들어간다 — 두 줄로
            s2.setContentsMargins(3, 3, 3, 3)
            s2.setSpacing(3)
            top2 = QHBoxLayout()
            top2.setSpacing(3)
            for a in ["버스끼리", "시간끼리"]:
                b = QPushButton(a)
                b.setObjectName("seg_on" if self.compare_axis == a else "seg_off")
                b.clicked.connect(lambda _, x=a: self.set_axis(x))
                top2.addWidget(b)
            s2.addLayout(top2)
            b3 = QPushButton("시나리오끼리")
            b3.setObjectName(
                "seg_on" if self.compare_axis == "시나리오끼리" else "seg_off")
            b3.clicked.connect(lambda: self.set_axis("시나리오끼리"))
            s2.addWidget(b3)
            v.addWidget(seg2)
            v.addSpacing(8)

            if self.compare_axis == "시나리오끼리":
                lb3 = QLabel("겹쳐 볼 시나리오")
                lb3.setStyleSheet(f"color:{c['muted']};font-size:13px;font-weight:700;")
                v.addWidget(lb3)
                n3 = QLabel("위의 시나리오 목록에서 체크한 것을 겹쳐 그립니다.\n"
                            "전압·위상각은 x축이 버스, 주파수·손실은 x축이 시간입니다.")
                n3.setWordWrap(True)
                n3.setStyleSheet(f"color:{c['muted']};font-size:12px;")
                v.addWidget(n3)
                v.addStretch(1)
                return sb

            lb2 = QLabel("비교할 " + ("버스" if self.compare_axis == "버스끼리" else "시간"))
            lb2.setStyleSheet(f"color:{c['muted']};font-size:13px;font-weight:700;")
            v.addWidget(lb2)
            le = QLineEdit(self.compare_targets)
            le.setPlaceholderText("예: 3, 7, 12   (최대 50개)")
            le.textChanged.connect(self.set_targets)
            v.addWidget(le)
            n = QLabel("최대 50개 · 한 그래프에 겹쳐 그립니다")
            n.setWordWrap(True)
            n.setStyleSheet(f"color:{c['muted']};font-size:13px;")
            v.addWidget(n)
            v.addSpacing(10)

            lb3 = QLabel("볼 항목")
            lb3.setStyleSheet(f"color:{c['muted']};font-size:13px;font-weight:700;")
            v.addWidget(lb3)
            for name, always in COMPARE_ITEMS:
                usable = always or self.compare_axis == "시간끼리"
                cb = QCheckBox(name)
                cb.setChecked(name in self.picked and usable)
                cb.setEnabled(usable)
                cb.stateChanged.connect(
                    lambda st, n2=name: self.toggle_item(n2, st))
                v.addWidget(cb)
                if not usable:
                    w = QLabel("   시간끼리 비교에서만")
                    w.setStyleSheet(f"color:{c['warn']};font-size:12px;")
                    v.addWidget(w)
            # 비교 그림 저장은 위쪽 "내보내기"와 **따로** 둔다(사용자 요청).
            # 원본 앱도 비교는 별도 버튼이었다(ExportComparisonButtonPushed).
            # ⚠️ 이름을 sb 로 쓰면 안 된다 — 이 함수의 sb 는 사이드바 자체다.
            #    덮어쓰면 사이드바가 파이썬 참조를 잃고 사라진다(실제로 겪음).
            v.addSpacing(10)
            savebtn = QPushButton("이 비교 그림 저장  (PNG · PDF)")
            savebtn.clicked.connect(self.save_compare_figures)
            v.addWidget(savebtn)

        v.addStretch()
        return sb

    # ── 가운데 ──
    def center(self):
        c = self.c
        w = QWidget()
        v = QVBoxLayout(w)
        v.setContentsMargins(16, 14, 16, 12)
        v.setSpacing(11)

        if self.task == "PV·QV 곡선":
            bar = self.change_bar()      # 계통 조건은 조류계산과 **같이 쓴다**
            if bar is not None:
                v.addWidget(bar)
            v.addWidget(self.curve_page(), 1)
            return w

        if self.mode == "비교":
            # 겹쳐 보기 체크칸이 이 목록에 있다 — 비교 모드에서 오히려 더 필요하다
            sb = self.scenario_bar()
            if sb is not None:
                v.addWidget(sb)
            v.addWidget(self.compare_area(), 1)
            return w

        bar = self.change_bar()          # 바꾼 것이 있을 때만 나온다
        if bar is not None:
            v.addWidget(bar)
        # 🚨 시나리오 목록은 **여기 있지 않다** (2026-08-15). 위에 두면 줄마다 자리를 먹어
        #    (4줄 248px) 그래프와 표를 밀어냈다 ⇒ **아래 탭 하나로** 내려보냈다.
        #    위에는 `change_bar` 가 「지금 조건 + [⟲ 원본으로]」 한 줄만 남긴다.

        split = QSplitter(Qt.Vertical)
        split.setChildrenCollapsible(False)
        split.setHandleWidth(10)
        # 🚨 `findChildren` 으로 찾으면 **지워지기를 기다리는 옛 것**이 먼저 잡힌다.
        #    지금 화면의 것은 여기에 들고 있는다(표 `_grid_tb` 와 같은 이유).
        self._split = split

        # ── 그래프 (접혀 있으면 안내 띠로 바뀐다) ──
        if not self.numbers:
            gt = QTabWidget()
            for name, plots, layout in GRAPHS[self.mode]:
                page = QWidget()
                lay = QHBoxLayout(page) if layout == "h" else QVBoxLayout(page)
                lay.setContentsMargins(10, 10, 10, 10)
                lay.setSpacing(9)
                for pname in plots:
                    # 시간과 버스를 둘 다 넘긴다 — 어느 쪽을 쓸지는 그래프가
                    # 이름(x축이 버스냐 시간이냐)을 보고 고른다.
                    # (한때 하나만 넘겨서 고른 버스가 아니라 늘 첫 버스를 그렸다)
                    real = charts.build(pname, c, self.sol, self.t, self.bus_row,
                                        self.show_violations, self.set_violations,
                                        self.show_line_profile)
                    lay.addWidget(real if real is not None else PlotBox(pname, c))
                gt.addTab(page, name)
            # 보고 있던 탭을 되살린다 — 위반 보기 토글 등이 rebuild() 로 화면을
            # 다시 만드는데, 안 되살리면 늘 0번(전압·위상)으로 튀어 버렸다
            # (토폴로지에서 토글을 누르면 전압 그래프로 넘어가던 버그).
            gt.setCurrentIndex(min(self.graph_tab, gt.count() - 1))
            gt.currentChanged.connect(
                lambda i: setattr(self, "graph_tab", int(i)))
            # 그래프가 낮으면 QtCharts 가 x축 글자를 "..." 로 줄여 버린다. 그래서
            # **평소 높이**(아래 setSizes 620)는 넉넉히 준다. 다만 이걸 최소치로
            # 잡아 두면 창을 화면보다 작게 못 만든다 — 470 + 표 579 로 창 최소가
            # 1201px 이 돼 맥북 화면에서 아래가 잘렸다(2026-08-13 실측·사용자 지적).
            # 최소는 낮추고, 좁게 볼지는 가운데 손잡이로 사용자가 정한다.
            #
            # 🚨 **계통 데이터 탭에서는 낮춘다** (2026-08-15). 아래는 이미 표에 66% 를
            #    주도록 돼 있는데(`_apply_split`), 이 최소치가 그걸 막고 있었다 —
            #    `min(…, room - 260)` 에 걸려 표가 353 을 받으려다 275 로 깎였고,
            #    위쪽 띠가 늘어난 만큼의 손실이 **전부 표로** 갔다.
            #
            # ⚠️ 단 **얼마나 낮출 수 있는지는 차트가 몇 개 쌓이느냐가 정한다.**
            #    처음에 그냥 150 으로 낮췄더니 「전압·위상」(차트 둘이 위아래로)에서
            #    한 개당 90px 밖에 안 돌아가 **x축 글자가 `...` 로 뭉개지고 아래가 잘렸다**
            #    (사용자가 보고 *"여기가 이상한데?"*). 옆으로 놓는 탭(조류 P·Q)과 한 장짜리
            #    (부하율·토폴로지)는 한 줄이라 낮춰도 멀쩡하다.
            gt.setMinimumHeight(self._graph_floor())
            # 🚨 **펼치기가 있으면 접기도 있어야 한다** (2026-08-15 사용자 지적).
            #    접혀 있을 때는 [그래프 펼치기] 가 그래프 자리에 바로 있는데, 펼치고 나면
            #    되접는 길이 저 멀리 위쪽 [숫자만] 뿐이었다 — 이름도 달라 같은 일인 줄
            #    몰랐다. **2026-08-18 에 그 위쪽 단추를 아예 없애 여기가 유일한 길이 됐다.**
            #    탭 줄 오른쪽 구석에 둔다 — **세로 자리를 안 먹는다.**
            fold = QPushButton("그래프 접기")
            fold.setToolTip("그래프를 접고 표를 넓게 씁니다.")
            fold.setCursor(Qt.PointingHandCursor)
            fold.clicked.connect(lambda: self.set_numbers(True))
            gt.setCornerWidget(fold, Qt.TopRightCorner)
            split.addWidget(gt)
        else:
            note = QFrame()
            note.setObjectName("card")
            note.setMaximumHeight(58)
            nv = QHBoxLayout(note)
            nv.setContentsMargins(16, 10, 16, 10)
            # ⚠️ 예전 문구는 "약 5~7배 빠름" 이라고 말했는데 **사실이 아니다** —
            #    실측 0.86~1.49배다(`실측_R4_버스수대시간.csv`). 표가 그만큼 넓어져
            #    줄을 더 그리기 때문이다. 안 빨라지는 것을 빨라진다고 말하지 않는다.
            if self.numbers_why == "changed":
                msg = ("계통을 바꿔 계산해서 그래프를 접고 표를 넓게 폈습니다 — "
                       "펼치면 표가 그만큼 줄어듭니다.")
            elif self.numbers_why == "narrow":
                msg = ("표를 고치는 화면이라 그래프를 접고 표를 넓게 폈습니다 — "
                       "펼치면 표가 그만큼 줄어듭니다.")
            elif self.numbers_auto and self.sol is not None:
                n_bus = int(self.sol.AC.shape[0]) + \
                    int(self.sol.DC.shape[0] if self.sol.DC.size else 0)
                msg = (f"버스가 {n_bus:,}개라 그래프를 접어 두었습니다 — "
                       f"점이 겹쳐 읽기 어렵습니다. 값은 [엑셀로 만들기] 로 보세요.")
            else:
                msg = "그래프를 접고 표를 넓게 쓰는 중입니다"
            t = QLabel(msg)
            t.setStyleSheet(f"color:{c['muted']};font-size:14px;")
            nv.addWidget(t)
            nv.addStretch()
            b = QPushButton("그래프 펼치기")
            b.setToolTip("큰 계통에서는 점이 겹쳐 덩어리로 보입니다. 그래도 보시려면 누르세요.")
            b.clicked.connect(lambda: self.set_numbers(False))
            nv.addWidget(b)
            v.addWidget(note)

        # ── 표 ──
        tw = QWidget()
        tv = QVBoxLayout(tw)
        tv.setContentsMargins(0, 0, 0, 0)
        tv.setSpacing(7)

        # 🚨 이 줄은 **따로 한 줄을 쓰지 않는다** (2026-08-15). 표 묶음 186px 중 표에 남는
        #    것이 19px 뿐이었는데, 그 줄 하나가 46px 을 먹고 있었다.
        #    표 탭바 오른쪽 구석으로 옮기면 **세로 자리를 안 쓴다**(그래프 접기 단추와 같은 수법).
        head_w = QWidget()
        head = QHBoxLayout(head_w)
        head.setContentsMargins(0, 0, 0, 0)
        head.setSpacing(8)
        lab = QLabel("VSC 표")
        lab.setStyleSheet(f"color:{c['muted']};font-size:13px;")
        head.addWidget(lab)

        if not self.case_has_vsc:
            off = QLabel("  없음  ")
            off.setStyleSheet(
                f"background:{c['bg']};color:{c['muted']};border:1px solid "
                f"{c['border']};border-radius:7px;padding:5px 4px;font-size:12px;")
            off.setToolTip("이 케이스에는 변환기(VSC)가 없습니다")
            head.addWidget(off)
        else:
            seg = QFrame()
            seg.setObjectName("segwrap")
            seg.setFixedHeight(34)
            seg.setStyleSheet(
                f"#segwrap {{ background:{c['bg']};border:1px solid {c['border']};"
                f"border-radius:9px; }}")
            sh = QHBoxLayout(seg)
            sh.setContentsMargins(3, 3, 3, 3)
            sh.setSpacing(3)
            for txt, val in [("ON", True), ("OFF", False)]:
                b = QPushButton(txt)
                b.setObjectName("seg_on" if self.show_vsc == val else "seg_off")
                b.setCursor(Qt.PointingHandCursor)
                b.setFixedWidth(52)
                b.clicked.connect(lambda _, x=val: self.set_vsc(x))
                sh.addWidget(b)
            head.addWidget(seg)
        cb = QPushButton("열 선택")
        cb.clicked.connect(self.pick_columns)
        head.addWidget(cb)

        tt = QTabWidget()
        tt.setCornerWidget(head_w, Qt.TopRightCorner)
        self._tabs = tt
        tt.setTabPosition(QTabWidget.North)

        if self.sol is not None:
            key = self.bus_row if self.mode == "다이나믹" else self.t
            specs = real_tables(self.sol, self.mode, key,
                                self.show_vsc and self.case_has_vsc)
            bad = self.violating_buses()
            for name, cols, arr in specs:
                t = QTableWidget(arr.shape[0], len(cols))
                t.setHorizontalHeaderLabels(cols)
                t.verticalHeader().setVisible(False)
                t.verticalHeader().setDefaultSectionSize(30)
                t.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
                t.setAlternatingRowColors(True)
                warn = QColor(self.c["warn"])
                for r in range(arr.shape[0]):
                    flag = name in ("AC 결과", "DC 결과") and \
                        (name[:2], int(arr[r, 0])) in bad
                    for cc in range(len(cols)):
                        val = arr[r, cc]
                        txt = f"{val:.0f}" if cc == 0 and float(val).is_integer() \
                            else f"{val:,.4f}".rstrip("0").rstrip(".")
                        it = NumItem(txt, val)
                        if cc > 0:
                            it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                        if flag:
                            it.setForeground(warn)
                        t.setItem(r, cc, it)
                # 🚨 **칸을 다 채운 뒤에 켠다.** 채우기 전에 켜면 한 칸 넣을 때마다
                #    표를 다시 늘어놓아 느리고, 넣는 자리가 밀려 값이 어긋난다.
                # 🚨 그리고 **켜는 순간 Qt 가 제멋대로 한 번 늘어놓는다** —
                #    `setSortingEnabled(True)` 가 지금 표시자(기본 0열)로 `sortByColumn`
                #    을 부른다. 실측: 아무것도 안 눌렀는데 버스가 302·301·224… 로
                #    뒤집혀 나왔다. ⇒ 켜기 전에 **표시자를 지운다**(-1 = 없음).
                self._applying_sort = True
                try:
                    t.horizontalHeader().setSortIndicator(-1, Qt.AscendingOrder)
                    t.setSortingEnabled(True)
                finally:
                    self._applying_sort = False
                self._apply_sort(name, t, cols)
                t.horizontalHeader().sortIndicatorChanged.connect(
                    lambda col, order, nm=name: self._sort_changed(nm, col, order))
                tt.addTab(self._with_viol_legend(name, t, bad, cols), name)
        else:
            for name in tables_for(self.mode, self.show_vsc and self.case_has_vsc):
                cols = [n for n, _ in TABLE_SPECS[name] if n in self.visible[name]]
                t = QTableWidget(14, len(cols))
                t.setHorizontalHeaderLabels(cols)
                t.verticalHeader().setVisible(False)
                t.verticalHeader().setDefaultSectionSize(32)
                t.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
                t.setAlternatingRowColors(True)
                for r in range(14):
                    for cc in range(len(cols)):
                        it = QTableWidgetItem(fake(cc, r))
                        if cc > 0:
                            it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                        t.setItem(r, cc, it)
                tt.addTab(t, name)

        # 점검 · 수렴 탭 — 카드를 세로로 쌓아서 키가 크다(점검 490 · 수렴 378).
        # 그대로 넣으면 그 키가 **창의 최소 높이**가 돼 창을 못 줄인다. 스크롤에
        # 담아 창이 작아도 아래까지 갈 수 있게 한다(2026-08-13).
        n = violation_count(self.viol())
        tt.addTab(_scrollable(self.check_page()), f"점검 ({n})" if n else "점검")
        tt.addTab(_scrollable(self.conv_page()), "수렴")
        n_ch = len(self.changes)
        tt.addTab(self.grid_page(),
                  f"계통 데이터 ({n_ch})" if n_ch else "계통 데이터")
        sp = self.scenario_bar()
        if sp is not None:
            tt.addTab(_scrollable(sp), f"시나리오 ({len(self.book.items)})")

        # 보고 있던 탭으로 되돌린다 — 조건을 하나 바꿀 때마다 화면을 다시 그리므로,
        # 이걸 안 하면 매번 첫 탭(AC 결과)으로 튄다.
        self._restore_tab(tt)
        tv.addWidget(tt)
        # 표 묶음도 최소치를 못 박는다 — 안 그러면 가장 키 큰 탭이 창의 최소
        # 높이를 정해 버린다(Qt 는 최소치를 손으로 정하면 그것을 먼저 본다).
        tw.setMinimumHeight(170)
        split.addWidget(tw)

        if not self.numbers:
            split.setStretchFactor(0, 3)      # 표가 sizeHint 로 밀고 올라오는 걸 막는다
            split.setStretchFactor(1, 2)
            self._apply_split(split)
            # 탭을 옮기면 **그 탭에 맞는 자리**로 바꾼다 (화면은 다시 안 그린다)
            tt.currentChanged.connect(
                lambda i, w=tt, s=split: self._table_tab_changed(
                    _tab_base(w.tabText(i)), s))
            split.splitterMoved.connect(lambda *_: self._save_split(split))
        else:
            tt.currentChanged.connect(
                lambda i, w=tt: setattr(self, "table_tab", _tab_base(w.tabText(i))))
        v.addWidget(split, 1)
        return w

    # ── 위아래 나눔 ────────────────────────────────────────────────────
    # 「계통 데이터」 는 값을 고치는 탭이라 **줄이 여럿 보여야** 하고, 결과 탭은
    # 그래프가 커야 한다. 한 자리로 둘을 다 맞출 수 없어 **탭마다 따로** 기억한다
    # (2026-08-13 사용자: "계통 데이터 탭 여기가 너무 작아서 불편해").

    def _graph_rows(self) -> int:
        """지금 그래프 탭에서 차트가 **몇 줄로 쌓이나**. 옆으로 놓는 탭은 한 줄이다."""
        try:
            _name, plots, layout = GRAPHS[self.mode][self.graph_tab]
        except (KeyError, IndexError):
            return 2                      # 모르면 넉넉한 쪽으로
        return len(plots) if layout == "v" else 1

    # 그래프 높이 두 가지 — **다시 실측했다**(2026-08-18, 50버스 · 창 950px).
    # 🚨 2026-08-15 에 적어 둔 *"320 → 제대로 나온다"* 는 **틀렸다.** 두 줄짜리 탭에서
    #    320·360·400·420 이 전부 **세로축 숫자를 `...` 로 뭉갠다.** 숫자가 살아나는
    #    문턱은 **425** 다(420 뭉개짐 · 430 나옴). 한 줄짜리는 150 에서도 멀쩡하다.
    #    ⇒ 두 줄 바닥을 320 → **425** 로 올린다. 이 값을 못 주는 탭에서는 그래프를
    #      **접는다**(`_graph_fits`) — 못 읽는 그래프를 자리만 차지한 채 두지 않는다.
    #    함께 버린 안 둘: **차트 제목 떼기**(380 에서도 여전히 뭉갠다 — 제목이 원인이
    #    아니었다) · **전압·위상을 좌우로 놓기**(세로축은 살지만 50버스의 x축 번호가
    #    `1... 1... ...` 로 통째로 죽는다).
    GRAPH_FLOOR = {1: 150, 2: 425}
    GRAPH_WANT = {1: 220, 2: 380}

    # 점검 탭 표 — 한 줄 높이 · 머리글이 먹는 몫 · 몇 줄에서 끊나 (2026-08-18)
    #   한 줄 30px 은 `setDefaultSectionSize(30)` 으로 못박혀 있어 폰트와 무관하다.
    #   머리글은 이 맥에서 **40px 실측**인데 폰트가 바뀌면 커질 수 있어 **6px 여유**를 둔다
    #   (모자라면 마지막 줄이 잘리고, 남으면 표 아래에 빈 띠가 6px 생길 뿐이다).
    #   가로 스크롤은 열이 `QHeaderView.Stretch` 라 뜨지 않으므로 셈에서 뺀다(실측 확인).
    CHECK_ROW_H = 30
    CHECK_CHROME = 46
    CHECK_MAX_ROWS = 10

    def _graph_floor(self) -> int:
        """그래프에 남겨 둘 **최소** 높이 — **읽을 수 있는 크기**다.

        얼마나 필요한지는 차트가 몇 줄로 쌓이느냐가 정한다(위아래로 둘이면 한 장이
        절반씩 나눠 갖는다). ⚠️ 예전에는 **탭마다 달랐다** — 계통 데이터 탭은
        `GRAPH_FLOOR`, 나머지는 260. 그런데 260 도 320 도 세로축을 뭉개는 값이라
        (2026-08-18 실측) 탭을 가릴 이유가 없어졌다. **어느 탭이든 읽을 수 있는
        크기가 바닥이고, 그걸 못 주면 접는다.**
        """
        return self.GRAPH_FLOOR.get(self._graph_rows(), self.GRAPH_FLOOR[2])

    def _room_for_graph(self) -> int:
        """위아래로 나눠 쓸 수 있는 높이. `_apply_split` 이 쓰는 셈과 같게 맞춘다."""
        sp = self._split
        tot = sum(sp.sizes()) if sp is not None else 0
        room = tot if (sp is not None and sp.isVisible() and tot > 300) \
            else self.height() - 190
        return max(430, room)

    def _graph_fits(self, tab=None) -> bool:
        """그 탭에서 **읽을 수 있는 그래프**와 표 몫을 함께 줄 수 있나.

        계통 데이터 탭은 표에 66% 를 주기로 돼 있어(표를 고치는 곳이다) 그래프에
        남는 것이 34% 뿐이다. 950px 창이면 271px — **바닥값 425 에 한참 못 미친다.**
        그럴 때 예전에는 그래프를 425 로 붙들어 표를 373 으로 깎았는데, 그러면
        **표도 좁고 그래프도 못 읽는다.** 접으면 표가 전부 갖는다.
        """
        tab = self.table_tab if tab is None else tab
        if self._split_slot(tab) != "grid":
            return True          # 결과 탭은 그래프 쪽에 62% 를 주므로 늘 넉넉하다
        room = self._room_for_graph()
        return self._graph_floor() <= room - int(room * 0.66)

    def _fold_for_room(self, tab=None) -> bool:
        """자리 때문에 접거나 펴야 하면 상태를 바꾸고 True. 부르는 쪽이 다시 그린다.

        사용자가 **직접 펼친 뒤**(`graph_kept`)에는 건드리지 않는다 — 그때는
        `_apply_split` 이 그래프를 우선으로 놓아 읽을 수 있는 크기를 준다.
        """
        if self.graph_kept:
            return False
        fits = self._graph_fits(tab)
        if not self.numbers and not fits:
            self.numbers, self.numbers_auto = True, True
            self.numbers_why = "narrow"
            return True
        # 좁아서 접은 것이면, 자리가 나는 탭으로 가면 도로 편다
        if self.numbers and self.numbers_why == "narrow" and fits:
            self.numbers, self.numbers_auto = False, False
            self.numbers_why = ""
            return True
        return False

    def _graph_want(self) -> int:
        """사용자가 **직접 펼쳤을 때** 그래프에 주는 높이 — 읽을 수 있는 크기."""
        return self.GRAPH_WANT.get(self._graph_rows(), 380)

    # 표가 주인공인 탭 — 그래프보다 표에 자리를 몰아 준다 (2026-08-18 사용자 확정).
    #   계통 데이터 = 표를 **고치는** 곳 · 점검 = 무엇이 걸렸는지 **읽는** 곳.
    # 🚨 둘이 접히는 까닭은 서로 다르다. 계통 데이터는 표에 66% 를 주면 그래프에
    #    425px 이 안 남는다는 **자리 계산**으로 접힌다(`_graph_fits`). 점검은 자리로만
    #    보면 431px 이 나오지만, 표 셋(전압 위반·과부하 선로·발전기 한계)이 352px 에
    #    들어가질 않아 **한 표도 온전히 안 보였다**(실측: 다 보이는 표 0개 → 접으면 2개).
    #    그래서 여기 목록에 넣어 같은 길을 타게 한다.
    TABLE_FIRST = ("계통 데이터", "점검")

    def _split_slot(self, tab=None):
        tab = self.table_tab if tab is None else tab
        # 점검 탭은 이름이 "점검 (26)" 처럼 건수를 달고 다니므로 앞부분으로 본다
        return "grid" if any(tab.startswith(n) for n in self.TABLE_FIRST) else "other"

    def _apply_split(self, split, tab=None):
        slot = self._split_slot(tab)
        saved = (self.split_sizes or {}).get(slot)
        if saved:
            split.setSizes(saved)
            return
        # 화면에 붙어 있으면 진짜 높이를, 아직 만드는 중이면 창 높이로 어림한다
        tot = sum(split.sizes())
        room = tot if (split.isVisible() and tot > 300) else self.height() - 190
        room = max(430, room)
        # 🚨 **직접 펼쳤으면 그래프가 우선이다** (2026-08-15 사용자 확정 —
        #    *"사용자가 키고 싶으면 그때 그래프를 띄우고 표는 작게 줄이자"*).
        #    계통 데이터 탭은 평소 표에 66% 를 주는데, 그 규칙을 그대로 두면 그래프가
        #    늘 바닥값만 받아 **축 글자가 뭉개진 채**로 보인다. 보겠다고 누른 사람에게
        #    못 읽는 그래프를 주는 것은 안 켜 준 것과 같다.
        # ⚠️ 예전에는 `slot == "grid"` 로 잠겨 있어 **계통 데이터 탭에만** 걸렸다.
        #    그 바람에 결과 탭에서 [그래프 펼치기] 를 누르면 393px 밖에 안 받아
        #    **펼쳐 놓고도 세로축이 뭉개졌다**(2026-08-18 실측). 바로 위 주석이
        #    금지한 그 상태다. ⇒ 어느 탭에서 눌러도 읽을 수 있는 크기를 준다.
        if self.graph_kept and not self.numbers:
            # ⚠️ `room` 으로 깎지 않는다 — 다시 그리는 도중에는 그 값이 실제보다 작게
            #    잡혀(470 으로 잡힌 적이 있다) 그래프 요청이 270 까지 깎였고, 결국 바닥값만
            #    받았다. **원하는 비율로 넘기고 남는 자리는 Qt 가 나눠 준다.**
            split.setSizes([self._graph_want(), 200])
            return
        # 데이터 고칠 땐 아래를 크게, 결과 볼 땐 그래프를 크게.
        share = 0.66 if slot == "grid" else 0.38
        # 그래프에 남겨 둘 최소치. **위 `gt.setMinimumHeight` 과 같은 값이어야 한다** —
        # 어긋나면 그래프 최소를 낮춰 놓고도 여기서 도로 깎아 표가 안 넓어진다(2026-08-15).
        keep = self._graph_floor()
        bottom = min(max(int(room * share), 350), room - keep)
        split.setSizes([room - bottom, bottom])

    def _save_split(self, split):
        """손으로 끈 자리는 **그 탭 것으로만** 기억한다."""
        if not isinstance(self.split_sizes, dict):
            self.split_sizes = {}
        self.split_sizes[self._split_slot()] = split.sizes()

    def _table_tab_changed(self, name, split):
        if name == self.table_tab:
            return
        self._save_split(split)          # 떠나는 탭의 자리를 먼저 적어 둔다
        self.table_tab = name
        # 자리가 모자라 접거나, 자리가 나서 도로 펴야 하면 화면을 다시 그린다
        # (그래프 자리가 통째로 안내 띠로 바뀌므로 split 만 고쳐서는 안 된다)
        if self._fold_for_room(name):
            self.rebuild()
            return
        self._apply_split(split)

    # ── 시작 화면 (파일을 아직 안 불러왔을 때) ──
    def start_page(self):
        c = self.c
        page = QWidget()
        outer = QVBoxLayout(page)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        body = QWidget()
        v = QVBoxLayout(body)
        v.setContentsMargins(60, 0, 60, 0)
        v.setSpacing(0)
        v.addStretch()

        title = QLabel("UNIGRID")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet(
            f"color:{c['text']};font-size:44px;font-weight:800;letter-spacing:4px;")
        v.addWidget(title)
        sub = QLabel("AC/DC 통합 조류계산")
        sub.setAlignment(Qt.AlignCenter)
        sub.setStyleSheet(f"color:{c['muted']};font-size:16px;")
        v.addWidget(sub)
        v.addSpacing(34)

        # 파일 놓는 자리
        drop = QFrame()
        drop.setObjectName("dropzone")
        drop.setMinimumHeight(170)
        drop.setMaximumWidth(720)
        dv = QVBoxLayout(drop)
        dv.setSpacing(7)
        d1 = QLabel("계통 파일을 여기로 끌어다 놓으세요")
        d1.setAlignment(Qt.AlignCenter)
        d1.setStyleSheet(f"color:{c['text']};font-size:19px;font-weight:700;")
        dv.addWidget(d1)
        self.dropzone = drop        # 끌어다 놓을 때 밝히려고 들고 있는다
        self.drop_label = d1
        d2 = QLabel("UNIGRID 엑셀 (.xlsx)  ·  MATPOWER (.m)  ·  PSS/E (.raw)"
                    "     — 형식은 자동으로 알아냅니다")
        d2.setAlignment(Qt.AlignCenter)
        d2.setStyleSheet(f"color:{c['muted']};font-size:13px;")
        dv.addWidget(d2)
        dv.addSpacing(10)
        row = QHBoxLayout()
        row.addStretch()
        pick = QPushButton("파일 고르기")
        pick.setObjectName("primary")
        pick.setMinimumSize(150, 44)
        pick.clicked.connect(self.do_import)
        row.addWidget(pick)
        row.addStretch()
        dv.addLayout(row)
        hb = QHBoxLayout()
        hb.addStretch(); hb.addWidget(drop); hb.addStretch()
        v.addLayout(hb)
        v.addSpacing(12)

        # 작은 링크 — 계산이 아니라 "케이스 파일을 새로 만드는" 도구
        lk = QHBoxLayout()
        lk.addStretch()
        conv = QPushButton("AC 계통 파일을 UNIGRID 엑셀로 만들기 (DC·변환기를 직접 넣으려면)")
        conv.setObjectName("link")
        conv.setCursor(Qt.PointingHandCursor)
        conv.clicked.connect(lambda: ConvertDialog(self, c).exec())
        lk.addWidget(conv)
        lk.addStretch()
        v.addLayout(lk)
        v.addSpacing(22)

        # 최근에 연 파일
        recent = load_recent()
        if recent:
            cap = QLabel("최근에 연 파일")
            cap.setStyleSheet(f"color:{c['muted']};font-size:13px;font-weight:700;")
            cw = QWidget()
            cw.setMaximumWidth(720)
            cvv = QVBoxLayout(cw)
            cvv.setContentsMargins(0, 0, 0, 0)
            cvv.setSpacing(7)
            cvv.addWidget(cap)
            for item in recent:
                name = Path(item["path"]).name
                b = QPushButton(f"{name}\n{item.get('info', '')}")
                b.setMinimumHeight(54)
                b.setStyleSheet(
                    f"text-align:left;padding:9px 14px;font-size:14px;")
                b.clicked.connect(lambda _, p=item["path"]: self.open_path(p))
                cvv.addWidget(b)
            hb2 = QHBoxLayout()
            hb2.addStretch(); hb2.addWidget(cw); hb2.addStretch()
            v.addLayout(hb2)

        v.addStretch()
        outer.addWidget(body, 1)

        # 아래 — 엔진 준비 상태
        bar = QFrame()
        bar.setObjectName("statusbar")
        bar.setFixedHeight(42)
        bh = QHBoxLayout(bar)
        bh.setContentsMargins(20, 0, 20, 0)
        self.warm_dot = QLabel("●")
        self.warm_dot.setStyleSheet(f"color:{c['muted']};font-size:13px;")
        self.warm_txt = QLabel("계산 엔진 준비 중...")
        self.warm_txt.setStyleSheet(f"color:{c['muted']};font-size:13px;")
        bh.addWidget(self.warm_dot)
        bh.addWidget(self.warm_txt)
        bh.addStretch()
        ver = QLabel("UNIGRID Desktop — 개발 중")
        ver.setStyleSheet(f"color:{c['muted']};font-size:12px;")
        bh.addWidget(ver)
        outer.addWidget(bar)

        if ENGINE.is_ready():
            self._warm_done(True)
        elif getattr(self, "warm", None) is None:
            self.warm = WarmThread()
            self.warm.ready.connect(self._warm_done)
            self.warm.start()
        return page

    def _warm_done(self, ok):
        if getattr(self, "warm_dot", None) is None:
            return
        c = self.c
        self.warm_dot.setStyleSheet(
            f"color:{c['ok'] if ok else c['warn']};font-size:13px;")
        self.warm_txt.setText("계산 엔진 준비됨" if ok
                              else "계산 엔진을 띄우지 못했습니다")
        self.warm_txt.setStyleSheet(
            f"color:{c['ok'] if ok else c['warn']};font-size:13px;")

    # ── 파일 끌어다 놓기 ──
    def dragEnterEvent(self, e):
        if dropped_path(e):
            e.acceptProposedAction()
            self.set_hot(True)
        else:
            e.ignore()

    def dragMoveEvent(self, e):
        if dropped_path(e):
            e.acceptProposedAction()
        else:
            e.ignore()

    def dragLeaveEvent(self, e):
        self.set_hot(False)

    def dropEvent(self, e):
        path = dropped_path(e)
        self.set_hot(False)
        if not path:
            e.ignore()
            return
        e.acceptProposedAction()
        self.open_path(path)

    def set_hot(self, on):
        """끌어온 파일이 창 위에 있을 때 놓는 자리를 밝게."""
        if self.dropzone is None:
            return
        c = self.c
        if on:
            self.dropzone.setStyleSheet(
                f"#dropzone {{ background:{c['accent_soft']};"
                f" border:2px dashed {c['accent']}; border-radius:14px; }}")
            if self.drop_label is not None:
                self.drop_label.setText("놓으면 바로 계산합니다")
                self.drop_label.setStyleSheet(
                    f"color:{c['accent']};font-size:19px;font-weight:700;")
        else:
            self.dropzone.setStyleSheet("")
            if self.drop_label is not None:
                self.drop_label.setText("계통 파일을 여기로 끌어다 놓으세요")
                self.drop_label.setStyleSheet(
                    f"color:{c['text']};font-size:19px;font-weight:700;")

    def open_path(self, path):
        """최근 파일에서 바로 열기."""
        if not Path(path).exists():
            QMessageBox.warning(self, "불러오기", "파일을 찾을 수 없습니다.")
            return
        self._start_solve(path)

    def case_card(self):
        """지금 무엇을 보고 있는지 — 파일 · 계통 종류 · 버스 수."""
        c = self.c
        name, info = self.case
        mode_txt = info.split("·")[0].strip()          # 예: "AC/DC 혼합"
        counts = info.split("·")[1].strip() if "·" in info else ""

        box = QFrame()
        box.setObjectName("card")
        v = QVBoxLayout(box)
        v.setContentsMargins(14, 12, 14, 13)
        v.setSpacing(7)

        top = QHBoxLayout()
        cap = QLabel("현재 케이스")
        cap.setStyleSheet(f"color:{c['muted']};font-size:12px;font-weight:700;")
        top.addWidget(cap)
        top.addStretch()
        ch = QPushButton("바꾸기")
        ch.setFixedHeight(24)
        ch.setStyleSheet(
            f"border:none;background:transparent;color:{c['accent']};"
            f"font-size:12px;padding:0;")
        ch.setCursor(Qt.PointingHandCursor)
        ch.clicked.connect(self.do_import)
        top.addWidget(ch)
        v.addLayout(top)

        stem = name.replace(".xlsx", "").replace(".m", "").replace(".raw", "")
        # Qt 는 밑줄을 줄 끊는 자리로 안 쳐서 긴 이름이 그냥 잘린다.
        # 밑줄 뒤에 폭 0짜리 끊김표(U+200B)를 넣어 두 줄까지 보이게 한다.
        f = QLabel(stem.replace("_", "_​"))
        f.setWordWrap(True)
        f.setToolTip(str(getattr(self, "_last_path", "") or name))
        f.setStyleSheet(f"color:{c['text']};font-size:16px;font-weight:700;")
        v.addWidget(f)

        ext = QLabel(name.split(".")[-1].upper() + " 파일")
        ext.setStyleSheet(f"color:{c['muted']};font-size:11px;")
        v.addWidget(ext)

        # 계통 종류 = 알약 표시
        pill = QLabel("  " + mode_txt + "  ")
        pill.setStyleSheet(
            f"background:{c['accent_soft']};color:{c['accent']};"
            f"border-radius:9px;padding:4px 6px;font-size:12px;font-weight:700;")
        pr = QHBoxLayout()
        pr.addWidget(pill)
        pr.addStretch()
        v.addLayout(pr)

        # 버스 수
        if counts:
            v.addWidget(hline_soft(c))
            cr = QHBoxLayout()
            cr.setSpacing(18)
            for part in counts.split("/"):
                part = part.strip()
                if not part:
                    continue
                bits = part.split()
                lab = bits[0] if bits else part
                num = bits[-1] if len(bits) > 1 else ""
                col = QVBoxLayout()
                col.setSpacing(0)
                a = QLabel(lab + " 버스")
                a.setStyleSheet(f"color:{c['muted']};font-size:11px;")
                b = QLabel(num)
                b.setStyleSheet(
                    f"color:{c['text']};font-size:18px;font-weight:700;")
                col.addWidget(a)
                col.addWidget(b)
                cr.addLayout(col)
            # baseMVA — 계산 결과가 아니라 계통 속성이라 여기가 맞는 자리
            if self.sol is not None:
                col = QVBoxLayout()
                col.setSpacing(0)
                a = QLabel("baseMVA")
                a.setStyleSheet(f"color:{c['muted']};font-size:11px;")
                b = QLabel(f"{self.sol.baseMVA:g}")
                b.setStyleSheet(
                    f"color:{c['text']};font-size:18px;font-weight:700;")
                col.addWidget(a)
                col.addWidget(b)
                cr.addLayout(col)
            cr.addStretch()
            v.addLayout(cr)
        return box

    def freq_card(self):
        """시스템 주파수 — 계통 전체에 하나뿐인 값이라 크게 보여준다."""
        c = self.c
        if self.sol is not None and self.sol.freq.size:
            f = float(self.sol.freq[min(self.t, self.sol.freq.size - 1)])
        else:
            f = 60.02
        # 기준 주파수는 케이스마다 다르다 (60 Hz / 50 Hz) — 못 박으면 안 된다
        nominal = self.sol.freq_nominal if self.sol is not None else 60.0
        # 데드밴드도 케이스 파일에서 읽는다. 예전엔 ±0.05 Hz 라고 내가 정한 값을
        # 썼는데, 실제 값은 0.036 Hz 이거나 아예 0 이다(app_engine._freq_deadband).
        db = self.sol.freq_db if self.sol is not None else 0.0
        dev = f - nominal
        box = QFrame()
        box.setObjectName("card")
        v = QVBoxLayout(box)
        v.setContentsMargins(14, 11, 14, 13)
        v.setSpacing(3)
        t = QLabel("시스템 주파수")
        t.setStyleSheet(f"color:{c['muted']};font-size:12px;font-weight:600;")
        v.addWidget(t)
        row = QHBoxLayout()
        row.setSpacing(5)
        big = QLabel(f"{f:.2f}")
        big.setStyleSheet(f"color:{c['text']};font-size:30px;font-weight:800;")
        row.addWidget(big)
        unit = QLabel("Hz")
        unit.setStyleSheet(f"color:{c['muted']};font-size:14px;")
        unit.setAlignment(Qt.AlignBottom)
        row.addWidget(unit)
        row.addStretch()
        v.addLayout(row)
        d = QLabel(f"기준 {nominal:.0f} Hz 대비 {dev:+.2f} Hz")
        d.setStyleSheet(f"color:{c['muted']};font-size:12px;")
        v.addWidget(d)
        # 데드밴드는 "발전기가 주파수에 응동하기 시작하는 폭"이다. 밖에 있다고
        # 잘못된 상태가 아니라 **발전기가 응동 중**이라는 뜻이라 경고색을 안 쓴다.
        # 진짜 위반(전압·과부하·변환기 한계)은 상태바와 점검 탭이 따로 센다.
        if db > 0:
            where = "안 — 발전기 응동 없음" if abs(dev) <= db else "밖 — 발전기 응동 중"
            txt = f"데드밴드 ±{db:g} Hz {where}"
        else:
            txt = "데드밴드 없음 — 작은 편차에도 발전기가 응동"
        s = QLabel(txt)
        s.setStyleSheet(f"color:{c['muted']};font-size:12px;")
        v.addWidget(s)
        return box

    def viol(self):
        """지금 화면의 위반 목록 (실제 결과가 있으면 실제값)."""
        if self.sol is None:
            return VIOLATIONS
        return real_violations(self.sol, self.t)

    def _apply_sort(self, name, table, cols):
        """적어 둔 정렬을 다시 건다. 없으면 원래 순서 그대로 둔다."""
        want = self.sort_by.get(name)
        if not want:
            return
        col, order = want
        if not (0 <= col < len(cols)):
            self.sort_by.pop(name, None)      # 열 선택으로 그 열이 사라졌다
            return
        # 🚨 빗장 — `sortItems` 도 `sortIndicatorChanged` 를 낸다. 안 막으면
        #    `_sort_changed` → `rebuild` → 여기 → 또 신호 로 **끝없이 돈다.**
        self._applying_sort = True
        try:
            table.sortItems(col, order)
        finally:
            self._applying_sort = False

    def _sort_changed(self, name, col, order):
        """열 머리를 눌렀다 — 적어 두고 **띠만** 다시 채운다.

        🚨 처음엔 여기서 `rebuild()` 를 불렀는데 **못 쓸 만큼 느렸다**(실측:
           50버스 0.17초 · 1,888버스 2.05초 · **6,495버스 5.97초**). 정렬은 "이 열로
           봐야지" 하고 자주 누르는 것이고, **정렬이 가장 필요한 게 큰 계통인데
           거기서 제일 느리다.** 게다가 다시 그릴 까닭도 없다 — 표는 Qt 가 이미
           늘어놓았고, 달라지는 것은 위 띠 한 줄뿐이다.
        """
        if self._applying_sort:
            return
        self.sort_by[name] = (int(col), order)
        self._fill_strip(name)

    def _clear_sort(self, name):
        self.sort_by.pop(name, None)
        self.rebuild()

    def _fill_strip(self, name):
        """표 위 띠의 내용을 다시 채운다. 보일 것이 없으면 숨긴다.

        띠를 **미리 만들어 두고 내용만 갈아 끼운다** — 정렬을 누를 때마다 화면을
        통째로 다시 그리지 않으려면 이 방법뿐이다(위 `_sort_changed` 참고).
        """
        keep = self._strips.get(name)
        if keep is None:
            return
        bar, cols = keep
        h = bar.layout()
        while h.count():                      # 옛 내용을 비운다
            it = h.takeAt(0)
            w = it.widget()
            if w is not None:
                w.setParent(None)
                w.deleteLater()

        c = self.c
        bad = self.violating_buses()
        n = (sum(1 for grid, _bus in bad if grid == name[:2])
             if name in ("AC 결과", "DC 결과") else 0)
        sorted_by = self.sort_by.get(name)
        # 🚨 **정렬 중이면 위반이 없어도 띄운다** (2026-08-18 사용자 확정).
        #    안 그러면 성한 계통에서 정렬했을 때 되돌릴 길이 화면에 없다 —
        #    열 머리를 다시 눌러도 오름 ↔ 내림만 오갈 뿐 「원래 순서」로는 못 간다.
        if not n and not sorted_by:
            bar.hide()
            return
        bar.show()

        if n:
            sq = QLabel("■")
            sq.setStyleSheet(f"color:{c['warn']};font-size:13px;")
            h.addWidget(sq)
            txt = QLabel(f"주황 = 전압이 한계를 벗어난 버스 {n}곳")
            txt.setStyleSheet(f"color:{c['text']};font-size:13px;")
            h.addWidget(txt)

        if sorted_by:
            col, order = sorted_by
            col_name = cols[col] if cols and 0 <= col < len(cols) else "?"
            way = "작은 값부터" if order == Qt.AscendingOrder else "큰 값부터"
            if n:
                sep = QLabel("·")
                sep.setStyleSheet(f"color:{c['border']};font-size:13px;")
                h.addWidget(sep)
            st = QLabel(f"{col_name} {way} 늘어놓음")
            st.setStyleSheet(f"color:{c['accent']};font-size:13px;font-weight:600;")
            h.addWidget(st)
            back = QPushButton("원래 순서로")
            back.setCursor(Qt.PointingHandCursor)
            back.setStyleSheet(
                f"border:none;background:transparent;color:{c['muted']};"
                f"font-size:13px;padding:0 4px;")
            back.clicked.connect(lambda _=False, nm=name: self._clear_sort(nm))
            h.addWidget(back)

        h.addStretch()

        if n:
            go = QPushButton("자세히 보기  →")
            go.setCursor(Qt.PointingHandCursor)
            go.setStyleSheet(
                f"border:none;background:transparent;color:{c['accent']};"
                f"font-size:13px;font-weight:600;padding:0;")
            go.clicked.connect(self.go_check)
            h.addWidget(go)

    def _with_viol_legend(self, name, table, bad, cols=None):
        """표 위에 띠를 얹어 돌려준다 — **주황이 무슨 뜻인지**와 **정렬 상태**.

        계기(주황) — AC·DC 결과 표에서 전압 한계를 벗어난 버스는 **줄 전체가 주황**
        인데, 그 규칙이 화면 어디에도 안 적혀 있었다. 표에 `Vmin[pu]`·`Vmax[pu]` 열이
        있어 눈으로 짚으면 짐작은 되지만, **짐작해야 한다는 게 문제**다.
        계기(정렬) — 열 머리로 늘어놓고 나면 **원래 순서로 돌아갈 길**이 필요하다.

        자리는 **탭 줄과 표 사이**다. 표 위로 끼어들므로 어떤 줄도 가리지 않는다.
        ⚠️ 보일 것이 없어도 **띠는 만들어 둔다(숨김)** — 정렬을 누를 때 이 띠만
           갈아 끼우려면 미리 있어야 한다(`_fill_strip`). 안 그러면 화면을 통째로
           다시 그려야 하고, 그건 6,495버스에서 6초다.
        """
        c = self.c
        box = QWidget()
        v = QVBoxLayout(box)
        v.setContentsMargins(0, 0, 0, 0)
        v.setSpacing(0)

        bar = QFrame()
        bar.setObjectName("violbar")
        bar.setFixedHeight(38)
        bar.setStyleSheet(
            f"#violbar {{ background:{c['accent_soft']};"
            f"border-bottom:1px solid {c['border']}; }}")
        h = QHBoxLayout(bar)
        h.setContentsMargins(14, 0, 12, 0)
        h.setSpacing(8)

        self._strips[name] = (bar, list(cols or []))
        v.addWidget(bar)
        v.addWidget(table, 1)
        # 🚨 **채우기는 레이아웃에 넣은 뒤에** (2026-08-18). Qt 는 위젯을 레이아웃에
        #    넣을 때 부모가 보이면 **자동으로 다시 보이게** 한다. 먼저 채우면서
        #    `bar.hide()` 를 불러도 그 다음 `addWidget` 이 도로 켜 버려서,
        #    위반도 정렬도 없는 계통에 **빈 파란 띠**가 그대로 남았다(시험이 잡았다).
        self._fill_strip(name)
        return box

    def violating_buses(self):
        """표에서 빨갛게 칠할 (계통, 버스번호) 집합."""
        out = set()
        if self.sol is None:
            return out
        cols, rows = self.viol().get("전압 위반", ([], []))
        for r in rows:
            parts = r[0].split()
            if len(parts) == 2 and parts[1].isdigit():
                out.add((parts[0], int(parts[1])))
        return out

    # ── 점검 탭 ──
    # ══════════════════════════ 계통 조건 (PDR §7 2단계) ══════════════════════════
    # 바꾸는 동안은 **계산하지 않는다.** 다 바꾸고 [이 조건으로 계산] 을 누를 때 한 번 푼다
    # (2026-08-06 사용자 확정). 버튼 한 번 = 조류계산 한 번 = 시나리오 한 줄.

    def change_bar(self):
        """무엇을 바꿨는지 + [이 조건으로 계산] · [되돌리기] · [원본으로].

        ⚠️ **예전에는 `self.changes` 가 있을 때만 만들었다.** 그래서 [이 조건으로 계산] 을
           누르는 순간 띠가 통째로 사라졌고, **원본으로 돌아갈 길이 화면에서 없어졌다.**
           길이 없던 게 아니라 시나리오 목록 안에 숨어 있었는데, 그 목록은 시나리오가
           2개 이상일 때만 그려지고 거기가 *조건을 되돌리는 곳* 이라고 말하지도 않는다.
           (사용자: *"선로 4번을 껐는데 5번을 끄고 싶다. 직접 되돌리는 게 불안하다"*)

        ⚠️ **2026-08-15 자리를 옮겼다.** 처음엔 굳은 상태에서도 이 띠를 남겼는데,
           **시나리오 카드와 같은 말을 두 곳에서** 하면서 44px + 간격을 먹었고 그만큼
           아래 표가 눌렸다(사용자: *"표가 너무 작아서 보기 힘들다"*). ⇒ 「⟲ 원본으로」는
           **시나리오 카드 머리로** 옮기고, 이 띠는 원래대로 *아직 안 푼 것이 있을 때만* 뜬다.
           단 카드가 안 그려지는 경우(시나리오가 하나뿐 — 원본을 지웠을 때)에는
           돌아갈 길이 또 사라지므로 그때는 여기가 맡는다.
        """
        # 시나리오 목록이 아래 탭으로 내려갔으므로(2026-08-15), 굳은 조건이 있으면
        # **늘** 이 한 줄이 나온다 — 지금 무슨 조건을 보고 있는지와 돌아갈 길.
        if not self.changes and not self.applied:
            return None
        c = self.c
        pending = bool(self.changes)          # 아직 안 푼 것이 있나
        edge = c["warn"] if pending else c["border"]
        bar = QFrame()
        bar.setObjectName("card")
        bar.setStyleSheet(
            f"#card {{ background:{c['surface']};border:1px solid {edge};"
            f"border-radius:10px; }}")
        h = QHBoxLayout(bar)
        h.setContentsMargins(15, 9, 11, 9)
        h.setSpacing(11)

        if pending:
            tag = QLabel(f"바꾼 것 {len(self.changes)}건")
            tag.setStyleSheet(
                f"color:{c['warn']};font-size:12px;font-weight:600;padding:2px 9px;"
                f"border:1px solid {c['warn']};border-radius:9px;")
            h.addWidget(tag)

            what = QLabel(SC.describe(self.changes))
            what.setStyleSheet(f"color:{c['text']};font-size:13px;font-weight:600;")
            h.addWidget(what)
        else:
            # 다 계산해 놓은 상태 — 경고가 아니라 **지금 무슨 조건을 보고 있나**를 말한다.
            tag = QLabel("지금 조건")
            tag.setStyleSheet(
                f"color:{c['muted']};font-size:12px;padding:2px 9px;"
                f"border:1px solid {c['border']};border-radius:9px;")
            h.addWidget(tag)

            what = QLabel(SC.describe(self.applied))
            what.setStyleSheet(f"color:{c['text']};font-size:13px;font-weight:600;")
            h.addWidget(what)

        if pending:
            if self.task == "PV·QV 곡선":
                # 곡선은 **이 조건을 이미 쓴다**(`curve_case` 가 changes 까지 얹는다).
                # 조류계산 쪽 문구("아직 계산 안 함")를 그대로 두면 곡선이 옛 조건으로 그려진
                # 줄 알게 된다 — 실제로는 여유가 이미 달라져 있다.
                wait = QLabel("곡선은 이 조건으로 그립니다")
                wait.setStyleSheet(f"color:{c['muted']};font-size:12px;")
            else:
                base = "원본" if not self.applied else SC.describe(self.applied)
                wait = QLabel(f"아직 계산 안 함 — 지금 화면은 「{base}」 결과입니다")
                wait.setStyleSheet(f"color:{c['warn']};font-size:12px;")
            h.addWidget(wait)

            # 쪼개짐은 **막지 않고 알려만 준다** (2026-08-06 확정 — 71bus 는 모든 선로가 쪼갠다)
            if self.base_case is not None:
                msg = SC.splits(self.base_case, self.changes)
                if msg:
                    warn = QLabel("계통이 쪼개집니다")
                    warn.setStyleSheet(
                        f"color:{c['warn']};font-size:12px;font-weight:600;"
                        f"background:{c['bg']};border-radius:8px;padding:3px 9px;")
                    warn.setToolTip(msg + "\n\n막지는 않습니다. 떨어져 나간 쪽에 전원이 없으면 "
                                          "그 답은 뜻이 없으니 결과를 볼 때 감안하세요.")
                    h.addWidget(warn)
        h.addStretch(1)

        if pending and self.task != "PV·QV 곡선":
            # 곡선 화면에는 안 붙인다 — 여기서 누르면 **조류계산**이 돈다.
            # 곡선은 왼쪽 [곡선 그리기] 가 자기 실행이다.
            run = QPushButton("▶  이 조건으로 계산")
            run.setObjectName("primary")
            run.setCursor(Qt.PointingHandCursor)
            run.clicked.connect(self.run_changes)
            h.addWidget(run)

        if pending:
            undo = QPushButton("↩ 되돌리기")
            undo.setToolTip("아직 계산 안 한 것만 물립니다.\n"
                            "이미 계산해서 보고 있는 조건은 그대로 둡니다.")
            undo.setCursor(Qt.PointingHandCursor)
            undo.clicked.connect(self.undo_changes)
            h.addWidget(undo)

        # 이미 계산해 굳은 조건이 있을 때만 — 「되돌리기」가 못 하는 일이다.
        if self.applied:
            home = QPushButton("⟲ 원본으로")
            home.setToolTip("파일을 열었을 때의 계통으로 돌아갑니다.\n"
                            "바꾼 것은 시나리오 목록에 그대로 남습니다.")
            home.setCursor(Qt.PointingHandCursor)
            home.clicked.connect(self.reset_to_base)
            h.addWidget(home)
        return bar

    def reset_to_base(self):
        """파일을 열었을 때의 계통으로 돌아간다.

        「↩ 되돌리기」와 하는 일이 다르다 — 저건 **아직 계산 안 한 것**만 지운다.
        [이 조건으로 계산] 을 누르는 순간 그 조건은 굳고, 그때부터 되돌리기로는 못 푼다.
        """
        base = self.book.base() if self.book is not None else None
        # 🚨 `book.base()` 는 원본이 없으면 **첫 시나리오로 떨어진다**(`scenario.py:481`).
        #    그걸 그대로 쓰면 「원본으로」가 엉뚱한 조건으로 데려간다 ⇒ 바꾼 것이 없는지 직접 본다.
        if base is not None and base.solved and not base.changes:
            self.show_scenario(base)      # 이미 푼 답을 들고 있으므로 **다시 안 푼다**
            return
        # 원본 결과가 없다(파일을 열고 아직 한 번도 원본으로 안 풀었다) — 조건만 비운다.
        self.changes = []
        self.applied = []
        self.rebuild()

    # ── ② 부하 일괄 증감 ──────────────────────────────────────────────
    # 칸을 하나씩 고치는 대신 **부하 전체에 한 수를 곱한다**. 부하 여유(margin)를 보는
    # 가장 흔한 방법이라 슬라이더 하나로 둔다. 발전은 넣지 않았다 — 슬랙이 차액을 다 받아
    # 무엇 때문에 답이 달라졌는지 흐려진다.

    def load_factor(self, changes=None):
        """지금 걸려 있는 배수 (원본 대비). 곱하기가 여러 개면 다 곱한 값이다."""
        f = 1.0
        for ch in (self.applied + self.changes if changes is None else changes):
            if isinstance(ch, SC.Scale):
                f *= float(ch.factor)
        return f

    def load_total(self, changes):
        """지금 보고 있는 시각의 총 부하 [MW] — P 만 센다 (Q 는 따로 안 보여 준다)."""
        case = SC.apply(self.base_case, changes) if changes else self.base_case
        total = 0.0
        for key in ("AC_PLoad_dat", "DC_PLoad_dat"):
            arr = SC._values(case, key)
            if arr.size == 0 or arr.shape[1] < 2:
                continue
            col = min(self.t + 1, arr.shape[1] - 1)      # 0열은 버스 번호
            total += float(np.nansum(arr[:, col]))
        return total / 1e6

    def load_times(self):
        """부하표가 들고 있는 시각 수 (1이면 한 시각짜리 계통)."""
        n = 1
        for key in SC.LOAD_TABLES:
            arr = SC._values(self.base_case, key)
            if arr.size and arr.shape[1] > 1:
                n = max(n, arr.shape[1] - 1)          # 0열은 버스 번호
        return n

    def has_load(self):
        return self.base_case is not None and any(
            SC._values(self.base_case, k).size for k in SC.LOAD_TABLES)

    def load_bar(self, inline: bool = False):
        """부하 전체 ×배수 슬라이더. 여기서도 **바로 계산하지 않는다.**

        `inline` 이면 카드 테두리 없이 **표 고르기 줄 안에** 들어간다 — 따로 한 줄을
        쓰면 그만큼 표가 줄어든다(2026-08-15: 표에 19px 밖에 안 남아 있었다).
        """
        if not self.has_load():
            return None
        c = self.c
        now = self.load_factor()
        if inline:
            bar = QWidget()
            h = QHBoxLayout(bar)
            h.setContentsMargins(0, 0, 0, 0)
            h.setSpacing(8)
        else:
            bar = QFrame()
            bar.setObjectName("card")
            bar.setStyleSheet(f"#card {{ background:{c['surface']};"
                              f"border:1px solid {c['border']};border-radius:10px; }}")
            h = QHBoxLayout(bar)
            h.setContentsMargins(15, 7, 11, 7)
            h.setSpacing(11)

        tag = QLabel("부하 전체")
        tag.setStyleSheet(f"color:{c['text']};font-size:13px;font-weight:600;")
        h.addWidget(tag)

        sl = QSlider(Qt.Horizontal)
        sl.setRange(50, 200)
        sl.setValue(int(round(now * 100)))
        sl.setFixedWidth(150 if inline else 240)
        sl.setTickPosition(QSlider.TicksBelow)
        sl.setTickInterval(25)
        n_t = self.load_times()
        sl.setToolTip("모든 부하에 같은 수를 곱합니다."
                      + (f"\n이 계통은 {n_t}시각짜리이고, **모든 시각에 함께** 걸립니다."
                         if n_t > 1 else ""))
        h.addWidget(sl)

        val = QLabel(f"×{now:.2f}")
        val.setFixedWidth(52)
        val.setStyleSheet(f"color:{c['accent']};font-size:14px;font-weight:700;")
        h.addWidget(val)

        # 🚨 곱하기는 **모든 시각**에 걸리는데 여기 뜨는 합계는 **보고 있는 시각 하나**다.
        #    시간을 바꾸면 이 숫자도 따라 움직여서 "이 시각에만 걸리나?" 로 읽힌다
        #    (2026-08-06 사용자 질문) ⇒ 여러 시각짜리면 어느 시각인지 라벨에 밝힌다.
        base_mw = self.load_total([])
        now_mw = self.load_total(self.applied + self.changes)
        when = f" ({min(self.t, n_t - 1) + 1} H)" if n_t > 1 else ""
        tot = QLabel(f"총 부하 {base_mw:,.1f} → {now_mw:,.1f} MW{when}"
                     if abs(now - 1.0) > 1e-9 else f"총 부하 {base_mw:,.1f} MW{when}")
        tot.setStyleSheet(f"color:{c['muted']};font-size:12px;")
        if n_t > 1:
            tot.setToolTip(f"지금 보고 있는 시각의 합입니다. "
                           f"곱하기는 {n_t}시각 전부에 걸립니다.")
        h.addWidget(tot)
        if not inline:
            h.addStretch(1)

        back = QPushButton("원래대로 (×1)")
        back.setCursor(Qt.PointingHandCursor)
        back.setEnabled(abs(now - 1.0) > 1e-9)
        back.clicked.connect(lambda: self.scale_loads(1.0))
        h.addWidget(back)

        def moved(v):
            val.setText(f"×{v / 100:.2f}")
            self._load_pending = v / 100
            if not sl.isSliderDown():            # 화살표키·홈 클릭 — 잠깐 뒤에 반영
                self._load_timer.start(250)

        sl.valueChanged.connect(moved)
        sl.sliderReleased.connect(lambda: self.scale_loads(sl.value() / 100))
        return bar

    def scale_loads(self, want):
        """부하를 원본의 want 배로 만든다. 이미 푼 조건에 걸린 배수는 빼고 얹는다."""
        self._load_timer.stop()
        if self.base_case is None:
            return
        done = self.load_factor(self.applied)          # 이미 계산해 놓은 몫
        self.changes = [ch for ch in self.changes if not isinstance(ch, SC.Scale)]
        need = float(want) / done
        if abs(need - 1.0) > 1e-9:
            self.changes.append(SC.Scale(tables=SC.LOAD_TABLES, factor=need,
                                         label=f"부하 전체 ×{want:g}"))
        self.rebuild()

    def grid_page(self):
        """계통 데이터 탭 — 엑셀 값을 그대로 보여 주고, 여기서 켜고 끈다."""
        c = self.c
        w = QWidget()
        v = QVBoxLayout(w)
        # 여백·간격을 줄인다 — 표에 19px 밖에 안 남던 시절의 값이었다(2026-08-15).
        v.setContentsMargins(12, 6, 12, 8)
        v.setSpacing(6)

        if self.base_case is None:
            note = QLabel("케이스를 열면 여기에 계통 데이터가 나옵니다.")
            note.setStyleSheet(f"color:{c['muted']};font-size:14px;")
            v.addWidget(note)
            v.addStretch(1)
            return w

        # 어느 표를 볼까 — 그 계통에 실제로 있는 것만 (설비가 없으면 칸도 안 만든다)
        picks = []
        for key, label in GRID_TABLES:
            arr = SC._values(self.base_case, key)
            if arr.size and arr.ndim == 2 and not np.all(np.isnan(arr)):
                picks.append((key, label, arr.shape[0]))
        if not picks:
            v.addWidget(QLabel("보여 줄 표가 없습니다."))
            return w
        if self.grid_key not in [k for k, _, _ in picks]:
            self.grid_key = picks[0][0]

        row = QHBoxLayout()
        row.setSpacing(6)
        for key, label, n in picks:
            b = QPushButton(f"{label} {n}")
            b.setObjectName("seg_on" if key == self.grid_key else "seg_off")
            b.setCursor(Qt.PointingHandCursor)
            b.clicked.connect(lambda _, k=key: self.set_grid_table(k))
            row.addWidget(b)
        # 찾기 칸을 **같은 줄에** 붙인다. 따로 한 줄을 쓰면 바가 셋이 되어
        # (표 고르기·부하·찾기) 표에 줄이 한 줄도 안 남는다(2026-08-13 실측).
        row.addSpacing(10)
        row.addWidget(self.find_bar(inline=True))
        row.addStretch(1)
        # 안내는 **글줄로 두지 않고** 표 고르기 단추의 설명으로 옮겼다 — 그 자리에
        # 부하 슬라이더를 들여야 한 줄이 준다.
        for i in range(row.count()):
            wd = row.itemAt(i).widget()
            if isinstance(wd, QPushButton):
                wd.setToolTip("켜고 끄기는 바로 계산하지 않습니다 — "
                              "다 바꾼 뒤 위의 [이 조건으로 계산] 을 누르세요.")
        load = self.load_bar(inline=True)       # ② 부하 일괄 증감 — **같은 줄에**
        if load is not None:
            row.addWidget(load)
        v.addLayout(row)

        v.addWidget(self.grid_table_widget(), 1)
        return w

    # ── 버스 번호로 찾기 ──────────────────────────────────────────────
    # 🚨 큰 계통에서는 표를 다 그리는 것 자체가 비싸다 — 6,495버스 케이스의 AC 선로 표는
    #    9,019줄이고 그리는 데 1.7초다(2026-08-06 실측). 그런데 그때 하려는 일은
    #    "9천 줄 훑어보기" 가 아니라 "그 선로 하나 끊어 보기" 다. 번호로 좁힌다.

    def find_rows(self, key, arr):
        """찾는 번호에 걸리는 줄 번호들. 찾는 게 없으면 None (= 전부)."""
        want = {int(x) for x in re.findall(r"\d+", self.grid_find or "")}
        if not want:
            return None
        sw = SC.SWITCHES.get(key)
        # 그 표에서 **버스 번호가 든 열** — 스위치가 아는 것이 있으면 그것을, 없으면 0열
        idc = [ci for ci in (sw.ident if sw else (0,)) if ci < arr.shape[1]]
        out = []
        for r in range(arr.shape[0]):
            for ci in idc:
                v = arr[r, ci]
                if np.isfinite(v) and int(v) in want:
                    out.append(r)
                    break
        return out

    def find_bar(self, inline=False):
        """버스 번호로 찾기. `inline` 이면 표 고르기 줄에 얹으므로 뒤 여백을 안 둔다."""
        c = self.c
        bar = QFrame()
        h = QHBoxLayout(bar)
        h.setContentsMargins(0, 0, 0, 0)
        h.setSpacing(8)

        lab = QLabel("버스 번호로 찾기")
        lab.setStyleSheet(f"color:{c['muted']};font-size:12px;")
        h.addWidget(lab)

        box = QLineEdit(self.grid_find or "")
        box.setPlaceholderText("예: 38 39 40 (비우면 전부)" if inline
                              else "예: 38   ·   38 39 40   (비우면 전부)")
        box.setFixedWidth(200 if inline else 260)
        box.returnPressed.connect(lambda: self.set_find(box.text()))
        box.editingFinished.connect(lambda: self.set_find(box.text()))
        h.addWidget(box)

        arr = SC._values(self.base_case, self.grid_key)
        rows = self.find_rows(self.grid_key, arr)
        if rows is not None:
            got = QLabel(f"{arr.shape[0]:,}줄 중 **{len(rows):,}줄**"
                         .replace("**", ""))
            got.setStyleSheet(f"color:{c['accent']};font-size:12px;font-weight:600;")
            h.addWidget(got)
            clr = QPushButton("전부 보기")
            clr.setCursor(Qt.PointingHandCursor)
            clr.clicked.connect(lambda: self.set_find(""))
            h.addWidget(clr)
        else:
            got = QLabel(f"{arr.shape[0]:,}줄")
            got.setStyleSheet(f"color:{c['muted']};font-size:12px;")
            h.addWidget(got)
        if not inline:
            h.addStretch(1)
        return bar

    def set_find(self, text):
        text = (text or "").strip()
        if text == (self.grid_find or ""):
            return
        self.grid_find = text
        self.rebuild()

    def grid_table_widget(self):
        """지금 고른 표 하나를 그린다.

        · 켤 수 있는 표면 첫 칸이 스위치
        · **화면 단위로 바꿔서** 보여 준다 (엔진은 W, 화면은 MW — 머리글이 [MW] 니까)
        · ③ 운전 조건 칸만 고칠 수 있고, 나머지는 회색이다
        """
        c = self.c
        key = self.grid_key
        sw = SC.SWITCHES.get(key)
        heads = GRID_HEADERS.get(key, [])
        scales = GRID_SCALES.get(key, {})
        editable = GRID_EDITABLE.get(key, set())
        eff = self.applied + self.changes      # 화면에 보이는 조건 = 푼 것 + 얹은 것
        arr = SC._values(SC.apply(self.base_case, eff), key)
        if key in GRID_PAD_TO_HEADERS and heads and arr.shape[1] < len(heads):
            # 파일에 없는 열은 빈 칸으로 보여 준다 — 값을 넣는 순간 표가 늘어난다
            arr = np.hstack([arr, np.full((arr.shape[0], len(heads) - arr.shape[1]),
                                          np.nan)])
        ncol = min(arr.shape[1], len(heads)) if heads else arr.shape[1]
        head_of = (lambda i: heads[i] if i < len(heads) else f"{i + 1}열")

        # 🚨 찾기로 좁히면 **보이는 줄 번호와 진짜 줄 번호가 달라진다.**
        #    켜고 끄기·값 고치기는 진짜 번호로 해야 하므로 그 대응을 들고 있는다.
        picked = self.find_rows(key, arr)
        self._grid_rows = list(range(arr.shape[0])) if picked is None else picked
        nrow = len(self._grid_rows)

        touched = {(ch.row, ch.col) for ch in self.changes
                   if isinstance(ch, SC.Cell) and ch.table == key}

        def cell(r, j):
            """진짜 줄 r · 데이터 열 j 의 칸 하나."""
            # 🚨 위상 조정기(Ctrl Mode = 2)에는 `Ctrl Bus` 가 없다 — 그 선로 **자신의**
            #    조류를 보기 때문이다. 빈 칸으로 두면 "여기 뭘 넣어야 하나" 로 읽히므로
            #    잠그고 「—」로 보여 준다(2026-08-13 사용자: "비워 둠 이거를 해결해야").
            #    엔진도 모드 2 에서는 이 칸을 안 읽는다.
            if key == "AC_Line_dat" and j == 14 and ncol > 13 and arr[r, 13] == 2:
                it = QTableWidgetItem("—")
                it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                it.setFlags(it.flags() & ~Qt.ItemIsEditable)
                it.setForeground(QColor(c["muted"]))
                it.setToolTip("위상 조정기는 그 선로 자신의 조류를 봅니다 — "
                              "볼 버스가 따로 없습니다")
                return it
            val = arr[r, j] * scales.get(j, 1.0)
            txt = "" if np.isnan(val) else (
                f"{val:.0f}" if float(val).is_integer() and abs(val) < 1e9
                else f"{val:,.4f}".rstrip("0").rstrip("."))
            it = QTableWidgetItem(txt)
            if j > 0:
                it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            if j in editable:
                it.setToolTip("고칠 수 있는 값입니다 — 바꾸면 계산은 안 돌고 "
                              "위의 [이 조건으로 계산] 을 눌러야 풉니다")
            else:
                it.setFlags(it.flags() & ~Qt.ItemIsEditable)
                it.setForeground(QColor(c["muted"]))   # 여기부터는 엑셀에서
            if (r, j) in touched:
                it.setForeground(QColor(c["warn"]))
            return it

        def new_table(labels):
            t = QTableWidget(nrow, len(labels))
            t.setHorizontalHeaderLabels(labels)
            t.verticalHeader().setVisible(False)
            t.verticalHeader().setDefaultSectionSize(30)
            t.setAlternatingRowColors(True)
            return t

        # ── 왼쪽(고정) / 오른쪽(미는 것) ────────────────────────────────
        # 오른쪽으로 밀면 **어느 줄인지 알 수 없다** — Ctrl 열은 19열 중 14~19열이라
        # 선로 번호가 화면 밖으로 나간다(2026-08-13 사용자: "첫 열을 고정시켜줘").
        # 그래서 상태 + 첫 번호 열을 **따로 만든 표**로 왼쪽에 세운다.
        # 🚨 데이터 열 0 은 어느 표에서도 고칠 수 있는 열이 아니다(GRID_EDITABLE 확인) —
        #    그래서 왼쪽 표에는 itemChanged 를 안 잇는다.
        freeze = ncol >= 2
        self._grid_loading = True              # 그리는 동안의 itemChanged 는 무시

        left = None
        if freeze:
            left = new_table((["상태"] if sw else []) + [head_of(0)])
            left.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
            left.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
            left.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
            if sw:
                # 🚨 순서가 중요하다 — `ResizeToContents` 인 채로 폭을 정하면 **무시된다.**
                #    상태 칸에는 글자가 아니라 단추가 들어서 자동 폭이 22px 로 눌린다
                #    (2026-08-13 렌더로 확인: 머리글이 "태", 단추 글자 안 보임).
                left.horizontalHeader().setSectionResizeMode(0, QHeaderView.Fixed)
                left.setColumnWidth(0, 66)
            for at, r in enumerate(self._grid_rows):
                left.setItem(at, 1 if sw else 0, cell(r, 0))

        body = [j for j in range(ncol)] if not freeze else list(range(1, ncol))
        off = (1 if sw else 0) if not freeze else -1   # 화면 열 = 데이터 열 + off
        labels = ((["상태"] if sw else []) + [head_of(j) for j in body]) if not freeze \
            else [head_of(j) for j in body]
        tb = new_table(labels)
        # 🚨 열이 많으면 늘려 맞추기(Stretch)가 머리글을 잘라 버린다 — IC 는 20열이라
        #    "Rating Power [MW]" 가 "ng Po" 로 보였다. 그럴 땐 글자에 맞추고 옆으로 넘긴다.
        tb.horizontalHeader().setSectionResizeMode(
            QHeaderView.Stretch if len(labels) <= 9 else QHeaderView.ResizeToContents)
        for at, r in enumerate(self._grid_rows):     # at = 화면 줄, r = 진짜 줄
            if sw:
                on = SC.is_on(self.base_case, key, r, eff)
                b = QPushButton("켜짐" if on else "꺼짐")
                b.setObjectName("seg_on" if on else "seg_off")
                b.setCursor(Qt.PointingHandCursor)
                b.clicked.connect(lambda _, rr=r: self.flip_row(rr))
                (left if freeze else tb).setCellWidget(at, 0, b)
            for j in body:
                tb.setItem(at, j + off, cell(r, j))
        self._grid_loading = False
        tb.itemChanged.connect(lambda item: self.grid_edited(key, item, off, scales))

        self._grid_tb = tb          # 스크롤 자리·다음 칸은 **미는 쪽** 기준
        self._grid_off = off        # 시험이 화면 열을 계산할 때 쓴다
        self._grid_frozen = left
        QTimer.singleShot(0, lambda: self._restore_grid_view(tb))
        if not freeze:
            return tb

        # 세로로 함께 움직인다. 줄 높이가 둘 다 30 이고 줄 단위로 구르므로 값이 같다.
        tb.verticalScrollBar().valueChanged.connect(left.verticalScrollBar().setValue)
        left.verticalScrollBar().valueChanged.connect(tb.verticalScrollBar().setValue)

        box = QWidget()
        h = QHBoxLayout(box)
        h.setContentsMargins(0, 0, 0, 0)
        h.setSpacing(0)
        lwrap = QWidget()
        lv = QVBoxLayout(lwrap)
        lv.setContentsMargins(0, 0, 0, 0)
        lv.setSpacing(0)
        lv.addWidget(left, 1)
        # 오른쪽에 가로 막대가 뜨면 그만큼 보이는 높이가 줄어든다. 왼쪽에도 같은
        # 높이를 비워 둬야 **마지막 줄이 어긋나지 않는다.**
        pad = QWidget()
        pad.setFixedHeight(0)
        lv.addWidget(pad)
        h.addWidget(lwrap)
        h.addWidget(tb, 1)

        def fit():
            try:
                left.setFixedWidth(left.horizontalHeader().length() + 2)
                hb = tb.horizontalScrollBar()
                pad.setFixedHeight(hb.sizeHint().height() if hb.isVisible() else 0)
            except RuntimeError:
                pass

        tb.horizontalScrollBar().rangeChanged.connect(lambda *_: fit())
        QTimer.singleShot(0, fit)
        return box

    def grid_edited(self, key, item, off, scales):
        """운전 조건 칸을 고쳤다. **계산은 안 한다** — 바꾼 목록에만 얹는다."""
        if getattr(self, "_grid_loading", False):
            return
        col = item.column() - off
        if col < 0 or col not in GRID_EDITABLE.get(key, set()):
            return
        txt = item.text().strip().replace(",", "")
        if txt == "":
            # 지워서 비운 것 — 「안 적음」이다. 비워도 되는 칸에서만 받는다.
            if col not in GRID_CLEARABLE.get(key, set()):
                QMessageBox.information(
                    self, "비울 수 없는 칸입니다",
                    "이 칸은 값이 있어야 합니다. 숫자를 넣어 주세요.\n"
                    "(비울 수 있는 것은 조정 칸입니다 — 비우면 「안 걸었다」는 뜻이 됩니다.)")
                self.rebuild()
                return
            shown = value = float("nan")
        else:
            try:
                shown = float(txt)
            except ValueError:
                QMessageBox.information(self, "숫자를 넣어 주세요",
                                        f"'{item.text()}' 는 숫자가 아닙니다.")
                self.rebuild()
                return
            scale = scales.get(col, 1.0)
            value = shown * (1.0 / scale) if scale != 1.0 else shown
        # 찾기로 좁혀 놓았으면 화면 줄 ≠ 진짜 줄
        seen = getattr(self, "_grid_rows", None)
        row = seen[item.row()] if seen and item.row() < len(seen) else item.row()
        cur = SC._values(SC.apply(self.base_case, self.applied + self.changes), key)
        before = float(cur[row, col]) if col < cur.shape[1] else float("nan")
        if np.isnan(before) and np.isnan(value):
            return                              # 비어 있던 것을 또 비웠다 — 아무 일도 아니다
        if not np.isnan(before) and abs(before - value) <= abs(before) * 1e-12:
            return                              # 안 바뀐 값 — 목록을 더럽히지 않는다
        head = GRID_HEADERS.get(key, [])
        name = head[col] if col < len(head) else f"{col + 1}열"
        self.changes = [ch for ch in self.changes
                        if not (isinstance(ch, SC.Cell) and ch.table == key
                                and ch.row == row and ch.col == col)]
        self.changes.append(SC.Cell(
            table=key, row=row, col=col, value=value,
            label=(f"{SC.describe_row(self.base_case, key, row)} {name} → "
                   + ("(비움)" if np.isnan(shown) else f"{shown:g}")),
            mark=SC.row_mark(self.base_case, key, row)))
        # 같은 줄의 **다음 고칠 수 있는 칸**을 미리 골라 둔다. 조정 열은 여섯 칸을
        # 잇달아 채우는 일이라(Mode·Bus·Target·Min·Max·Steps) 한 칸 칠 때마다
        # 오른쪽으로 다시 찾아가야 했다. 이제 바로 다음 칸에 숫자를 치면 된다.
        nxt = [x for x in sorted(GRID_EDITABLE.get(key, set())) if x > col]
        # 위상 조정기(Ctrl Mode = 2)는 `Ctrl Bus` 를 안 쓴다 — 그 선로 자신의 조류를
        # 보기 때문이다. 방금 2 를 쳤으면 그 칸을 건너뛰고 `Ctrl Target` 으로 간다.
        if key == "AC_Line_dat" and col == 13 and value == 2:
            nxt = [x for x in nxt if x != 14]
        self._grid_focus = (item.row(), (nxt[0] if nxt else col) + off)
        self.rebuild()

    def set_grid_table(self, key):
        self.grid_key = key
        self.rebuild()

    def flip_row(self, row):
        """그 줄을 켜거나 끈다. **계산은 안 한다** — 목록에만 얹는다."""
        key = self.grid_key
        eff = self.applied + self.changes
        now = SC.is_on(self.base_case, key, row, eff)
        # 같은 줄에 대한 옛 기록은 지운다 (껐다 켜면 아무것도 안 바꾼 것이 되게)
        self.changes = [ch for ch in self.changes
                        if not (isinstance(ch, SC.Cell) and ch.table == key
                                and ch.row == row)]
        try:
            ch = SC.toggle(self.base_case, key, row, on=not now)
        except SC.NotSupported as exc:
            QMessageBox.information(self, "못 바꿉니다", str(exc))
            return
        # **이미 푼 조건**과 같아지면 얹을 것이 없다 (껐다 켜면 아무것도 안 바꾼 것)
        if SC.is_on(self.base_case, key, row, self.applied) != (not now):
            self.changes.append(ch)
        self.rebuild()

    def undo_changes(self):
        """아직 안 푼 것만 물린다. 이미 계산해서 보고 있는 조건은 그대로 둔다."""
        self.changes = []
        self.rebuild()

    def run_changes(self):
        """[이 조건으로 계산] — 여기서만 푼다."""
        if not self.changes or self.base_case is None:
            return
        self._pending = self.applied + self.changes   # 풀고 나서 시나리오로 담으려고
        # 이름은 **이번에 새로 얹은 것**으로 짓는다. 전체 목록으로 지으면 맨 앞의
        # 오래된 조건("부하 전체 ×1.3")이 계속 이름을 차지한다.
        self._pending_new = list(self.changes)
        self._start_solve(getattr(self, "_last_path", self.base_case.case_name),
                          case=SC.apply(self.base_case, self._pending))

    def overlay_pairs(self):
        """겹쳐 그릴 (이름, 결과) 목록. 체크한 것 중 **풀린 것만**."""
        out = []
        for i, s in enumerate(self.book.items):
            if self.overlay and i not in self.overlay:
                continue
            if s.solved:
                out.append((s.name, s.solution))
        return out

    def toggle_overlay(self, i, on):
        if not self.overlay:                       # 비어 있으면 '전부' 라는 뜻
            self.overlay = set(range(len(self.book.items)))
        self.overlay = (self.overlay | {i}) if on else (self.overlay - {i})
        self.rebuild()

    def compare_scenarios_area(self, picked):
        """시나리오끼리 비교 — 목록에서 체크한 것을 한 그래프에 겹쳐 그린다."""
        c = self.c
        tabs = QTabWidget()
        pairs = self.overlay_pairs()
        if len(self.book.items) < 2:
            page = QWidget(); pv = QVBoxLayout(page)
            lb = QLabel("아직 담아 둔 시나리오가 없습니다.\n"
                        "계통 데이터 탭에서 조건을 바꾸고 [이 조건으로 계산] 을 누르면 쌓입니다.")
            lb.setAlignment(Qt.AlignCenter)
            lb.setStyleSheet(f"color:{c['muted']};font-size:15px;")
            pv.addWidget(lb)
            tabs.addTab(page, "결과")
            return tabs
        for name in picked:
            page = QWidget()
            pv = QVBoxLayout(page)
            pv.setContentsMargins(10, 10, 10, 10)
            pv.setSpacing(9)
            pv.addWidget(charts.compare_scenarios(c, pairs, name, self.t), 2)
            tb = self.scenario_table(name, pairs)
            if tb is not None:
                pv.addWidget(tb, 1)
            tabs.addTab(page, name)
        return tabs

    def scenario_table(self, item, pairs):
        """겹쳐 그린 시나리오의 요약 표 — 최저·최고·원본 대비."""
        if item in ("주파수", "손실") or not pairs:
            return None
        col = {"전압 크기": "VM[pu]", "위상각": "Angle[deg]"}.get(item)
        if col is None:
            return None
        rows = []
        base_lo = None
        for name, sol in pairs:
            vals = []
            for kind in ("AC", "DC"):
                arr = sol.at(kind, self.t)
                cols = sol.cols(kind)
                if arr.size and col in cols:
                    vals.append(np.asarray(arr[:, cols.index(col)], dtype=float))
            if not vals:
                continue
            y = np.concatenate(vals)
            lo, hi = float(np.nanmin(y)), float(np.nanmax(y))
            if base_lo is None:
                base_lo = lo
            rows.append((name, lo, hi, lo - base_lo))
        if not rows:
            return None
        heads = ["시나리오", "최저", "최고", "첫 줄 대비"]
        tb = QTableWidget(len(rows), len(heads))
        tb.setHorizontalHeaderLabels(heads)
        tb.verticalHeader().setVisible(False)
        tb.verticalHeader().setDefaultSectionSize(28)
        tb.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        tb.setAlternatingRowColors(True)
        for r, (name, lo, hi, d) in enumerate(rows):
            for cc, txt in enumerate([name, f"{lo:.4f}", f"{hi:.4f}",
                                      "—" if r == 0 else f"{d:+.4f}"]):
                it = QTableWidgetItem(txt)
                if cc:
                    it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                if cc == 3 and r and d < -1e-6:
                    it.setForeground(QColor(self.c["warn"]))
                tb.setItem(r, cc, it)
        return tb

    def scenario_bar(self):
        """담아 둔 시나리오 목록. 원본뿐이면 안 만든다(줄 하나짜리 목록은 군더더기다).

        결과를 이미 들고 있으므로 **다시 계산하지 않고** 오갈 수 있다.
        칸이 줄마다 어긋나지 않게 **격자**로 놓는다.
        """
        if len(self.book.items) < 2:
            return None
        c = self.c
        card = QFrame()
        card.setObjectName("card")
        v = QVBoxLayout(card)
        v.setContentsMargins(14, 9, 12, 10)
        v.setSpacing(6)

        head = QHBoxLayout()
        cap = QLabel(f"시나리오 {len(self.book.items)}")
        cap.setStyleSheet(f"color:{c['muted']};font-size:12px;font-weight:600;")
        head.addWidget(cap)
        head.addStretch(1)
        tip = QLabel("이름을 두 번 누르면 고칠 수 있습니다")
        tip.setStyleSheet(f"color:{c['muted']};font-size:11px;")
        head.addWidget(tip)
        v.addLayout(head)

        g = QGridLayout()
        g.setHorizontalSpacing(12)
        g.setVerticalSpacing(3)
        for j, (name, w) in enumerate(
                [("겹쳐", 46), ("이름", 0), ("", 0), ("", 0),
                 ("전압 최저", 82), ("원본 대비", 74), ("", 34)]):
            if name:
                q = QLabel(name)
                q.setStyleSheet(f"color:{c['muted']};font-size:11px;")
                g.addWidget(q, 0, j)
            if w:
                g.setColumnMinimumWidth(j, w)
        g.setColumnStretch(1, 1)

        base = self.book.base()
        for i, s in enumerate(self.book.items, start=1):
            self._scenario_row(g, i, s, base)

        # 🚨 **몇 줄까지만 보이고 넘으면 목록 안에서 스크롤한다** (2026-08-15 사용자 확정).
        #    그전에는 시나리오 한 줄에 **25px 씩 카드가 계속 자랐다** — 7개면 235px 이고,
        #    그만큼 그래프와 표가 줄었다. 6개부터는 **둘 다 바닥에 닿아** 표가 통째로 잘렸다
        #    (실측: 카드 0→110→160→235 / 그래프 523→408→353→320 / 표 275→215→186→170).
        #    ⇒ 상한을 씌워 **쌓일수록 나빠지는 것 자체를 없앤다.**
        #    시나리오는 보통 최근 것만 보므로 네 줄이면 실제로 쓸 만하다.
        v.addLayout(g)
        v.addStretch(1)
        return card

    def _scenario_row(self, g, i, s, base):
        c = self.c
        here = list(s.changes) == list(self.applied)     # 지금 화면이 이것인가

        left0 = QHBoxLayout()
        left0.setSpacing(4)
        dot = QLabel("●" if here else "○")
        dot.setStyleSheet(f"color:{c['accent'] if here else c['border']};font-size:12px;")
        left0.addWidget(dot)
        if s.solved:
            cb = QCheckBox()
            cb.setChecked((i - 1) in self.overlay if self.overlay else True)
            cb.setToolTip("비교 모드에서 겹쳐 그릴지")
            cb.toggled.connect(lambda on, k=i - 1: self.toggle_overlay(k, on))
            left0.addWidget(cb)
        left0.addStretch(1)
        g.addLayout(left0, i, 0)

        left = QHBoxLayout()
        left.setSpacing(7)
        name = _ClickLabel(s.name)
        name.setStyleSheet(
            f"color:{c['accent'] if here else c['text']};font-size:13px;"
            f"font-weight:{'700' if here else '500'};")
        name.setToolTip(SC.describe(s.changes))
        name.double_clicked.connect(lambda _s=s: self.rename_scenario(_s))
        left.addWidget(name)
        if here:
            tag = QLabel("지금 보는 것")
            tag.setStyleSheet(
                f"color:{c['accent']};background:{c['accent_soft']};font-size:11px;"
                f"border-radius:8px;padding:1px 8px;")
            left.addWidget(tag)
        left.addStretch(1)
        g.addLayout(left, i, 1)

        def cell(col, text, color=None, bold=False, right=False, tip=""):
            q = QLabel(text)
            q.setStyleSheet(f"color:{color or c['muted']};font-size:12px;"
                            f"font-weight:{'600' if bold else '400'};")
            if right:
                q.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            if tip:
                q.setToolTip(tip)
            g.addWidget(q, i, col)

        # 🚨 「바꾼 것 N건」과 「반복 N회」는 뺐다 (2026-08-15) — 앞엣것은 이름이 이미
        #    말해 주고(`AC 선로 2-4 끔`), 뒤엣것은 시나리오를 고르는 데 안 쓰인다
        #    (지금 보는 것의 반복 횟수는 아래 상태줄에 있다). 대신 **못 푼 것**은
        #    반드시 보여야 하므로 그때만 자리를 쓴다.
        if s.error:
            cell(2, s.summary, c["warn"], bold=True)
        vmin = s.vmin()
        cell(4, "—" if np.isnan(vmin) else f"{vmin:.4f} pu", c["text"], right=True)
        d = s.against(base) if (base is not None and not s.base) else float("nan")
        cell(5, "—" if np.isnan(d) else f"{d:+.4f}",
             c["warn"] if (not np.isnan(d) and d < -1e-6) else c["muted"],
             right=True, tip="원본 대비 전압 최저 변화")

        # 🚨 [결과 보기] 단추를 뺐다 — **이름을 누르면 간다.** 단추(34px)가 줄 높이를
        #    42px 로 만들고 있었고, 목록이 그만큼 자리를 먹었다(4줄 248px).
        #    [지우기] 는 작은 ✕ 로 줄인다.
        if not here:
            name.setCursor(Qt.PointingHandCursor)
            name.clicked.connect(lambda _s=s: self.show_scenario(_s))
            name.setToolTip(SC.describe(s.changes) +
                            ("\n\n눌러서 이 결과를 봅니다" if s.solved
                             else "\n\n눌러서 이 조건만 불러옵니다"))
        act = QHBoxLayout()
        act.setSpacing(6)
        act.addStretch(1)
        if not s.base:
            x = QPushButton("✕")
            x.setFixedSize(22, 22)
            x.setCursor(Qt.PointingHandCursor)
            x.setToolTip("이 시나리오를 목록에서 지웁니다 (계통은 안 건드립니다)")
            x.clicked.connect(lambda _, _s=s: self.drop_scenario(_s))
            act.addWidget(x)
        g.addLayout(act, i, 6)

    def show_scenario(self, s):
        """그 시나리오로 화면을 옮긴다. 결과를 들고 있으면 **다시 안 푼다.**"""
        self.applied = list(s.changes)
        self.changes = []
        if s.solved:
            self.sol = s.solution
            self.t = min(self.t, max(int(s.solution.n_time) - 1, 0))
        else:
            # 안 풀린 시나리오 — 조건만 깔아 준다. 화면 결과는 건드리지 않는다.
            QMessageBox.information(
                self, "안 풀린 시나리오",
                f"「{s.name}」 은 답을 못 찾은 조건입니다.\n조건만 깔아 두었으니 "
                f"여기서 더 바꿔 다시 계산해 보세요.\n\n화면의 결과는 그대로 둡니다.")
        self.rebuild()

    def rename_scenario(self, s):
        name, ok = QInputDialog.getText(self, "이름 바꾸기", "시나리오 이름", text=s.name)
        if ok and name.strip():
            self.book.rename(s, name.strip())
            self.rebuild()

    def drop_scenario(self, s):
        if list(s.changes) == list(self.applied):
            back = self.book.base()
            if back is not None and back.solved:
                self.show_scenario(back)          # 보고 있던 것을 지우면 원본으로
        self.book.remove(s)
        self.rebuild()

    def check_page(self):
        c = self.c
        w = QWidget()
        outer = QVBoxLayout(w)
        outer.setContentsMargins(14, 12, 14, 12)
        outer.setSpacing(11)

        n = violation_count(self.viol())
        head = QFrame()
        head.setObjectName("card")
        hv = QHBoxLayout(head)
        hv.setContentsMargins(14, 10, 14, 10)
        dot = QLabel("●")
        dot.setStyleSheet(
            f"color:{c['warn'] if n else c['ok']};font-size:15px;")
        hv.addWidget(dot)
        msg = QLabel(f"한계를 벗어난 항목 {n}건" if n else "한계를 벗어난 항목 없음")
        msg.setStyleSheet(
            f"color:{c['warn'] if n else c['ok']};font-size:15px;font-weight:700;")
        hv.addWidget(msg)
        hv.addStretch()
        sub = QLabel("전압 한계 · 선로 용량 · 변환기 한계 · 발전기 한계를 계산 결과에서 걸러낸 것입니다")
        sub.setStyleSheet(f"color:{c['muted']};font-size:12px;")
        hv.addWidget(sub)
        outer.addWidget(head)

        # ── 탭 자동 조정 결과 (2026-08-13, §7 5단계 A1) ──────────────────
        # 탭은 **계산이 정한 값**이라 사용자가 볼 데가 없으면 결과를 못 읽는다.
        # 특히 한계에 걸려 목표를 못 맞춘 경우를 모르면 "목표대로 됐겠지" 하고 넘어간다.
        tap = getattr(self.sol, "tap_ctrl", None) if self.sol else None
        if tap is not None and len(tap):
            tap_a = np.asarray(tap, dtype=float)
            miss = int((tap_a[:, 4] == 0).sum())
            # ── 계단 (2026-08-14, §7 5단계 ③) ────────────────────────────
            # 🚨 계단이면 **목표를 정확히 못 맞추는 것이 정상**이다(설 자리가
            #    정해져 있으니까). 그래서 「목표 맞춤」이라 쓰면 거짓말이 되고,
            #    반대로 경고를 달면 멀쩡한 것을 결함처럼 보이게 한다.
            #    ⇒ 11열을 보고 **계단은 계단이라고** 쓴다.
            #    2 = 계단 자리로 옮긴 뒤 다시 푸는 데 실패해 연속값이 그대로다.
            #        이것만은 경고다 — 사용자는 계단인 줄 알고 있을 테니까.
            nofit = int((tap_a[:, 10] == 2).sum()) if tap_a.shape[1] > 10 else 0
            warn_on = bool(miss or nofit)
            # 한계를 안 적어 앱이 0.9~1.1 로 잡은 것 (2026-08-13 사용자 확정).
            # **내가 정하지 않은 값이 답을 가두고 있으므로** 반드시 밝힌다.
            auto = int((tap_a[:, 7] == 1).sum()) if tap_a.shape[1] > 7 else 0
            card = QFrame()
            card.setObjectName("card")
            if warn_on:
                wc2 = QColor(c["warn"])
                card.setStyleSheet(
                    f"#card{{background:rgba({wc2.red()},{wc2.green()},{wc2.blue()},0.12);"
                    f"border:1px solid {c['warn']};}}")
            cv = QVBoxLayout(card)
            cv.setContentsMargins(14, 10, 14, 12)
            cv.setSpacing(8)

            hh = QHBoxLayout()
            ic = QLabel("⚠" if warn_on else "⚙")
            ic.setStyleSheet(
                f"color:{c['warn'] if warn_on else c['muted']};font-size:15px;")
            hh.addWidget(ic)
            kind_col = tap_a[:, 8] if tap_a.shape[1] > 8 else np.ones(len(tap_a))
            names = [n for k, n in ((1, "탭"), (2, "위상"), (3, "SVC"))
                     if (kind_col == k).any()]
            what = "·".join(names) + " 자동 조정"
            ttl = QLabel(f"{what} {len(tap)}대"
                         + (f" — {miss}대가 목표를 못 맞췄습니다" if miss else "")
                         + (f" — {nofit}대는 계단으로 못 내렸습니다" if nofit else ""))
            ttl.setStyleSheet(
                f"color:{c['warn'] if warn_on else c['text']};"
                f"font-size:14px;font-weight:700;")
            hh.addWidget(ttl)
            hh.addStretch()
            note = QLabel("굵은 값은 계산이 정한 것입니다"
                          + (" · 한계에 걸리면 목표를 포기합니다" if miss else "")
                          + (" · 계단은 정해진 자리에만 서므로 목표에 근접합니다"
                             if (np.asarray(tap, dtype=float).shape[1] > 9
                                 and (np.asarray(tap, dtype=float)[:, 9] > 0).any())
                             else ""))
            note.setStyleSheet(f"color:{c['muted']};font-size:12px;")
            hh.addWidget(note)
            cv.addLayout(hh)

            heads = ["방식", "선로", "맞추는 곳", "목표", "정해진 값",
                     "움직일 수 있는 범위", "결과"]
            tt2 = QTableWidget(len(tap), len(heads))
            tt2.setHorizontalHeaderLabels(heads)
            tt2.verticalHeader().setVisible(False)
            tt2.verticalHeader().setDefaultSectionSize(28)
            tt2.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
            tt2.setAlternatingRowColors(True)
            tt2.setFixedHeight(34 + 28 * len(tap))
            for r, row in enumerate(tap_a):
                live = row[4] != 0
                lo, hi = (row[5], row[6]) if len(row) > 6 else (np.nan, np.nan)
                mine = len(row) > 7 and row[7] == 1
                kind = int(row[8]) if len(row) > 8 else 1
                unit = {1: "", 2: "°", 3: " Mvar"}[kind]
                # 계단 (2026-08-14) — 10열 = 한 단 크기 · 11열 = 내렸나
                sz = float(row[9]) if len(row) > 9 else 0.0
                state = int(row[10]) if len(row) > 10 else 0
                lim = "—" if np.isnan(lo) else (
                    f"{lo:g} ~ {hi:g}{unit}" + ("  (자동)" if mine else ""))
                vals = [{1: "탭", 2: "위상", 3: "SVC"}[kind],
                        "—" if kind == 3 else f"{int(row[0])}",
                        "이 선로" if kind == 2 else f"버스 {int(row[1])}",
                        f"{row[2]:,.3f} MW" if kind == 2 else f"{row[2]:.4f} pu",
                        f"{row[3]:.4f}{unit}",
                        lim,
                        # 🚨 계단은 「목표 맞춤」이라 쓰면 거짓말이다 — 설 자리가
                        #    정해져 있어 목표에 **가까운 자리**에 설 뿐이다.
                        ("한계에 걸림 — 목표 포기" if not live
                         else "계단으로 못 내림 — 연속값 그대로" if state == 2
                         else f"계단 자리 (한 단 {sz:g}{unit})" if state == 1
                         else "목표 맞춤")]
                for cc, txt in enumerate(vals):
                    it = QTableWidgetItem(txt)
                    if cc > 0:
                        it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    if not live or state == 2:
                        it.setForeground(QColor(c["warn"]))
                    elif cc == 4 and mine:
                        it.setForeground(QColor(c["muted"]))
                    tt2.setItem(r, cc, it)
            cv.addWidget(tt2)
            if auto:
                # 사용자가 안 적어서 앱이 정한 값이다 — 어디를 고치면 되는지까지 적는다.
                # 탭이면 0.9~1.1, 위상이면 ±30° — 어느 쪽인지 갈라 적는다
                kinds = []
                if ((tap_a[:, 7] == 1) & (tap_a[:, 8] == 1)).any():
                    kinds.append("탭은 0.9 ~ 1.1")
                if ((tap_a[:, 7] == 1) & (tap_a[:, 8] == 2)).any():
                    kinds.append("위상은 -30 ~ 30°")
                if ((tap_a[:, 7] == 1) & (tap_a[:, 8] == 3)).any():
                    kinds.append("SVC 는 -50 ~ 50 Mvar")
                am = QLabel(
                    f"ⓘ {auto}대는 움직일 범위(Ctrl Min·Ctrl Max)를 적지 않아 "
                    f"{' · '.join(kinds)} 로 잡았습니다 — "
                    f"「계통 데이터」 에서 바꿀 수 있습니다.")
                am.setWordWrap(True)
                am.setStyleSheet(f"color:{c['muted']};font-size:12px;")
                cv.addWidget(am)
            outer.addWidget(card)

        # 무효출력 한계를 걸면 수렴하지 못하는 계통 — 아래 표는 한계를 적용하지 않은
        # 값이므로 눈에 띄게 밝힌다 (2026-07-31).
        qmsg = getattr(self.sol, "qlim_message", "") if self.sol else ""
        if qmsg:
            warn = QFrame()
            warn.setObjectName("card")
            # 경고색을 옅게 깐 배경. Qt 스타일시트는 '#RRGGBBAA' 를 못 읽으므로
            # rgba() 로 준다(그냥 색을 깔면 글자가 배경에 묻힌다).
            wc = QColor(c["warn"])
            warn.setStyleSheet(
                f"#card{{background:rgba({wc.red()},{wc.green()},{wc.blue()},0.12);"
                f"border:1px solid {c['warn']};}}")
            wv = QHBoxLayout(warn)
            wv.setContentsMargins(14, 10, 14, 10)
            wicon = QLabel("⚠")
            wicon.setStyleSheet(f"color:{c['warn']};font-size:16px;font-weight:700;")
            wv.addWidget(wicon)
            # 엔진이 보낸 문구를 **그대로** 쓴다. 앞에 "한계 적용 시 수렴 실패 —" 를
            # 붙이던 것은 문구 안에 이미 그 말이 들어 있어 겹쳤다 (2026-08-12).
            wtxt = QLabel(qmsg)
            wtxt.setWordWrap(True)
            wtxt.setStyleSheet(f"color:{c['warn']};font-size:13px;font-weight:700;")
            wv.addWidget(wtxt, 1)
            outer.addWidget(warn)

        # ── 한계로 묶인 발전기가 있으면 알린다 (2026-08-12, §7.6 G8)
        #    ⚠️ 이건 **경고가 아니라 참고**다. 한계로 묶는 것 자체는 정상 동작이라
        #    빨간 경고로 띄우면 정상 계통마다 떠서 경고가 무뎌진다.
        #    다만 **흡수 쪽(Qmin)에 걸리면 전압이 올라가는데** 처음 보면 놀랄 일이라 밝힌다.
        note = self._qlim_note()
        if note:
            info = QFrame()
            info.setObjectName("card")
            iv = QHBoxLayout(info)
            iv.setContentsMargins(14, 9, 14, 9)
            icon = QLabel("ⓘ")
            icon.setStyleSheet(f"color:{c['muted']};font-size:15px;font-weight:700;")
            iv.addWidget(icon)
            itxt = QLabel(note)
            itxt.setWordWrap(True)
            itxt.setStyleSheet(f"color:{c['muted']};font-size:12px;")
            iv.addWidget(itxt, 1)
            outer.addWidget(info)

        # ── 정격이 안 적혀 부하율을 못 재는 선로가 있으면 밝힌다 (2026-08-18)
        #    ⚠️ **조용히 빼면 그것대로 못 믿는다.** 과부하 판정에서 뺀 것은 맞지만,
        #       뺐다는 사실을 화면이 말해야 사용자가 "왜 0건이지?" 를 안 겪는다.
        #       (반대 방향의 옛 결함 = 이 선로들을 전부 과부하로 세어 IEEE 118버스에서
        #        거짓 경보 186건을 띄우던 것.)
        n_unrated = unrated_lines(self.sol, self.t) if self.sol is not None else 0
        if n_unrated:
            info2 = QFrame()
            info2.setObjectName("card")
            iv2 = QHBoxLayout(info2)
            iv2.setContentsMargins(14, 9, 14, 9)
            ic2 = QLabel("ⓘ")
            ic2.setStyleSheet(f"color:{c['muted']};font-size:15px;font-weight:700;")
            iv2.addWidget(ic2)
            it2 = QLabel(
                f"선로 {n_unrated}개는 정격(용량)이 안 적혀 있어 부하율을 재지 못했습니다 — "
                f"과부하 판정에서 뺐습니다.")
            it2.setWordWrap(True)
            it2.setStyleSheet(f"color:{c['muted']};font-size:12px;")
            iv2.addWidget(it2, 1)
            outer.addWidget(info2)

        for title, (cols, rows) in self.viol().items():
            box = QFrame()
            box.setObjectName("card")
            bv = QVBoxLayout(box)
            bv.setContentsMargins(14, 11, 14, 13)
            bv.setSpacing(8)
            th = QHBoxLayout()
            t = QLabel(title)
            t.setStyleSheet(f"color:{c['text']};font-size:14px;font-weight:700;")
            th.addWidget(t)
            cnt = QLabel(f"{len(rows)}건")
            cnt.setStyleSheet(
                f"color:{c['warn'] if rows else c['muted']};font-size:12px;")
            th.addWidget(cnt)
            th.addStretch()
            bv.addLayout(th)

            if rows:
                tb = QTableWidget(len(rows), len(cols))
                tb.setHorizontalHeaderLabels(cols)
                tb.verticalHeader().setVisible(False)
                tb.verticalHeader().setDefaultSectionSize(30)
                tb.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
                # 🚨 **상한만 걸면 표가 눌린다** (2026-08-18 실측).
                #    이 화면은 카드를 세로로 쌓고 끝에 `outer.addStretch()` 를 두는데,
                #    `addWidget` 으로 붙은 것은 **늘림 몫이 0**이라 남는 자리를 전부 그
                #    stretch 가 가져간다. 그래서 `setMaximumHeight` 는 아무 일도 안 하고
                #    표는 늘 제 기본 크기(88px)로 눌린 채 그려졌다 — **19건짜리 표가 한 줄**.
                #    26건을 잡아 놓고 화면이 3줄만 보여 주니 기능이 없는 것과 같았다.
                #    ⇒ **높이를 못박는다**(`setFixedHeight` 는 최소·최대를 함께 건다).
                #    다 보여 주되 한 표가 화면을 통째로 먹지 않게 **10줄에서 끊는다** —
                #    그보다 많으면 표 안에서 스크롤하고, 전체 건수는 제목 옆 "N건" 이 말한다.
                #    (건수에 상한이 없다 — `checks.real_violations` 는 걸린 것을 다 담는다.)
                tb.setFixedHeight(self.CHECK_ROW_H * min(len(rows), self.CHECK_MAX_ROWS)
                                  + self.CHECK_CHROME)
                for r, row in enumerate(rows):
                    for cc, val in enumerate(row):
                        it = QTableWidgetItem(val)
                        if cc > 1:
                            it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                        if cc == len(row) - 1:
                            it.setForeground(Qt.red)
                        tb.setItem(r, cc, it)
                bv.addWidget(tb)
            else:
                ok = QLabel("없음")
                ok.setStyleSheet(f"color:{c['muted']};font-size:12px;")
                bv.addWidget(ok)
            outer.addWidget(box)

        outer.addStretch()
        return w

    def _qlim_note(self):
        """무효출력 한계 때문에 알릴 것이 있으면 그 문구, 없으면 빈 글자.

        **아래 "발전기 한계" 표가 세는 것과 같은 것을 센다** — 표는 `satQ ~= 0` 인 줄만
        보여주므로 여기서도 그 열을 본다. 두 해법이 같은 표를 내므로 안내도 같아진다.
        ⚠️ 예전에는 엔진이 따로 보내 준 값을 썼는데, 그 값을 Gauss-Seidel 만 보내는 바람에
           **안내가 한쪽에만 뜨고 표와 숫자가 어긋났다**(2026-08-12 사용자 지적).

        내용은 표가 말하지 못하는 것만 담는다 — **흡수 쪽(Qmin)에 걸리면 전압이 올라간다**.
        발전기가 빨아들이던 무효전력을 더 못 빨아들이기 때문인데, 보통은 한계를 걸면
        전압이 내려간다고 생각하므로 처음 보면 놀랄 만한 일이다.
        """
        sol = self.sol
        tbl = getattr(sol, "gen_limit", None) if sol is not None else None
        if tbl is None or len(tbl) == 0:
            return ""
        try:
            sat_q = [int(r[9]) for r in tbl]
        except (IndexError, TypeError, ValueError):
            return ""
        dn = sum(1 for s in sat_q if s < 0)
        if not any(sat_q):
            return ""

        s = ""
        if dn:
            s = (f"무효전력을 빨아들이던 발전기 {dn}대가 흡수 한계에 걸렸습니다. "
                 "더 못 빨아들이므로 그 부근 전압이 오히려 올라갑니다.")
        if getattr(sol, "method", "nr") == "gs":
            s += (" " if s else "") + (
                "한계를 건 계통은 답이 여럿일 수 있어 Newton 방식과 다른 답이 나올 수 있습니다 "
                "— 결과가 이상해 보이면 Newton 으로도 풀어 견줘 보십시오.")
        return s.strip()

    # ── 해법 고르기 (2026-08-12, §7.6 G8) ──
    def _solver_picker(self):
        """Newton / Gauss-Seidel 을 고르는 자리. 못 쓰는 계통이면 흐리게 하고 까닭을 보여준다.

        왜 못 쓰는지 판정은 `app_engine.gs_refusal(case)`(여기서는 `ENGINE`) 이 한다 — 엔진도 같은 것을 검사해
        오류를 내지만(`runpfGS_app.m`), 여기서 **미리** 막아 사용자가 눌러 보고 실패를 겪지 않게 한다.
        """
        c = self.c
        box = QVBoxLayout()
        box.setSpacing(2)
        lab = QLabel("해법")
        lab.setStyleSheet(f"color:{c['muted']};font-size:11px;")
        box.addWidget(lab)

        pick = QComboBox()
        pick.addItem("Newton-Raphson", "nr")
        pick.addItem("Gauss-Seidel", "gs")
        pick.setCurrentIndex(1 if getattr(self, "solver", "nr") == "gs" else 0)

        why = None
        case = getattr(self, "_case_for_solver", None)
        if case is not None:
            try:
                why = ENGINE.gs_refusal(case)
            except Exception:
                why = None
        if why:
            # 고를 수 없게 흐리게 하고, 마우스를 올리면 까닭이 뜨게 한다
            model = pick.model()
            item = model.item(1)
            if item is not None:
                item.setEnabled(False)
            pick.setItemData(1, why, Qt.ToolTipRole)
            pick.setToolTip(why)
            if pick.currentIndex() == 1:
                pick.setCurrentIndex(0)

        pick.currentIndexChanged.connect(self._solver_changed)
        self._solver_pick = pick
        box.addWidget(pick)
        return box

    def _solver_changed(self, _idx):
        """해법을 바꾸면 **같은 계통을 그 해법으로 다시 푼다.**"""
        pick = getattr(self, "_solver_pick", None)
        if pick is None:
            return
        method = pick.currentData() or "nr"
        if method == getattr(self, "solver", "nr"):
            return
        path = getattr(self, "_last_path", None)
        if not path:
            self.solver = method
            return
        # 조건을 바꿔 푼 상태면 그 케이스를 그대로 쓴다 (원본으로 되돌리지 않는다)
        case = getattr(self, "_case_for_solver", None)
        self._start_solve(path, case, method)

    # ── 수렴 탭 ──
    def conv_page(self):
        c = self.c
        sol = self.sol
        conv = dict(
            converged=sol.converged if sol else CONV["converged"],
            iters=sol.iters if sol else CONV["iters"],
            threshold=sol.threshold if sol else CONV["threshold"],
            mis=list(sol.mis_history) if sol else CONV["mis"],
            blocks=list(sol.block_names) if sol and sol.block_names else CONV["blocks"],
            block_hist=(sol.block_history.tolist()
                        if sol is not None and sol.block_history.size
                        else CONV["block_hist"]),
            dominant=list(sol.dominant_block) if sol and sol.dominant_block
            else CONV["dominant"],
            seconds=sol.seconds if sol else CONV["seconds"],
        )
        w = QWidget()
        outer = QVBoxLayout(w)
        outer.setContentsMargins(14, 12, 14, 12)
        outer.setSpacing(11)

        # 요약 줄
        head = QFrame()
        head.setObjectName("card")
        hv = QHBoxLayout(head)
        hv.setContentsMargins(14, 11, 14, 11)
        hv.setSpacing(26)

        def kv(k, v, color=None):
            b = QVBoxLayout()
            b.setSpacing(2)
            a = QLabel(k)
            a.setStyleSheet(f"color:{c['muted']};font-size:11px;")
            d = QLabel(v)
            d.setStyleSheet(
                f"color:{color or c['text']};font-size:17px;font-weight:700;")
            b.addWidget(a)
            b.addWidget(d)
            return b

        hv.addLayout(kv("수렴", "성공" if conv["converged"] else "실패",
                        c["ok"] if conv["converged"] else c["warn"]))
        hv.addLayout(kv("반복 횟수", f"{conv['iters']}회"))
        hv.addLayout(kv("계산 시간", f"{conv['seconds']:.2f} s"))
        hv.addLayout(kv("수렴 기준", f"{conv['threshold']:g}"))
        hv.addLayout(kv("최종 불평형", f"{CONV['mis'][-1]:.2e}"))
        hv.addStretch()
        hv.addLayout(self._solver_picker())
        outer.addWidget(head)

        # 불평형이 줄어드는 과정
        box = QFrame()
        box.setObjectName("card")
        bv = QVBoxLayout(box)
        bv.setContentsMargins(14, 11, 14, 13)
        t = QLabel("반복에 따른 최대 불평형")
        t.setStyleSheet(f"color:{c['text']};font-size:14px;font-weight:700;")
        bv.addWidget(t)
        sub = QLabel("실제 그래프가 들어갈 자리 — 세로축은 로그 눈금")
        sub.setStyleSheet(f"color:{c['muted']};font-size:11px;")
        bv.addWidget(sub)
        bar = QHBoxLayout()
        bar.setSpacing(10)
        for i, m in enumerate(conv["mis"]):
            cell = QFrame()
            cell.setObjectName("plot")
            cv = QVBoxLayout(cell)
            cv.setContentsMargins(10, 8, 10, 8)
            a = QLabel(f"{i}회")
            a.setStyleSheet(f"color:{c['muted']};font-size:11px;")
            d = QLabel(f"{m:.2e}")
            d.setStyleSheet(
                f"color:{c['ok'] if m < conv['threshold'] else c['text']};"
                f"font-size:15px;font-weight:700;")
            cv.addWidget(a)
            cv.addWidget(d)
            bar.addWidget(cell)
        bar.addStretch()
        bv.addLayout(bar)
        outer.addWidget(box)

        # 블록별 + 발목 잡은 곳
        box2 = QFrame()
        box2.setObjectName("card")
        b2 = QVBoxLayout(box2)
        b2.setContentsMargins(14, 11, 14, 13)
        b2.setSpacing(8)
        t2 = QLabel("무엇이 수렴을 늦추나 — 반복별 블록 최대 불평형")
        t2.setStyleSheet(f"color:{c['text']};font-size:14px;font-weight:700;")
        b2.addWidget(t2)
        tb = QTableWidget(len(conv["block_hist"]), len(conv["blocks"]) + 2)
        tb.setHorizontalHeaderLabels(["반복"] + conv["blocks"] + ["가장 큰 블록"])
        tb.verticalHeader().setVisible(False)
        tb.verticalHeader().setDefaultSectionSize(30)
        tb.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        tb.setMinimumHeight(30 * len(conv["block_hist"]) + 46)
        tb.setMaximumHeight(30 * len(conv["block_hist"]) + 46)
        for r, row in enumerate(conv["block_hist"]):
            tb.setItem(r, 0, QTableWidgetItem(str(r + 1)))
            worst = max(range(len(row)), key=lambda i: row[i])
            for cc, v in enumerate(row):
                it = QTableWidgetItem("—" if v == 0 else f"{v:.1e}")
                it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                if cc == worst and v > 0:
                    it.setForeground(Qt.red)
                tb.setItem(r, cc + 1, it)
            tb.setItem(r, len(row) + 1, QTableWidgetItem(conv["dominant"][r]))
        b2.addWidget(tb)
        note = QLabel("조류계산은 방정식 묶음 8종의 오차를 함께 줄여 나갑니다. "
                      "그중 오차가 가장 큰 묶음이 수렴 속도를 결정합니다 (빨간 값). "
                      "계산이 안 끝날 때 어느 방정식이 문제인지 여기서 보입니다.")
        note.setStyleSheet(f"color:{c['muted']};font-size:11px;")
        b2.addWidget(note)
        outer.addWidget(box2)

        outer.addStretch()
        return w

    def set_bus(self, idx):
        self.bus_row = int(idx)
        self.rebuild()

    def set_time(self, idx):
        self.t = int(idx)
        self.rebuild()

    def set_vsc(self, on):
        self.show_vsc = bool(on)
        self.rebuild()

    def toggle_vsc(self):
        self.set_vsc(not self.show_vsc)

    def set_violations(self, on):
        """계통도 '위반 보기' 켜기/끄기. 계통도 토글 버튼이 부른다."""
        self.show_violations = bool(on)
        self.rebuild()

    def show_line_profile(self, g, ei):
        """계통도에서 선로를 클릭하면 그 선로의 24시간 부하율을 팝업으로 띄운다.
        시간별 데이터가 없는 케이스(스냅샷 1시각)면 그 사실을 적어 준다."""
        if self.sol is None:
            return
        import topology
        c = self.c
        title = topology.edge_label(g, ei)
        ser = topology.loading_series(g, self.sol, ei)
        d = QDialog(self)
        d.setWindowTitle(f"{title} — 24시간 부하율")
        d.setStyleSheet(self.styleSheet())
        d.setMinimumSize(560, 380)
        v = QVBoxLayout(d)
        v.setContentsMargins(16, 14, 16, 14)
        v.setSpacing(10)
        if ser is None:
            body = charts._note(c, f"{title} 는 부하율 결과가 없습니다.\n"
                                   "(변환기·3권선 지선이거나 Branch 표에 없는 선로)")
        elif len(ser[0]) <= 1:
            pct = ser[1][0] if ser[1] else float("nan")
            body = charts._note(
                c, f"{title}\n이 케이스는 시간별(24h) 데이터가 없습니다.\n"
                   f"현재 부하율 = {pct:.1f}%")
        else:
            body = charts.loading_profile_view(c, self.sol, ser[0], ser[1],
                                               f"{title}  ·  24시간 부하율")
        v.addWidget(body, 1)
        close = QPushButton("닫기")
        close.clicked.connect(d.accept)
        row = QHBoxLayout(); row.addStretch(); row.addWidget(close)
        v.addLayout(row)
        d.exec()

    def pick_columns(self):
        name = self._tabs.tabText(self._tabs.currentIndex())
        d = QDialog(self)
        d.setWindowTitle(f"열 선택 — {name}")
        d.setStyleSheet(self.styleSheet())
        d.setMinimumWidth(320)
        v = QVBoxLayout(d)
        v.setContentsMargins(20, 18, 20, 18)
        v.setSpacing(7)
        info = QLabel("표에 보일 열을 고르세요")
        info.setStyleSheet(f"color:{self.c['muted']};font-size:13px;")
        v.addWidget(info)
        boxes = []
        for col, _ in TABLE_SPECS[name]:
            b = QCheckBox(col)
            b.setChecked(col in self.visible[name])
            v.addWidget(b)
            boxes.append((col, b))
        row = QHBoxLayout()
        row.addStretch()
        cancel = QPushButton("취소")
        cancel.clicked.connect(d.reject)
        ok = QPushButton("적용")
        ok.setObjectName("primary")
        ok.clicked.connect(d.accept)
        row.addWidget(cancel)
        row.addWidget(ok)
        v.addLayout(row)
        if d.exec():
            picked = {col for col, b in boxes if b.isChecked()}
            if picked:
                self.visible[name] = picked
                self.rebuild()

    def compare_area(self):
        c = self.c
        wide = self.compare_axis in ("시간끼리", "시나리오끼리")
        picked = [n for n, always in COMPARE_ITEMS
                  if n in self.picked and (always or wide)]
        tabs = QTabWidget()
        if not picked:
            page = QWidget()
            pv = QVBoxLayout(page)
            lb = QLabel("왼쪽에서 볼 항목을 하나 이상 고르세요")
            lb.setAlignment(Qt.AlignCenter)
            lb.setStyleSheet(f"color:{c['muted']};font-size:16px;")
            pv.addWidget(lb)
            tabs.addTab(page, "결과")
            return tabs
        if self.compare_axis == "시나리오끼리":
            return self.compare_scenarios_area(picked)
        targets = [t.strip() for t in self.compare_targets.split(",") if t.strip()]
        unit = "버스" if self.compare_axis == "버스끼리" else "시간"
        for name in picked:
            page = QWidget()
            pv = QVBoxLayout(page)
            pv.setContentsMargins(10, 10, 10, 10)
            pv.setSpacing(9)
            # 원본 앱은 "내보내기"를 눌러야 팝업창에 비교 그림을 그려 줬지만,
            # 여기서는 화면에 바로 그린다 (charts.compare_chart).
            box = charts.compare_chart(c, self.sol, name, self.compare_axis,
                                       targets)
            pv.addWidget(box, 2)
            tb = self.compare_table(name, targets)
            if tb is not None:
                pv.addWidget(tb, 1)
            tabs.addTab(page, name)
        return tabs

    def save_compare_figures(self):
        """지금 보고 있는 비교 그림들을 그대로 파일로. (일반 내보내기와 따로)"""
        if self.sol is None:
            QMessageBox.information(self, "비교 그림 저장",
                                    "먼저 케이스를 불러와 계산하세요.")
            return
        items = [n for n, always in COMPARE_ITEMS
                 if n in self.picked and (always or self.compare_axis == "시간끼리")]
        if not items:
            QMessageBox.information(self, "비교 그림 저장",
                                    "왼쪽에서 볼 항목을 하나 이상 고르세요.")
            return
        targets = [t.strip() for t in self.compare_targets.split(",") if t.strip()]
        folder = exporter.default_folder(self.sol.case_name)
        QApplication.setOverrideCursor(Qt.WaitCursor)
        try:
            files = exporter.save_compare_figures(
                self.c, self.sol, self.compare_axis, targets, items, folder)
        except Exception as exc:
            QApplication.restoreOverrideCursor()
            QMessageBox.critical(self, "비교 그림 저장 실패", f"{exc}")
            return
        QApplication.restoreOverrideCursor()
        if not files:
            QMessageBox.information(
                self, "비교 그림 저장",
                "저장할 그림이 없습니다. 비교할 번호를 확인해 주세요.")
            return
        names = "\n".join(f"  · {p.name}" for p in files)
        box = QMessageBox(self)
        box.setWindowTitle("비교 그림 저장 완료")
        box.setText(f"{len(files)}개 파일을 저장했습니다.\n\n{folder}\n\n{names}")
        box.addButton("확인", QMessageBox.AcceptRole)
        opn = box.addButton("폴더 열기", QMessageBox.ActionRole)
        box.exec()
        if box.clickedButton() is opn:
            import subprocess
            subprocess.run(["open", str(folder)], check=False)

    def compare_rows(self, item, targets):
        """비교 대상들의 값 — (열이름들, 행들). 화면 표와 내보내기가 같이 쓴다.

        표를 만드는 부분과 값을 뽑는 부분을 갈라 둔 이유는, 내보내기에서
        **화면 위젯을 읽지 않고** 같은 숫자를 그대로 엑셀로 쓰기 위해서다.
        """
        sol = self.sol
        if sol is None or not targets or not sol.AC.size:
            return None
        col = {"전압 크기": "VM[pu]", "위상각": "Angle[deg]"}.get(item)
        if col is None or col not in sol.cols("AC"):
            return None
        ci = sol.cols("AC").index(col)
        bus_ids = [int(b) for b in sol.AC[:, 0, 0]]

        if self.compare_axis == "버스끼리":
            rows, head = [], ["Time[h]"]
            picks = []
            for tgt in targets:
                if tgt.isdigit() and int(tgt) in bus_ids:
                    picks.append(bus_ids.index(int(tgt)))
                    head.append(f"버스 {tgt}")
            if not picks:
                return None
            for t in range(sol.AC.shape[2]):
                rows.append([t + 1] + [sol.AC[p, ci, t] for p in picks])
        else:
            head = ["Bus"] + [f"{t}H" for t in targets if t.isdigit()]
            times = [int(t) - 1 for t in targets if t.isdigit()]
            times = [t for t in times if 0 <= t < sol.AC.shape[2]]
            if not times:
                return None
            rows = [[bus_ids[r]] + [sol.AC[r, ci, t] for t in times]
                    for r in range(sol.AC.shape[0])]
        return head, rows

    def compare_table(self, item, targets):
        """비교 대상들의 실제 값을 표로 (그래프가 들어가기 전까지 숫자로 확인)."""
        got = self.compare_rows(item, targets)
        if got is None:
            return None
        head, rows = got
        tb = QTableWidget(len(rows), len(head))
        tb.setHorizontalHeaderLabels(head)
        tb.verticalHeader().setVisible(False)
        tb.verticalHeader().setDefaultSectionSize(28)
        tb.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        tb.setAlternatingRowColors(True)
        for r, row in enumerate(rows):
            for cc, val in enumerate(row):
                txt = f"{val:.0f}" if cc == 0 else f"{val:.4f}"
                it = QTableWidgetItem(txt)
                if cc > 0:
                    it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                tb.setItem(r, cc, it)
        return tb

    # ── 하단 ──
    def statusbar(self):
        c = self.c
        bar = QFrame()
        bar.setObjectName("statusbar")
        bar.setFixedHeight(46)
        h = QHBoxLayout(bar)
        h.setContentsMargins(18, 0, 18, 0)
        h.setSpacing(20)

        def item(k, val, color=None):
            box = QHBoxLayout()
            box.setSpacing(6)
            a = QLabel(k)
            a.setStyleSheet(f"color:{c['muted']};font-size:14px;")
            b = QLabel(val)
            b.setStyleSheet(
                f"color:{color or c['text']};font-size:14px;font-weight:600;")
            box.addWidget(a)
            box.addWidget(b)
            return box

        # 곡선 화면에서는 **곡선 것**을 말한다. 조류계산의 위반·반복 횟수를 그대로 두면
        # 지금 보고 있는 그림과 상관없는 숫자가 아래에 남는다 (F1d 는 둘을 갈랐다).
        if self.task == "PV·QV 곡선":
            cur = self.cur
            dot = QLabel("●")
            good = cur is not None
            dot.setStyleSheet(
                f"color:{c['ok'] if good else c['muted']};font-size:13px;")
            h.addWidget(dot)
            if cur is None:
                h.addLayout(item("PV·QV 곡선", "아직 안 그림"))
            else:
                h.addLayout(item("버틸 수 있는 부하", f"{cur.nose_MW:,.1f} MW"))
                h.addLayout(item("남은 여유", f"{cur.lam_crit * 100:.1f} %"))
                h.addLayout(item("걸음", f"{cur.lam.size:,}회"))
                h.addLayout(item("한계에 걸린 발전기", f"{cur.switched.size}대",
                                 c["warn"] if cur.switched.size else None))
                h.addLayout(item("계산 시간", f"{cur.seconds:.2f} s"))
            h.addStretch(1)
            return bar

        sol = self.sol
        n = violation_count(self.viol())
        # 점 하나로 전체 상태를 알린다 — 수렴 실패든 위반이든 있으면 주황
        healthy = (sol is None or sol.converged) and n == 0
        dot = QLabel("●")
        dot.setStyleSheet(f"color:{c['ok'] if healthy else c['warn']};font-size:13px;")
        h.addWidget(dot)

        # 봐야 할 순서대로 — 위반이 있으면 그게 제일 앞
        vb = QPushButton(f"위반 {n}건" if n else "위반 없음")
        vb.setCursor(Qt.PointingHandCursor)
        vb.setToolTip("누르면 점검 탭으로 갑니다")
        col = c["warn"] if n else c["ok"]
        vb.setStyleSheet(
            f"border:none;background:transparent;color:{col};"
            f"font-size:14px;font-weight:700;padding:0;text-align:left;")
        vb.clicked.connect(self.go_check)
        h.addWidget(vb)

        if sol is not None:
            h.addLayout(item("수렴", "성공" if sol.converged else "실패",
                             c["ok"] if sol.converged else c["warn"]))
            h.addLayout(item("반복", f"{sol.iters}회"))
            # 파일 읽기 + 계산 전체. 엔진 기동이 섞였으면 그렇다고 밝힌다.
            tt = f"{sol.seconds:.2f} s"
            if not sol.warm_start:
                tt += "  (첫 계산 — 준비 시간 포함)"
            h.addLayout(item("계산 시간", tt))
        else:
            h.addLayout(item("수렴", "—", c["muted"]))
            h.addLayout(item("반복", "—"))
            h.addLayout(item("계산 시간", "—"))
        h.addStretch()
        return bar

    def _restore_tab(self, tt):
        """다시 그린 뒤 보고 있던 표 탭으로 되돌린다. 그 탭이 사라졌으면 첫 탭."""
        want = getattr(self, "table_tab", None)
        if not want:
            return
        for i in range(tt.count()):
            if _tab_base(tt.tabText(i)) == want:
                tt.setCurrentIndex(i)
                return

    def go_check(self):
        """상태바의 위반 건수 → 점검 탭으로."""
        if self.mode == "비교":       # 비교 모드엔 표가 없으니 스냅샷으로 돌아간다
            self.mode = "스냅샷"
            self.rebuild()
        tt = getattr(self, "_tabs", None)
        if tt is None:
            return
        for i in range(tt.count()):
            if tt.tabText(i).startswith("점검"):
                tt.setCurrentIndex(i)
                self.table_tab = "점검"      # 다시 그려도 여기 머문다
                return

    # ── 동작 ──
    def set_mode(self, m):
        self.mode = m
        self.rebuild()

    def set_axis(self, a):
        self.compare_axis = a
        self.rebuild()

    # ── PV·QV 곡선 (F1d) ──────────────────────────────────────────────
    def curve_case(self):
        """곡선을 그릴 케이스 — **지금 화면의 조건이 전부 얹힌 것**.

        조류계산을 돌렸는지와 무관하다(아직 계산 안 한 `changes` 까지 얹는다).
        곡선은 자기 실행이므로 "이 조건으로 계산" 을 먼저 누를 필요가 없다.
        """
        if self.base_case is None:
            return None
        want = self.applied + self.changes
        return SC.apply(self.base_case, want) if want else self.base_case

    def curve_why(self):
        """곡선을 못 그리면 그 까닭. 그릴 수 있으면 None."""
        case = self.curve_case()
        if case is None:
            return "계통 파일을 먼저 여십시오."
        try:
            return ENGINE.curve_refusal(case)
        except Exception as exc:                     # noqa: BLE001
            return f"확인할 수 없습니다 ({exc})."

    def _bus_list(self, text):
        out = []
        for part in (text or "").replace(";", ",").split(","):
            part = part.strip()
            if not part:
                continue
            try:
                out.append(int(float(part)))
            except ValueError:
                continue
        return out

    def curve_controls(self, v):
        c = self.c
        lb = QLabel("부하를 늘릴 버스")
        lb.setStyleSheet(f"color:{c['muted']};font-size:13px;font-weight:700;")
        v.addWidget(lb)
        le = QLineEdit(self.curve_load)
        le.setPlaceholderText("비우면 부하가 있는 버스 전부")
        le.textChanged.connect(lambda t: setattr(self, "curve_load", t))
        v.addWidget(le)
        n = QLabel("여기에 적은 버스의 부하만 함께 늘립니다")
        n.setWordWrap(True)
        n.setStyleSheet(f"color:{c['muted']};font-size:12px;")
        v.addWidget(n)
        v.addSpacing(10)

        lb2 = QLabel("곡선을 그릴 버스")
        lb2.setStyleSheet(f"color:{c['muted']};font-size:13px;font-weight:700;")
        v.addWidget(lb2)
        le2 = QLineEdit(self.curve_pick)
        le2.setPlaceholderText("비우면 늘린 버스와 같게 (최대 8개)")
        le2.textChanged.connect(self.set_curve_pick)
        v.addWidget(le2)
        v.addSpacing(10)

        lb3 = QLabel("가로축")
        lb3.setStyleSheet(f"color:{c['muted']};font-size:13px;font-weight:700;")
        v.addWidget(lb3)
        seg = QFrame()
        seg.setObjectName("segwrap")
        seg.setStyleSheet(
            f"#segwrap {{ background:{c['bg']};border:1px solid {c['border']};"
            f"border-radius:8px; }}")
        sh = QHBoxLayout(seg)
        sh.setContentsMargins(3, 3, 3, 3)
        sh.setSpacing(3)
        for key, name in (("MW", "합계 부하"), ("lambda", "배수 λ")):
            b = QPushButton(name)
            b.setObjectName("seg_on" if self.curve_x == key else "seg_off")
            b.setCursor(Qt.PointingHandCursor)
            b.clicked.connect(lambda _, x=key: self.set_curve_x(x))
            sh.addWidget(b)
        v.addWidget(seg)
        v.addSpacing(14)

        run = QPushButton("곡선 그리기" if not self.curve_busy else "그리는 중…")
        run.setObjectName("primary")
        run.setEnabled(not self.curve_busy)
        run.clicked.connect(self.run_curve)
        v.addWidget(run)
        n2 = QLabel("버스가 많으면 몇 분 걸립니다")
        n2.setWordWrap(True)
        n2.setStyleSheet(f"color:{c['muted']};font-size:12px;")
        v.addWidget(n2)

    def set_curve_pick(self, t):
        self.curve_pick = t
        if self.cur is not None:
            self.rebuild()          # 이미 그린 곡선이면 다시 계산하지 않고 골라 보기만

    def set_curve_x(self, x):
        self.curve_x = x
        self.rebuild()

    def set_task(self, t):
        self.task = t
        self.rebuild()

    def run_curve(self):
        case = self.curve_case()
        if case is None or self.curve_busy:
            return
        self.curve_busy = True
        self.curve_err = ""
        self.rebuild()
        self._cthread = CurveThread(case, self._bus_list(self.curve_load),
                                    self._bus_list(self.curve_pick))
        self._cthread.done.connect(self._curve_done)
        self._cthread.failed.connect(self._curve_failed)
        self._cthread.engine_missing.connect(self._curve_failed)
        self._cthread.start()

    def _curve_done(self, cur):
        self.curve_busy = False
        self.cur = cur
        self.curve_err = ""
        self.rebuild()

    def _curve_failed(self, msg):
        self.curve_busy = False
        self.curve_err = msg
        self.rebuild()

    def curve_page(self):
        c = self.c
        w = QWidget()
        v = QVBoxLayout(w)
        v.setContentsMargins(0, 0, 0, 0)
        v.setSpacing(11)

        why = self.curve_why()
        if why:
            v.addWidget(charts.NoData("곡선을 그릴 수 없습니다", why, c))
            return w
        if self.curve_err:
            v.addWidget(charts.NoData("곡선 실패", self.curve_err, c))
            return w
        if self.curve_busy:
            v.addWidget(charts.NoData(
                "곡선을 그리는 중입니다…",
                "부하를 조금씩 올리며 더는 안 풀리는 곳까지 갑니다.", c))
            return w
        if self.cur is None:
            v.addWidget(charts.NoData(
                "왼쪽에서 [곡선 그리기] 를 누르십시오",
                "부하를 늘려 가며 전압이 무너지는 지점(코 끝점)을 찾습니다.", c))
            return w

        v.addWidget(self.curve_summary())
        split = QSplitter(Qt.Vertical)
        split.setChildrenCollapsible(False)
        split.setHandleWidth(10)
        pick = self._bus_list(self.curve_pick)
        split.addWidget(charts.curve_chart(c, self.cur, pick, self.curve_x))
        split.addWidget(charts.curve_q_chart(c, self.cur, self.curve_x))
        split.setSizes([600, 380])
        v.addWidget(split, 1)
        return w

    def curve_summary(self):
        """코 끝점 한 줄 요약 — 곡선에서 사람이 실제로 가져가는 숫자."""
        c = self.c
        cur = self.cur
        card = Card(c)          # Card 는 이미 세로 layout 을 갖고 있다 — 그 안에 넣는다
        h = QHBoxLayout()
        h.setContentsMargins(2, 0, 2, 0)
        h.setSpacing(26)
        card.v.addLayout(h)

        base = cur.load_MW[0] if cur.load_MW.size else 0.0
        room = cur.nose_MW - base
        items = [
            ("지금 부하", f"{base:,.1f} MW"),
            ("버틸 수 있는 부하", f"{cur.nose_MW:,.1f} MW"),
            ("남은 여유", f"{room:,.1f} MW  ({cur.lam_crit * 100:.1f} %)"),
            ("코 끝점 최저 전압", f"{float(cur.v[:, min(cur.nose, cur.v.shape[1]-1)].min()):.4f} p.u."),
        ]
        for name, val in items:
            box = QVBoxLayout()
            k = QLabel(name)
            k.setStyleSheet(f"color:{c['muted']};font-size:12px;")
            val_l = QLabel(val)
            val_l.setStyleSheet(f"color:{c['text']};font-size:17px;font-weight:800;")
            box.addWidget(k)
            box.addWidget(val_l)
            h.addLayout(box)
        h.addStretch(1)

        if cur.switched.size:
            names = ", ".join(str(int(b)) for b in cur.switched[:6])
            more = "…" if cur.switched.size > 6 else ""
            note = QLabel(f"무효출력 한계에 걸린 발전기 {cur.switched.size}대\n"
                          f"버스 {names}{more}")
            note.setStyleSheet(f"color:{c['warn']};font-size:12px;")
            h.addWidget(note)
        return card

    def set_targets(self, t):
        self.compare_targets = t

    def toggle_item(self, name, state):
        if state:
            self.picked.add(name)
        else:
            self.picked.discard(name)
        self.rebuild()

    def set_numbers(self, v):
        self.numbers = v
        self.numbers_auto = False        # 사용자가 직접 골랐다 — 자동이 아니다
        # ⚠️ **이유도 함께 지운다.** 안 지우면 직접 접었는데도 앞서 자동으로 접혔던
        #    이유가 그대로 남아 *"계통을 바꿔 계산해서…"* 같은 남의 문구를 달게 된다
        #    (접힘 → 펼침 → 직접 접기 순서에서 실제로 그렇게 된다).
        self.numbers_why = ""
        if not v:
            # 직접 펼쳤다 ⇒ 이 케이스에서는 **더 이상 자동으로 접지 않는다**.
            # (조건을 바꿀 때마다 도로 접히면 못 쓴다)
            self.graph_kept = True
            # 🚨 **적어 둔 자리를 버린다.** 안 버리면 접히기 전의 「표 66%」 가 그대로
            #    되살아나 그래프가 바닥값만 받는다 — 펼쳐 놓고도 못 읽는 꼴이 된다.
            if isinstance(self.split_sizes, dict):
                self.split_sizes.pop(self._split_slot(), None)
        self.rebuild()

    def toggle_theme(self):
        self.dark = not self.dark
        self.rebuild()

    def do_import(self):
        # 파일 고르기 창이 처음 보여 줄 자리. 저장소의 검증용 케이스 폴더가 있으면 거기서
        # 시작한다(뼈대에서는 케이스가 널려 있던 v14 폴더였다).
        cases = Path(__file__).resolve().parent.parent / "cases"
        start = str(cases if cases.is_dir() else Path.home())
        path, _ = QFileDialog.getOpenFileName(
            self, "계통 파일 선택", start,
            "계통 파일 (*.xlsx *.m *.raw);;모든 파일 (*)")
        if not path:
            return
        if load_case is None:
            QMessageBox.warning(self, "불러오기",
                                "케이스 읽기 모듈(load_case)을 찾지 못했습니다.")
            return

        self._start_solve(path)

    def _start_solve(self, path, case=None, method=None):
        # 🚨 **다른 파일**을 열 때는 아직 계산 안 한 "바꾼 것"을 버린다 (2026-08-12 확인).
        #    안 버리면 `_solved` 의 원본 갱신 분기(`not self.changes`)가 건너뛰어져
        #    **화면은 새 계통인데 base_case·시나리오·곡선은 앞 계통 것**으로 남는다
        #    (case14 에서 선로를 끈 채 case118 을 여니 화면 118버스 / base_case case14).
        #    같은 파일을 다시 푸는 것([이 조건으로 계산]·다시 풀기)에는 손대지 않는다 —
        #    거기서 지우면 사용자가 방금 한 편집이 사라진다.
        if case is None and getattr(self, "_last_path", None) not in (None, path):
            self.changes = []
            self.cur = None
            self.curve_err = ""
        self._last_path = path
        if method is None:
            method = getattr(self, "solver", "nr")
        self.solver = method
        self.prog = QProgressDialog("조류계산 중입니다...", None, 0, 0, self)
        self.prog.setWindowTitle("UNIGRID")
        self.prog.setWindowModality(Qt.WindowModal)
        self.prog.setMinimumWidth(340)
        self.prog.setCancelButton(None)
        self.prog.show()

        self.thread = SolveThread(path, case, method)
        self.thread.done.connect(self._solved)
        self.thread.failed.connect(self._solve_failed)
        self.thread.engine_missing.connect(self._engine_missing)
        self.thread.start()

    def _solved(self, sol):
        if getattr(self, "prog", None) is not None:
            self.prog.close()
        # 조건을 안 바꾸고 푼 것이면 이 케이스가 **원본**이다 (바꿔서 푼 것은 원본이 아니다).
        loaded = getattr(getattr(self, "thread", None), "loaded_case", None)
        pending = getattr(self, "_pending", None)
        if pending:
            # [이 조건으로 계산] 으로 푼 것 — 시나리오 한 줄로 담고 바꾼 목록을 비운다
            self.book.add(self.base_case, pending, solution=sol,
                          name=self._new_name(pending))
            self.applied = list(pending)      # 이제 이것이 화면의 조건이다
            self.changes = []
            self._pending = None
            # 🚨 **계통을 바꿔 계산하면 그래프를 접고 표를 넓게 연다** (2026-08-15 사용자 확정).
            #    계통을 고칠 때 눈이 가는 것은 표인데, 그래프가 자리를 크게 먹어 표가
            #    한두 줄로 눌렸다(사용자: *"아래 표가 너무 작아서 보기 힘들다"*).
            #    보고 싶으면 [그래프 펼치기] 를 누르면 되고, 그때는 표가 도로 줄어든다.
            #    ⚠️ **한 번이라도 직접 펼쳤으면 다시 안 접는다** — 조건 하나 바꿀 때마다
            #       도로 접히면 못 쓴다(큰 계통 자동 접기에서 정한 것과 같은 이유, 2026-08-06).
            if not self.graph_kept:
                self.numbers = True
                self.numbers_auto = True
                self.numbers_why = "changed"
        elif loaded is not None and not self.changes:
            self.base_case = loaded
            self.applied = []
            self.changes = []
            self.book = SC.Book()
            self.book.add(loaded, [], solution=sol, name="원본")
            # 새 계통을 열면 곡선은 버린다 — **앞 계통의 곡선**이라 지금 화면과 상관없다.
            self.cur = None
            self.curve_err = ""
            # 🚨 큰 계통은 **그래프를 접은 채로 연다** (2026-08-06 사용자 확정).
            #    버스가 수천이면 점이 겹쳐 빨간 덩어리가 되어 읽을 수가 없다.
            #    보고 싶으면 [그래프 펼치기] 를 누르면 된다.
            #    조건을 바꿔 다시 풀 때는 손대지 않는다 — 사용자가 펼쳐 놓았으면 그대로 둔다.
            n_bus = int(sol.AC.shape[0]) + int(sol.DC.shape[0] if sol.DC.size else 0)
            self.numbers_auto = n_bus > BIG_BUSES
            self.numbers = self.numbers_auto
            self.numbers_why = "big" if self.numbers_auto else ""
            self.graph_kept = False       # 새 계통 — 「직접 펼쳤다」 기억도 새로
            # 정렬 기억도 새로 (2026-08-18). 계통이 달라지면 열 구성부터 달라질 수
            # 있고(AC 단독 ↔ AC/DC), 앞 계통을 보던 눈이 새 계통에 그대로 걸려 있으면
            # 파일을 열자마자 영문 모를 순서로 늘어서 있다. 위 `graph_kept` 와 같은 계열.
            self.sort_by = {}
        # 해법 고르기가 볼 케이스 — 방금 푼 그것이다(조건을 바꿔 푼 것이면 그 케이스).
        # 이 값으로 `gs_refusal` 을 물어 Gauss-Seidel 을 흐리게 할지 정한다 (2026-08-12, G8).
        self._case_for_solver = (getattr(getattr(self, "thread", None), "case", None)
                                 or loaded or getattr(self, "base_case", None))
        self.sol = sol
        # 🚨 곡선은 **AC 단독 계통에서만** 그린다(2026-08-12 사용자 확정 — 넓히지 않는다).
        #    곡선 화면에 있는 채로 AC/DC 파일을 열면 갈래가 곡선에 머물러 있는데
        #    그 버튼은 흐려져 있어 **죽은 화면에 앉게 된다** ⇒ 조류계산으로 되돌린다.
        if self.task == "PV·QV 곡선" and self.curve_why():
            self.task = "조류계산"
            self.cur = None
        self.t = 0
        self.bus_row = 0
        self.show_violations = False      # 새 케이스는 위반 보기 꺼진 채로 시작
        self.case = (Path(sol.case_name).name or "case",
                     f"{sol.mode_name} · AC {sol.AC.shape[0]} / DC {sol.DC.shape[0]}")
        self.case_has_vsc = sol.VSC_bus is not None and sol.VSC_bus.size > 0
        save_recent(str(getattr(self, "_last_path", sol.case_name)), self.case[1])
        if not self.case_has_vsc:
            self.show_vsc = False
        self.rebuild()

    def _new_name(self, pending):
        """시나리오 이름 — 이번에 새로 얹은 것으로 짓고, 없으면 전체로."""
        fresh = getattr(self, "_pending_new", None)
        self._pending_new = None
        return SC.auto_name(self.base_case, fresh if fresh else pending)

    def _solve_failed(self, msg):
        if getattr(self, "prog", None) is not None:
            self.prog.close()
        pending = getattr(self, "_pending", None)
        if pending:
            # 조건을 바꿔 풀다가 안 풀린 것 — **시나리오는 살려 둔다.**
            # 무엇을 바꿨는지가 목록에 남아야 다음 판단을 할 수 있다(PDR §4.3).
            # 화면은 그대로 두므로 직전 결과가 지워지지 않는다.
            self.book.add(self.base_case, pending, error=msg,
                          name=self._new_name(pending))
            self.changes = []          # applied 는 그대로 — 화면은 직전 결과 그대로다
            self._pending = None
            self.rebuild()
            QMessageBox.warning(
                self, "안 풀렸습니다",
                f"{SC.describe(pending)}\n\n이 조건으로는 답을 찾지 못했습니다. "
                f"시나리오 목록에 '안 풀림' 으로 남겨 두었고, 화면은 그대로 둡니다."
                f"\n\n{msg[:600]}")
            return
        QMessageBox.critical(self, "계산 실패", msg[:1500])

    def _engine_missing(self, msg):
        """계산 엔진을 못 찾았을 때 — 어디를 찾아봤는지까지 담긴 안내를 띄운다.

        안내문이 "[직접 고르기] 로 알려 주세요" 라고 말하므로 **그 버튼이 실제로 있어야 한다.**
        고른 자리는 `engine_path.remember()` 가 기억하므로 다음부터는 묻지 않는다.
        """
        if getattr(self, "prog", None) is not None:
            self.prog.close()
        box = QMessageBox(self)
        box.setIcon(QMessageBox.Warning)
        box.setWindowTitle("계산 엔진을 찾지 못했습니다")
        box.setText(msg[:2000])
        pick = box.addButton("직접 고르기…", QMessageBox.ActionRole)
        box.addButton("닫기", QMessageBox.RejectRole)
        box.exec()
        if box.clickedButton() is pick:
            self._pick_engine()

    def _pick_engine(self):
        """mwpython 자리를 직접 고른다. 쓸 수 있는 자리면 기억하고 다시 계산한다."""
        start = "/Applications"
        path, _ = QFileDialog.getOpenFileName(
            self, "mwpython 자리 고르기 (보통 <설치자리>/bin/mwpython)", start)
        if not path:
            return
        p = Path(path)
        if p.name != "mwpython":
            QMessageBox.warning(
                self, "다시 골라 주세요",
                f"고른 파일 이름이 'mwpython' 이 아닙니다: {p.name}\n"
                "보통 <MATLAB 또는 Runtime 설치자리>/bin/mwpython 입니다.")
            return
        if not os.access(p, os.X_OK):
            QMessageBox.warning(self, "다시 골라 주세요",
                                f"이 파일은 실행할 수 없습니다:\n{p}")
            return

        engine_path.remember(p)
        warn = engine_path.release_warning(p)
        QMessageBox.information(
            self, "자리를 기억했습니다",
            f"{p}\n\n다음부터는 묻지 않습니다."
            + (f"\n\n{warn}" if warn else ""))
        # 방금 고른 자리로 다시 풀어 본다 (계산 프로세스가 옛 자리로 떠 있을 수 있으니 정리하고)
        last = getattr(self, "_last_path", None)
        if last:
            ENGINE.shutdown()
            self._start_solve(last)

    def do_export(self):
        ExportDialog(self, self.c, self.mode, self.picked).exec()


def main():
    app = QApplication(sys.argv)
    w = Proto()
    w.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
