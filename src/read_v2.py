"""read_v2.py — 새 서식(v2) 엑셀을 읽어 **엔진이 아는 자리**로 되돌린다.

핵심 생각
    계산 엔진은 지금처럼 **열 위치**로 받는다. 그래서 v2 는 사람이 보는 서식일 뿐이고,
    여기서 머리글을 보고 값을 찾아 `format_v2` 에 적힌 `v1_col` 자리에 놓아 준다.
    ⇒ MATLAB 을 안 고쳐도 되고(재컴파일 0), 화면·엔진 어느 쪽도 서식 변화를 모른다.

읽는 방법이 v1 과 정반대다
    v1 : 머리글을 안 보고 **위치**로 읽는다 (그래서 열이 밀리면 조용히 틀렸다)
    v2 : **머리글로 찾는다** — 열 순서가 바뀌어도, 선택 열이 빠져 있어도 제대로 읽힌다.
         못 찾으면 **어느 열이 없는지 말한다**(PDR 위험 R5 가 여기서 사라진다).

    python src/read_v2.py <v2파일>          # 표 모양을 찍어 본다
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

import numpy as np
from openpyxl import load_workbook

import format_v2 as F

# 내부 키 ← v2 시트 이름 (엔진에 넘기는 차례는 app_worker.TABLE_ORDER 와 같다)
TABLE_OF_SHEET = {
    "Base": "Base_dat",
    "AC Bus Data": "AC_Bus_dat",
    "AC Line Data": "AC_Line_dat",
    "AC Gen Data": "AC_gen_dat",
    "AC 3w Transformer Data": "AC_3wtrans_dat",
    "DC Bus Data": "DC_Bus_dat",
    "DC Line Data": "DC_Line_dat",
    "DC Gen Data": "DC_gen_dat",
    "ACDC IC Data": "IC_dat",
    "MVDC LVDC Converter Data": "DCDC_Conv_dat",
    "AC P Consume Data": "AC_PLoad_dat",
    "AC Q Consume Data": "AC_QLoad_dat",
    "DC P Consume Data": "DC_PLoad_dat",
}


class MissingColumn(ValueError):
    """머리글을 못 찾았다 — 무엇이 없는지 말해 준다."""


def is_v2(path: str | Path) -> bool:
    """v2 파일인가. `읽어보기` 시트와 `Base` 시트로 가린다."""
    wb = load_workbook(path, read_only=True)
    try:
        names = set(wb.sheetnames)
    finally:
        wb.close()
    return "Base" in names or "읽어보기" in names


# ─────────────────────────────────────────────── 머리글 맞추기
def _key(text: str) -> str:
    """머리글을 견주기 좋게 다듬는다 — 단위·공백·대소문자를 무시한다."""
    t = re.sub(r"\[.*?\]", " ", str(text or ""))      # [MW] 같은 단위는 뗀다
    t = re.sub(r"\s+", " ", t).strip().lower()
    return t


def _find(headers: list[str], col: F.Col) -> int | None:
    """그 열이 몇 번째에 있나 (0부터). 없으면 None."""
    want = _key(col.name)
    keys = [_key(h) for h in headers]
    if want in keys:
        return keys.index(want)
    return None


# ─────────────────────────────────────────────── 읽기
def _sheet_rows(ws) -> tuple[list[str], list[list]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    head = [("" if c is None else str(c)) for c in rows[0]]
    body = [list(r) for r in rows[1:]
            if any(v is not None and v != "" for v in r)]
    return head, body


def _num(v) -> float:
    if v is None or v == "" or v == "-":
        return np.nan
    try:
        return float(v)
    except (TypeError, ValueError):
        return np.nan


# 🚨 뜻이 바뀌어 이름을 갈아 낀 열. **옛 이름을 만나면 막는다.**
#    A1 조정 열은 `required=False` 라 못 찾아도 조용히 넘어가는데(값이 NaN 이 되어
#    엔진에겐 없는 것과 같다), 여기서는 그 조용한 무시가 위험하다 — 사용자는
#    `Ctrl Steps = 4` 를 적어 놓고 계단이 걸린 줄 알지만 실제로는 연속으로 돈다.
#    ⭐ *조용히 무시하느니 분명히 막는다*(2026-08-13 AC/DC 경로에서 정한 것과 같은 규칙).
RENAMED = {
    "ctrl steps": ("Ctrl Step Size",
                   "계단 **수**(예: 4) 였는데 이제 **한 단 크기**(예: 0.00625) 입니다"),
    "shunt steps": ("Shunt Step Size",
                    "계단 **수** 였는데 이제 **한 단 크기**(Mvar, 예: 10) 입니다"),
}


def read_sheet(sheet: F.Sheet, ws) -> np.ndarray:
    """v2 시트 하나 → 엔진이 아는 자리에 놓인 배열."""
    head, body = _sheet_rows(ws)

    for h in head:
        old = RENAMED.get(_key(h))
        if old is None:
            continue
        new_name, what = old
        raise MissingColumn(
            f"'{sheet.name}' 시트의 열 이름 '{h}' 는 이제 안 씁니다.\n"
            f"  → '{new_name}' 로 바꿔 주십시오.\n"
            f"  ⚠️ 이름만 바뀐 게 아니라 **뜻이 바뀌었습니다** — {what}.\n"
            f"     적어 둔 값도 다시 보셔야 합니다.")

    if sheet.time_series:
        # 부하 시트: 첫 열이 버스, 그다음이 시각. MW → W 로 되돌린다.
        if not body:
            return np.full((1, max(len(head), 1)), np.nan)
        out = []
        for r in body:
            bus = _num(r[0])
            vals = [_num(x) * F.MW_TO_W for x in r[1:]]
            out.append([bus] + vals)
        return np.asarray(out, dtype=float)

    # 🚨 표의 **폭**을 파일에 실제로 있는 머리글로 정한다.
    #    엔진은 `size(표, 2) >= N` 으로 기능이 있나를 가리므로 폭이 곧 뜻이다 —
    #    한 칸만 넓거나 좁아도 없던 한계가 생기거나 있던 한계가 사라진다.
    #    (`ACDC_71bus_L2_qmax08`: 빈 `Qmin` 열까지 잘라 12열로 만들었더니
    #     엔진이 13열 미만이라 보고 **Qmax 를 통째로 무시**해 답이 1.7% 달라졌다.)
    present = [c.v1_col for c in sheet.cols
               if c.v1_col is not None and _find(head, c) is not None]
    width = max(present) if present else max((c.v1_col or 0) for c in sheet.cols)

    # 2026-08-11: 서식은 **선택 열까지 늘 만들어 둔다**(사람이 값을 채워 넣을 자리가 있어야
    # 하므로). 그래서 "머리글은 있는데 값이 하나도 없는" 끝쪽 선택 열이 생기는데, 그대로
    # 폭에 세면 위 주석대로 **없던 기능이 켜진다.** ⇒ 끝에서부터 훑어 잘라낸다.
    #   · 끝쪽만 자른다 — 중간 열을 빼면 뒤 열의 자리가 밀린다.
    #     (중간이 비어 있는 것은 안전하다. 엔진이 NaN 을 Inf=한계 없음 으로 읽는다:
    #      `preprocess_AC_gen.m` 의 `qmax_gen(~isfinite(qmax_gen)) = Inf`)
    #   · 필수 열과 기본값이 있는 열은 안 자른다(그 열은 값이 없어도 뜻이 있다).
    #   · 🚨 **아무 폭으로나 줄이면 안 된다.** 짝을 이루는 열(Qmax·Qmin / Pmax·Pmin)은 뒤쪽만
    #     비었다고 잘라 내면 앞쪽까지 죽는다 — `ACDC_71bus_L2_qmax08` 이 그렇다(Qmax 는 있고
    #     Qmin 이 비었는데, 13→12 로 줄이면 엔진이 13열 미만이라 보고 **Qmax 를 통째로 무시**한다).
    #     ⇒ `sheet.v1_widths` = 엔진이 아는 폭 목록. **거기 있는 값으로만** 줄인다.
    if body and sheet.v1_widths:
        by_v1 = {c.v1_col: c for c in sheet.cols if c.v1_col is not None}
        allowed = set(sheet.v1_widths)
        w, best = width, width
        while w > 0:
            col = by_v1.get(w)
            if col is None or col.required or col.default is not None:
                break
            at = _find(head, col)
            if at is None:
                break                       # 머리글이 없으면 애초에 폭에 안 들어간다
            has_value = any(
                at < len(r) and not np.isnan(_num(r[at])) for r in body)
            if has_value:
                break
            w -= 1
            if w in allowed:                # 엔진이 아는 폭에 닿을 때만 확정한다
                best = w
        width = best

    if not body:
        # 🚨 값이 없는 시트를 빈 배열로 주면 안 된다. 컴파일된 전처리기 일부가
        #    `any(isnan(...))` 로 "데이터 없음"을 가리므로 **1줄짜리 NaN 행**이어야 한다
        #    (`load_case._read_numeric_sheet` 가 같은 이유로 그렇게 만든다).
        return np.full((1, width), np.nan)
    out = np.full((len(body), width), np.nan)
    missing = []
    for col in sheet.cols:
        if col.v1_col is None:
            continue                                  # v2 에서만 있는 열(예: DC/DC Status)
        if col.v1_col > width:
            continue                                  # 위에서 폭에서 뺀 빈 선택 열
        at = _find(head, col)
        if at is None:
            if col.required and col.default is None:
                missing.append(col.header)
            elif col.default is not None and col.v1_col <= width:
                # 🚨 기본값 때문에 표를 **넓히지 않는다.** 폭이 곧 뜻이라(위 주석),
                #    없던 열을 채워 넣으면 v1 에서 꺼져 있던 기능이 켜진다.
                #    `case14_matpower` 의 AC 발전기 표는 10열인데 `|V| deadband` 는
                #    11번째라, 넓히려다 표 밖으로 나가 터졌다(2026-08-06).
                out[:, col.v1_col - 1] = col.default
            continue
        vals = np.array([_num(r[at]) if at < len(r) else np.nan for r in body])
        if col.scale != F.KEEP:
            # v2 → 엔진 단위로 되돌린다.
            # ⚠️ `/ 1e-6` 이 아니라 `* 1e6` 으로 한다 — 1e-6 은 이진수로 딱 안 떨어져
            #    나누면 오차가 더 붙는다(1e6 은 딱 떨어진다).
            vals = vals * (1.0 / col.scale)
        out[:, col.v1_col - 1] = vals

    if missing:
        raise MissingColumn(
            f"'{sheet.name}' 시트에서 못 찾은 열: {', '.join(missing)}\n"
            f"  있는 머리글: {', '.join(h for h in head if h)}")

    return out


def read_tables(path: str | Path) -> dict[str, np.ndarray]:
    """v2 파일 → 엔진에 넘길 표 묶음 (`load_case(...).tables` 와 같은 모양)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        tables: dict[str, np.ndarray] = {}
        for sheet in F.SHEETS:
            key = TABLE_OF_SHEET.get(sheet.name)
            if key is None:
                continue                              # Mode 는 따로 읽는다
            if sheet.name not in wb.sheetnames:
                tables[key] = np.zeros((0, 0))
                continue
            tables[key] = read_sheet(sheet, wb[sheet.name])
        return tables
    finally:
        wb.close()


def read_mode(path: str | Path) -> float:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["Mode"]
        for r in ws.iter_rows(min_row=2, values_only=True):
            v = _num(r[0])
            if not np.isnan(v):
                return float(v)
    finally:
        wb.close()
    return 0.0


if __name__ == "__main__":
    p = Path(sys.argv[1])
    print(f"{p.name}  (v2 = {is_v2(p)})  Mode = {read_mode(p)}")
    for k, v in read_tables(p).items():
        print(f"  {k:<16} {v.shape}")
