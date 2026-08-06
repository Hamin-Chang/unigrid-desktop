"""convert_case.py — 케이스 엑셀을 옛 서식(v1)에서 새 서식(v2)으로 다시 저장한다.

  python src/convert_case.py <파일 또는 폴더> [-o 내보낼폴더] [--check]

무엇을 하나
    `format_v2.py` 의 정의대로 머리글을 바꾸고, 단위를 환산하고, 죽은 열을 버리고,
    새로 생긴 열(예: DC/DC 의 `Status`)을 채워 **새 파일로 저장한다.**

🚨 원본은 절대 건드리지 않는다. 언제나 새 파일(`<이름>_v2.xlsx`)로 쓴다.

🚨 v1 은 **위치로만** 뜻이 정해진다(머리글을 아무도 안 읽었다). 그래서 여기서도 머리글을
   믿지 않고 위치로 읽되, **열 수가 아는 모양인지 먼저 확인**한다.
   세대가 다른 파일(발전기·부하를 버스 시트에 담던 22열짜리 `AC Bus Data`,
   MatACDC 에서 급히 만든 21열 IC 등)을 같은 자리로 읽으면 **조용히 엉뚱한 계통**이 된다.
   ⇒ 모르는 모양이면 **바꾸지 않고 그 사실을 말한다.**

`--check` 는 저장하지 않고 무엇이 걸리는지만 훑는다.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

import format_v2 as F


class Refused(Exception):
    """모르는 모양이라 바꾸지 않았다."""


# ─────────────────────────────────────────────── 읽기 (v1 = 위치)
def _rows(ws) -> list[list]:
    """머리글 줄을 뺀 숫자 줄들. 빈 줄은 버린다."""
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or v == "" for v in r):
            continue
        out.append(list(r))
    return out


def _width(ws) -> int:
    """표의 폭. **머리글이 비어 있어도 값이 있으면 그 열까지가 폭이다.**

    🚨 예전에는 머리글이 빈 끝 열을 그냥 잘랐다. 그런데 `ACDC_71bus_L2_ic15.xlsx` 는
       IC 표의 21·22열(`V_base [kV]`·`I_max [kA]` = 전류 한계)에 **머리글을 안 적고 값만**
       넣어 뒀다. 잘라 버리자 22열이 20열이 되어 **엔진이 전류 한계를 통째로 무시**했고,
       손실이 315,403 → 389,129 (23%) 어긋났다. 폭이 곧 뜻이므로(아래 주석) 값을 본다.
       (2026-08-06 v14 56개 전수 변환에서 발견)
    """
    hdr = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ()))
    n_head = len(hdr)
    while n_head and (hdr[n_head - 1] is None or hdr[n_head - 1] == ""):
        n_head -= 1

    n_data = 0                                   # 값이 든 마지막 열
    for r in ws.iter_rows(min_row=2, values_only=True):
        for j in range(len(r) - 1, n_data - 1, -1):
            if r[j] is not None and r[j] != "":
                n_data = j + 1
                break
    return max(n_head, n_data)


def _num(v):
    if v is None or v == "" or v == "-":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


# ─────────────────────────────────────────────── 한 시트 바꾸기
def convert_sheet(sheet: F.Sheet, ws) -> tuple[list[str], list[list]]:
    """(v2 머리글, v2 줄들) 을 돌려준다."""
    width = _width(ws)
    src = _rows(ws)

    if sheet.time_series:
        # 부하 시트: 첫 열이 버스, 그다음은 시각. 값 전부 W → MW
        n_time = max((len(r) for r in src), default=1) - 1
        head = [sheet.cols[0].header] + [str(i) for i in range(1, n_time + 1)]
        rows = []
        for r in src:
            bus = _num(r[0])
            vals = [(_num(x) or 0.0) * F.W_TO_MW for x in r[1:1 + n_time]]
            vals += [0.0] * (n_time - len(vals))
            rows.append([bus] + vals)
        return head, rows

    if not src:
        # 값이 없는 시트(머리글만)는 모양을 몰라도 잘못 읽을 것이 없다.
        # 계통에 그 설비가 없을 때 흔하다 — 여기서 막으면 파일 전체가 막힌다.
        keep = [c for c in sheet.cols if c.v1_col is None or c.v1_col <= max(width, 1)]
        return [c.header for c in (keep or sheet.cols)], []

    if sheet.v1_widths and width not in sheet.v1_widths:
        raise Refused(
            f"'{sheet.source_name}' 의 열 수가 {width} 인데 아는 모양은 "
            f"{list(sheet.v1_widths)} 입니다. 서식 세대가 다를 수 있어 바꾸지 않았습니다.")

    # 🚨 **옛 파일에 없던 선택 열은 만들지 않는다.**
    #    엔진은 `size(표, 2) >= N` 으로 기능이 있나를 가린다 — 빈 칸이라도 자리를 만들면
    #    없던 기능이 생긴 것으로 읽힌다(`AConly_case118` 에서 발전기 한계 표가 없다가 생겼다).
    #    반대로 자리를 없애면 있던 기능이 사라진다 — 엔진의 가름이 **짝으로** 걸려 있어서다
    #    (13 이상이면 Qmax·Qmin 둘 다 · 15 이상이면 Pmax·Pmin · 16 이상이면 S_N).
    cols = [c for c in sheet.cols if c.v1_col is None or c.v1_col <= width]

    head = [c.header for c in cols]
    rows = []
    for r in src:
        line = []
        for c in cols:
            if c.v1_col is None or c.v1_col > len(r):
                line.append(c.default)             # 신설이거나 옛 파일에 없던 열
                continue
            v = _num(r[c.v1_col - 1])
            if v is None:
                line.append(c.default)
            else:
                line.append(v * c.scale if c.scale != F.KEEP else v)
        rows.append(line)
    return head, rows


# ─────────────────────────────────────────────── 파일 하나
HEAD_FILL = PatternFill("solid", fgColor="E8EEF7")
NOTE_FILL = PatternFill("solid", fgColor="FFF6E5")


def _write_notes(wb: Workbook) -> None:
    ws = wb.create_sheet("읽어보기", 0)
    ws["A1"] = "UNIGRID 케이스 서식 v2 — 값이 무슨 뜻인지"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = "이 시트는 계산에 쓰이지 않습니다. 마음껏 적어 두세요."
    ws["A2"].font = Font(italic=True, color="6B7684")
    for i, name in enumerate(["시트", "열", "뜻"], start=1):
        cell = ws.cell(row=4, column=i, value=name)
        cell.font = Font(bold=True)
        cell.fill = HEAD_FILL
    for j, (sheet, col, text) in enumerate(F.VALUE_NOTES, start=5):
        ws.cell(row=j, column=1, value=sheet)
        ws.cell(row=j, column=2, value=col)
        ws.cell(row=j, column=3, value=text)
    for col, w in (("A", 26), ("B", 24), ("C", 72)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"


def convert_file(path: Path, out_dir: Path | None = None,
                 check_only: bool = False) -> tuple[Path | None, list[str]]:
    """한 파일을 바꾼다. (저장한 자리, 알림 목록) 을 돌려준다."""
    src = load_workbook(path, read_only=True, data_only=True)
    notes: list[str] = []
    wb = Workbook()
    wb.remove(wb.active)

    for sheet in F.SHEETS:
        name = sheet.source_name
        if name not in src.sheetnames:
            if not sheet.optional_sheet and not sheet.time_series:
                notes.append(f"⚠️ 시트 없음: {name}")
            continue
        try:
            head, rows = convert_sheet(sheet, src[name])
        except Refused as exc:
            notes.append(f"⛔ {exc}")
            continue

        ws = wb.create_sheet(sheet.name)
        ws.append(head)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = HEAD_FILL
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        for row in rows:
            ws.append(row)
        for i, c in enumerate(head, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = \
                max(10, min(24, len(str(c)) + 3))
        ws.freeze_panes = "A2"

    src.close()

    if any(n.startswith("⛔") for n in notes):
        return None, notes                       # 하나라도 거부되면 파일을 안 만든다

    _write_notes(wb)
    if check_only:
        return None, notes

    out_dir = out_dir or path.parent
    out_dir.mkdir(parents=True, exist_ok=True)
    out = out_dir / f"{path.stem}_v2.xlsx"
    wb.save(out)
    return out, notes


# ─────────────────────────────────────────────── 실행
def main() -> int:
    ap = argparse.ArgumentParser(description="케이스 엑셀을 v1 → v2 서식으로 다시 저장")
    ap.add_argument("target", help="엑셀 파일 또는 폴더")
    ap.add_argument("-o", "--out", help="내보낼 폴더 (기본: 원본 옆)")
    ap.add_argument("--check", action="store_true", help="저장하지 않고 훑어만 본다")
    args = ap.parse_args()

    t = Path(args.target)
    files = sorted(t.glob("*.xlsx")) if t.is_dir() else [t]
    files = [f for f in files if not f.stem.endswith("_v2")]
    if not files:
        print(f"엑셀을 찾지 못했습니다: {t}")
        return 2

    out_dir = Path(args.out) if args.out else None
    done = refused = 0
    for f in files:
        try:
            out, notes = convert_file(f, out_dir, args.check)
        except Exception as exc:
            print(f"  {f.name:<40} 실패 — {type(exc).__name__}: {exc}")
            refused += 1
            continue
        if out is None and any(n.startswith("⛔") for n in notes):
            print(f"  {f.name:<40} 바꾸지 않음")
            refused += 1
        else:
            print(f"  {f.name:<40} " + ("훑어봄" if args.check else f"→ {out.name}"))
            done += 1
        for n in notes:
            print(f"      {n}")

    print()
    print(f"바꾼 파일 {done} · 바꾸지 않은 파일 {refused}")
    return 1 if refused else 0


if __name__ == "__main__":
    sys.exit(main())
