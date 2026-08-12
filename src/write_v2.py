"""write_v2.py — 메모리에 든 케이스를 **새 서식(v2) 엑셀**로 저장한다.

무엇을 하나
    `unigrid_convert` 가 PSS/E `.raw`·MATPOWER `.m` 을 읽어 만든 `ACDCCase`(숫자 표 13종)를
    받아 사람이 보는 v2 엑셀로 쓴다. 변환 창(X1 (b))이 부르는 곳이다.

방향이 `read_v2.py` 와 정반대다
    read_v2 : v2 값 / scale  →  엔진 자리(`v1_col`)
    write_v2: 엔진 값 * scale →  v2 머리글

    ⇒ 두 방향 모두 `format_v2.py` 한 곳만 본다. 여기에 열을 따로 적지 않는다.

**서식의 전체 열을 쓴다** (2026-08-11 변경). 값이 없는 열도 머리글을 만들고 칸을 비워 둔다 —
   사람이 나중에 발전기 한계 같은 값을 채워 넣으려면 **자리가 있어야** 하기 때문이다.

   🚨 그전에는 반대로 했다(케이스 표의 실제 열 수까지만). 엔진이 `size(표, 2) >= N` 으로
   기능이 있나를 가려 **폭이 곧 뜻**이라, 빈 칸이라도 자리를 더 만들면 없던 기능이 켜졌다.
   ⇒ 이제 그 몫은 `read_v2` 가 맡는다 — **끝쪽의 "머리글만 있고 값이 없는 선택 열" 을 폭에서 뺀다.**
   그래서 값을 안 채우면 엔진에 옛 폭 그대로 넘어가고, 채우는 순간 기능이 켜진다.
"""

from __future__ import annotations

from pathlib import Path

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

import format_v2 as F
from convert_case import HEAD_FILL, _write_notes
from read_v2 import TABLE_OF_SHEET


def _cell(v):
    """숫자 하나를 엑셀 칸으로. NaN·무한대는 빈 칸으로 둔다."""
    if v is None:
        return None
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    if not np.isfinite(f):
        return None
    return f


def sheet_rows(sheet: F.Sheet, tab: np.ndarray) -> tuple[list[str], list[list]]:
    """표 하나 → (v2 머리글, v2 줄들)."""
    arr = np.asarray(tab, dtype=float) if tab is not None else np.zeros((0, 0))
    if arr.ndim != 2:
        arr = arr.reshape(1, -1) if arr.size else np.zeros((0, 0))

    if sheet.time_series:
        # 부하 시트: 첫 열이 버스, 그다음이 시각. W → MW.
        if not arr.size:
            return [sheet.cols[0].header, "1"], []
        n_time = arr.shape[1] - 1
        head = [sheet.cols[0].header] + [str(i) for i in range(1, n_time + 1)]
        rows = [[_cell(r[0])] + [_cell(x * F.W_TO_MW) for x in r[1:]] for r in arr]
        return head, rows

    # 값이 하나도 없으면(설비가 없는 계통) 머리글만 남긴다 — 되읽으면 NaN 한 줄이 되어
    # 엔진이 "그 설비 없음"으로 본다.
    if not arr.size or np.all(np.isnan(arr)):
        return [c.header for c in sheet.cols], []

    width = arr.shape[1]
    cols = list(sheet.cols)          # 2026-08-11: 서식의 **전체 열**을 쓴다 (아래 참고)

    head = [c.header for c in cols]
    rows = []
    for r in arr:
        line = []
        for c in cols:
            if c.v1_col is None:                 # v2 에서 새로 생긴 열
                line.append(_cell(c.default))
                continue
            if c.v1_col > width:
                # 케이스 표에 없던 열. 기본값이 있으면 그 값을, 없으면 **빈칸**으로 둔다.
                #   · 기본값 있는 열(Status=1 · 데드밴드=0)은 채워야 뜻이 선다. 채운 값이
                #     "그 기능 꺼짐" 과 같아서 답이 안 바뀐다.
                #   · 기본값 없는 선택 열(Qmax·Qmin·Pmax·Pmin·S_N)은 비운다. `read_v2` 가
                #     끝쪽의 빈 선택 열을 폭에서 빼므로 엔진에는 예전 폭 그대로 넘어간다.
                line.append(_cell(c.default))
                continue
            v = r[c.v1_col - 1]
            line.append(_cell(v if c.scale == F.KEEP else v * c.scale))
        rows.append(line)
    return head, rows


def write_case(case, out_path: str | Path) -> Path:
    """`ACDCCase` 를 v2 엑셀로 저장하고 저장한 자리를 돌려준다."""
    out = Path(out_path)
    if out.suffix.lower() != ".xlsx":
        out = out.with_suffix(".xlsx")

    wb = Workbook()
    wb.remove(wb.active)

    for sheet in F.SHEETS:
        if sheet.name == "Mode":
            head, rows = [sheet.cols[0].header], [[_cell(case.mode)]]
        else:
            key = TABLE_OF_SHEET.get(sheet.name)
            if key is None:
                continue
            tab = case.tables.get(key)
            head, rows = sheet_rows(sheet, None if tab is None else np.asarray(tab))
            # 🚨 값이 없는 시트도 **머리글만 남겨 만든다.** `.m`·`.raw` 는 AC 전용이라
            #    DC 시트가 늘 비는데, 여기서 빼 버리면 사람이 AC/DC 케이스를 만들려고
            #    시트 이름과 머리글을 손으로 지어내야 한다(변환 창이 그러라고 안내한다).
            #    되읽으면 NaN 한 줄이 되어 엔진은 "그 설비 없음"으로 본다.

        ws = wb.create_sheet(sheet.name)
        ws.append(head)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = HEAD_FILL
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        for row in rows:
            ws.append(row)
        for i, c in enumerate(head, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = \
                max(10, min(24, len(str(c)) + 3))
        ws.freeze_panes = "A2"

    _write_notes(wb)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out
