"""exporter.py — 결과를 엑셀·그림 파일로 저장한다.

원본 MATLAB 앱(`acdcapp_0404.mlapp` 의 `ExportButtonPushed`, 472줄)이 만들던 것과
**파일 이름·시트 구성을 똑같이** 맞춘다 — 그 결과를 받아 쓰던 기존 스크립트가
손 안 대고 그대로 돌아가게 하려는 것이다. 원본과 다른 점은 셋뿐이다:

  ① **무엇을 저장할지 고를 수 있다** (원본은 늘 전부 저장). 24시간 × 파일 4개면
     시간이 꽤 걸리는데 표만 필요할 때 그림까지 기다릴 이유가 없다.
  ② **그림을 PNG 와 PDF 로 둘 다** 낸다 (원본은 PNG 만). 여기 그림은 Qt 라
     벡터로 나가므로, 논문에 넣을 땐 PDF 가 확대해도 안 깨진다.
  ③ **저장 폴더 기본값이 케이스 이름을 딴 폴더**다 (원본은 매번 폴더를 물어봄).
     파일 이름이 케이스와 무관하게 늘 같아서, 케이스를 바꿔가며 돌리면
     한 폴더에 저장할 경우 **앞 결과를 덮어쓴다**.

원본 비교 모드 파일명의 오타 `Comaprison` 은 `Comparison` 으로 고쳤다.
"""
from __future__ import annotations

from pathlib import Path

import numpy as np
from PySide6.QtCore import Qt, QMarginsF, QPoint, QSizeF
from PySide6.QtGui import QPageSize, QPainter, QPdfWriter
from PySide6.QtWidgets import (QApplication, QHBoxLayout, QScrollArea,
                               QVBoxLayout, QWidget)

import charts

# 표 이름 → (Solution 에서 꺼낼 이름, 원본 앱이 쓰던 파일 이름)
TABLE_FILES = {
    "AC 결과":   ("AC", "AC_results.xlsx"),
    "DC 결과":   ("DC", "DC_results.xlsx"),
    "선로 조류": ("Branch", "Branch_results.xlsx"),
    "손실":      ("Loss", "Loss_results.xlsx"),
}

# 계통 종류(케이스 엑셀 `Mode` 시트)별로 나오는 표.
# 원본 `ExportButtonPushed` 의 switch 문과 같다 — 0 혼합 / 1 AC only / 2 DC only.
MODE_TABLES = {
    0: ["AC 결과", "DC 결과", "선로 조류", "손실"],
    1: ["AC 결과", "선로 조류", "손실"],
    2: ["DC 결과", "선로 조류", "손실"],
}

FIG_W, FIG_H = 1100, 620            # 그림 한 장 크기 [px]


def default_folder(case_name) -> Path:
    """기본 저장 위치 — `~/Desktop/결과/<케이스 이름>/`."""
    stem = Path(str(case_name)).stem or "결과"
    return Path.home() / "Desktop" / "결과" / stem


def table_names(sol) -> list[str]:
    """이 케이스에서 실제로 저장할 수 있는 표 이름들.

    계통 종류로 한 번 거르고, **값이 비어 있는 표는 또 거른다**
    (혼합 계통이라고 적혀 있어도 DC 결과가 없을 수 있다).
    """
    names = MODE_TABLES.get(int(getattr(sol, "mode", 0)), MODE_TABLES[0])
    keep = []
    for n in names:
        which, _ = TABLE_FILES[n]
        arr = sol.loss if which == "Loss" else getattr(sol, which, None)
        if arr is not None and getattr(arr, "size", 0):
            keep.append(n)
    return keep


def figure_names(sol, mode: str) -> list[tuple[str, str]]:
    """(화면 탭 이름, 저장할 파일 이름) — 원본 앱이 쓰던 이름 그대로.

    DC only 계통은 위상각도 주파수도 없어서 원본도 이름을 달리 썼다(563·567줄).
    "조류 (P·Q)" 만 원본에 없던 그림이라 이름을 새로 붙였다.
    """
    dc_only = int(getattr(sol, "mode", 0)) == 2
    if mode == "스냅샷":
        volt = "Voltage Result" if dc_only else "Voltage - Phase Angle Result"
        return [("전압·위상", volt),
                ("조류 (P·Q)", "Power Flow P-Q"),
                ("부하율", "Line Loading Bar"),
                ("토폴로지", "Topology")]
    if mode == "다이나믹":
        volt = "Voltage Dynamics" if dc_only else "Voltage - Phase Angle Dynamics"
        return [("전압·위상", volt),
                ("주파수", "Frequency Dynamic"),
                ("토폴로지", "Topology")]
    return []


# ─────────────────────────────────────────── 엑셀
def _sheet(wb, title, cols, arr):
    ws = wb.create_sheet(title)
    ws.append([str(c) for c in cols])          # 1행 = 열 이름 (writetable 과 같음)
    a = np.atleast_2d(np.asarray(arr, dtype=float))
    for r in a:
        # NaN·inf 는 빈 칸으로 — 엑셀에서 #NUM! 로 뜨는 것보다 낫다
        ws.append([float(v) if np.isfinite(v) else None for v in r])


def save_tables(sol, folder: Path, names, on_step=None) -> list[Path]:
    """표를 엑셀로. **한 시간에 시트 하나**, 이름은 `1H`·`2H`… (원본과 동일).

    손실만 시간축이 이미 행이라 `Loss` 시트 한 장으로 끝난다.
    """
    from openpyxl import Workbook
    folder.mkdir(parents=True, exist_ok=True)
    out = []
    for name in names:
        which, fname = TABLE_FILES[name]
        wb = Workbook()
        wb.remove(wb.active)                    # 기본으로 생기는 빈 시트 제거
        if which == "Loss":
            _sheet(wb, "Loss", sol.cols("Loss"), sol.loss)
        else:
            for t in range(max(1, int(sol.n_time))):
                arr = sol.at(which, t)
                if arr.size:
                    _sheet(wb, f"{t + 1}H", sol.cols(which), arr)
        if not wb.sheetnames:
            continue
        p = folder / fname
        wb.save(p)
        out.append(p)
        if on_step:
            on_step(fname)
    return out


# ─────────────────────────────────────────── 그림
def _realize(w):
    """화면에 안 띄우고 배치만 확정시킨다.

    이걸 건너뛰면 위젯 크기가 잡히기 전에 그려서 **빈 그림**이 나온다.
    `WA_DontShowOnScreen` 이라 사용자 화면에는 아무것도 안 뜬다.
    """
    w.setAttribute(Qt.WA_DontShowOnScreen, True)
    w.show()
    QApplication.processEvents()
    return w


def _page(c, sol, plots, layout, t, bus_row):
    """탭 한 장을 화면과 같은 짜임으로 다시 만든다 (그림 여러 개면 같이 담는다)."""
    page = QWidget()
    page.setStyleSheet(f"background:{c['plot']};")
    lay = QHBoxLayout(page) if layout == "h" else QVBoxLayout(page)
    lay.setContentsMargins(12, 12, 12, 12)
    lay.setSpacing(10)
    made = 0
    for pname in plots:
        wdg = charts.build(pname, c, sol, t, bus_row)
        if wdg is not None:
            lay.addWidget(wdg)
            made += 1
    if not made:
        page.deleteLater()
        return None
    # 계통도는 스크롤 상자에 담겨 있어 그대로 잡으면 **보이는 만큼만** 나온다.
    # 안쪽 위젯을 꺼내 제 크기(minimumWidth)로 키워서 통째로 뽑는다.
    box = page.findChild(QScrollArea)
    if box is not None and box.widget() is not None:
        inner = box.widget()
        inner.setParent(None)
        page.deleteLater()
        inner.resize(max(inner.minimumWidth(), FIG_W),
                     max(inner.minimumHeight(), 420))
        return _realize(inner)
    page.resize(FIG_W, FIG_H)
    return _realize(page)


def _save_png(w, path: Path):
    w.grab().save(str(path))


def _save_pdf(w, path: Path):
    """벡터 PDF — 확대해도 안 깨진다(논문·발표용).

    화면 좌표는 96 dpi 기준이라 그 비율로 종이 크기를 잡고,
    그린 다음 해상도만큼 키운다. 여백은 0 — 그림만 딱 들어가게.
    """
    wr = QPdfWriter(str(path))
    wr.setResolution(300)
    mm = 25.4 / 96.0
    wr.setPageSize(QPageSize(QSizeF(w.width() * mm, w.height() * mm),
                             QPageSize.Millimeter))
    wr.setPageMargins(QMarginsF(0, 0, 0, 0))
    p = QPainter(wr)
    p.scale(wr.resolution() / 96.0, wr.resolution() / 96.0)
    w.render(p, QPoint(0, 0))       # QPainter 로 그릴 땐 시작점을 꼭 줘야 한다
    p.end()


# 비교 그림 이름 — 원본 앱과 같다(오타 `Comaprison` 만 고쳤다).
CMP_FILES = {
    "전압 크기": "Voltage Comparison",
    "위상각":   "Phase Angle Comparison",
    "주파수":   "Frequency Comparison",
    "손실":     "Power Loss Comparison",
}


def compare_figure_name(item: str, axis: str) -> str:
    """비교 그림 파일 이름. 주파수·손실은 시간끼리에서만 그려서 늘 [Time]."""
    tag = "Time" if (item in ("주파수", "손실") or axis != "버스끼리") else "Bus"
    return f"{CMP_FILES.get(item, item)} [{tag}]"


def save_compare_figures(c, sol, axis, targets, items, folder: Path,
                         on_step=None) -> list[Path]:
    """비교 그래프를 PNG·PDF 로.

    일반 내보내기와 **따로** 둔다 — 원본 앱도 비교는 별도 버튼이었고
    (`ExportComparisonButtonPushed`), 보고 있는 그 그림을 그 자리에서
    저장하는 쪽이 자연스럽다.
    못 그리는 항목(예: DC 버스만 골라 놓고 위상각)은 **건너뛴다** —
    안내 문구를 그림 파일로 저장하면 안 되니까.
    """
    folder.mkdir(parents=True, exist_ok=True)
    out = []
    for item in items:
        w = charts.compare_chart(c, sol, item, axis, targets)
        if not charts.is_chart(w):
            if on_step:
                on_step(f"{item} (그릴 수 없어 건너뜀)")
            continue
        holder = QWidget()
        holder.setStyleSheet(f"background:{c['plot']};")
        lay = QVBoxLayout(holder)
        lay.setContentsMargins(12, 12, 12, 12)
        lay.addWidget(w)
        holder.resize(FIG_W, FIG_H)
        _realize(holder)
        base = compare_figure_name(item, axis)
        png, pdf = folder / f"{base}.png", folder / f"{base}.pdf"
        _save_png(holder, png)
        _save_pdf(holder, pdf)
        holder.close()
        holder.deleteLater()
        out += [png, pdf]
        if on_step:
            on_step(base)
    return out


def save_figures(c, sol, mode, folder: Path, wanted, t=0, bus_row=0,
                 on_step=None) -> list[Path]:
    """고른 탭을 PNG·PDF 로 각각 한 장씩."""
    from prototype import GRAPHS       # 늦게 부른다 — 서로 부르는 꼴을 피하려고
    spec = {n: (pl, la) for n, pl, la in GRAPHS.get(mode, [])}
    folder.mkdir(parents=True, exist_ok=True)
    out = []
    for tab, fname in figure_names(sol, mode):
        if tab not in wanted or tab not in spec:
            continue
        plots, layout = spec[tab]
        w = _page(c, sol, plots, layout, t, bus_row)
        if w is None:
            continue
        png, pdf = folder / f"{fname}.png", folder / f"{fname}.pdf"
        _save_png(w, png)
        _save_pdf(w, pdf)
        w.close()
        w.deleteLater()
        out += [png, pdf]
        if on_step:
            on_step(fname)
    return out
